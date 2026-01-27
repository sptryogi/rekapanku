import streamlit as st
st.set_page_config(
    page_title="Rekapanku",           # judul di tab browser
    page_icon="📊",                   # emoji atau file ikon (.png/.ico)
    layout="wide"
)
from datetime import datetime
import pandas as pd
import numpy as np
import io
import time
import re
from rapidfuzz import fuzz
import pdfplumber
from openpyxl import load_workbook

# --- FUNGSI-FUNGSI PEMROSESAN ---

def clean_and_convert_to_numeric(column):
    """Menghapus semua karakter non-digit (kecuali titik dan minus) dan mengubah kolom menjadi numerik."""
    if column.dtype == 'object':
        column = column.astype(str).str.replace(r'[^\d,\-]', '', regex=True)
        column = column.str.replace(',', '.', regex=False)
    return pd.to_numeric(column, errors='coerce').fillna(0)

def clean_order_all_numeric(column):
    """
    Fungsi khusus untuk membersihkan kolom di file order-all.
    Menghapus semua karakter non-digit dari string.
    """
    # Karena kita akan memastikan kolom dibaca sebagai string,
    # kita bisa langsung membersihkannya dengan aman.
    # Regex `\D` berarti "karakter apa pun yang bukan digit".
    # Ini akan menghapus '.' , ',' , spasi, 'Rp', dll.
    cleaned_column = column.astype(str).str.replace(r'\D', '', regex=True)
    
    # Ubah string angka yang sudah bersih (misal: "35750") ke tipe data numerik.
    return pd.to_numeric(cleaned_column, errors='coerce').fillna(0)

def clean_columns(df):
    """Menghapus spasi di awal dan akhir dari semua nama kolom DataFrame."""
    df.columns = df.columns.str.strip()
    return df

def extract_relevant_variation_part(var_str):
    """Mengekstrak bagian variasi yang relevan (A5, QPP, dll.) untuk DAMA.ID STORE."""
    if pd.isna(var_str):
        return None
    
    var_str_clean = str(var_str).strip().upper()
    parts = [p.strip() for p in var_str_clean.split(',')]
    # Gunakan keywords yang sama dengan logika di process_rekap
    size_keywords = {'QPP', 'A5', 'B5', 'A6', 'A7', 'HVS', 'KORAN'}
    
    for part in parts:
        if part in size_keywords:
            return part # Kembalikan bagian relevan pertama yang ditemukan
    
    return None # Kembalikan None (atau string kosong) jika tidak ada yang cocok

def extract_paper_and_size_variation(var_str):
    """
    Mengekstrak Jenis Kertas (HVS, QPP, KK, KORAN, dll.) ATAU 
    Ukuran/Paket (A5, B5, PAKET 10, dll.) dari string variasi.
    Mengembalikan bagian relevan yang ditemukan, dipisahkan spasi.
    """
    if pd.isna(var_str):
        return '' # Kembalikan string kosong jika input NaN

    var_str_clean = str(var_str).strip().upper()
    
    # Definisikan keywords dan patterns yang dicari
    # Anda bisa tambahkan jenis kertas atau pola lain di sini
    paper_types = {'HVS', 'QPP', 'KORAN', 'KK', 'KWARTO', 'BIGBOS', 'ART PAPER'} 
    size_package_patterns = [
        r'\b(PAKET\s*\d+)\b',      # Contoh: PAKET 10, PAKET 5
        r'\b((A|B)\d{1,2})\b'     # Contoh: A5, B5, A6, A7 (hanya kode ukuran)
    ]
    
    relevant_parts_found = []
    
    # 1. Cari Jenis Kertas (sebagai kata utuh)
    # Gunakan regex \b untuk memastikan kata utuh
    for paper in paper_types:
        if re.search(r'\b' + re.escape(paper) + r'\b', var_str_clean):
            # Map KK ke KORAN jika ditemukan
            relevant_parts_found.append('KORAN' if paper == 'KK' else paper)
            
    # 2. Cari Ukuran/Paket menggunakan pola regex
    for pattern in size_package_patterns:
        matches = re.findall(pattern, var_str_clean)
        # findall bisa mengembalikan tuple jika ada group, ambil group utama
        for match in matches:
             if isinstance(match, tuple):
                 # Ambil group pertama yang cocok (misal 'PAKET 10' atau 'A5')
                 relevant_parts_found.append(match[0].strip()) 
             else:
                 relevant_parts_found.append(match.strip())

    # Hilangkan duplikat (jika ada) dan gabungkan dengan spasi
    # Urutkan agar konsisten (misal selalu 'A5 HVS', bukan kadang 'HVS A5')
    unique_parts = sorted(list(set(relevant_parts_found)))
    
    return ' '.join(unique_parts) # Gabungkan bagian yang relevan
    
def process_rekap(order_df, income_df, seller_conv_df):
    """
    Fungsi untuk memproses dan membuat sheet 'REKAP' dengan file 'income' sebagai data utama.
    """
    # --- PERUBAIKAN 1: Mengubah agregasi untuk memisahkan produk per pesanan ---
    # Agregasi data dari order-all berdasarkan No. Pesanan DAN Nama Produk
    order_agg = order_df.groupby(['No. Pesanan', 'Nama Produk','Nama Variasi']).agg({
        'Jumlah': 'sum',
        'Harga Setelah Diskon': 'first',
        'Total Harga Produk': 'sum'
        #'Nama Variasi': 'first'
    }).reset_index()
    order_agg.rename(columns={'Jumlah': 'Jumlah Terjual'}, inplace=True)

    # Pastikan tipe data 'No. Pesanan' sama untuk merge
    income_df['No. Pesanan'] = income_df['No. Pesanan'].astype(str)
    order_agg['No. Pesanan'] = order_agg['No. Pesanan'].astype(str)
    seller_conv_df['Kode Pesanan'] = seller_conv_df['Kode Pesanan'].astype(str)
    
    # Gabungkan income_df dengan order_agg. Ini akan membuat duplikasi baris income untuk setiap produk.
    rekap_df = pd.merge(income_df, order_agg, on='No. Pesanan', how='left')

    # # 1. Pastikan 'Total Penghasilan' (dari income_df) adalah numerik
    # rekap_df['Total Penghasilan'] = clean_and_convert_to_numeric(rekap_df['Total Penghasilan'])
    
    # # 2. Tandai baris retur BARU: Dapatkan daftar No. Pesanan yang diretur
    # #    Ini adalah No. Pesanan di mana SETIDAKNYA SATU baris memiliki Total Penghasilan < 0
    # returned_orders_list = rekap_df[rekap_df['Total Penghasilan'] < 0]['No. Pesanan'].unique()
    
    # 1. Pastikan 'No. Pengajuan' ada dan bersih (di rekap_df, dari income_dilepas)
    if 'No. Pengajuan' not in rekap_df.columns:
        rekap_df['No. Pengajuan'] = np.nan # Buat kolomnya jika tidak ada
    rekap_df['No. Pengajuan'] = rekap_df['No. Pengajuan'].astype(str).str.strip()
    
    # 2. Dapatkan daftar No. Pesanan yang punya 'No. Pengajuan'
    potential_return_orders = rekap_df[
        rekap_df['No. Pengajuan'].notna() & 
        (rekap_df['No. Pengajuan'] != 'nan') & 
        (rekap_df['No. Pengajuan'] != '')
    ]['No. Pesanan'].unique()
    
    # 3. Siapkan list untuk menampung No. Pesanan Full vs Partial
    full_return_orders = set()
    partial_return_orders = set()
    
    # 4. Siapkan dict untuk menyimpan item-item yang diretur sebagian
    #    Format: { 'No. Pesanan': { 'keys': set(...), 'count': X } }
    partial_return_items_map = {}
    
    # 5. Iterasi HANYA pada No. Pesanan yang berpotensi retur
    for order_id in potential_return_orders:
        # 6. Cek di order_all_df (order_df ASLI)
        order_details = order_df[order_df['No. Pesanan'] == order_id]
        
        if order_details.empty:
            continue 
            
        total_items_in_order = len(order_details)
        
        # 7. Cek 'Status Pembatalan/ Pengembalian'
        returned_items = order_details[order_details['Status Pembatalan/ Pengembalian'] == 'Permintaan Disetujui']
        returned_items_count = len(returned_items)
        
        if returned_items_count == 0:
            # Punya No. Pengajuan tapi tidak ada 'Permintaan Disetujui'
            continue 
            
        # 8. Tentukan Tipe Retur
        if returned_items_count > 0 and returned_items_count == total_items_in_order:
            # FULL RETURN
            full_return_orders.add(order_id)
        elif returned_items_count > 0 and returned_items_count < total_items_in_order:
            # PARTIAL RETURN
            partial_return_orders.add(order_id)
            
            # 9. Simpan (Nama Produk, Nama Variasi) dari item yang diretur
            returned_item_keys = [
                (row['Nama Produk'], row['Nama Variasi']) 
                for _, row in returned_items.iterrows()
            ]
            partial_return_items_map[order_id] = {
                'keys': set(returned_item_keys),
                'count': returned_items_count # Simpan jumlah item retur
            }
    
    # REVISI 2: Gabungkan Nama Produk dan Variasi untuk produk spesifik
    produk_khusus_raw = [
        "CUSTOM AL QURAN MENGENANG/WAFAT 40/100/1000 HARI",
        "AL QUR'AN GOLD TERMURAH",
        "Alquran Cover Emas Kertas HVS Al Aqeel Gold Murah",
        "AL-QUR'AN SAKU A7 MAHEER HAFALAN AL QUR'AN",
        "AL QUR'AN NON TERJEMAH AL AQEEL A5 KERTAS KORAN WAKAF",
        "AL QUR'AN NON TERJEMAH Al AQEEL A5 KERTAS KORAN WAKAF",
        "AL-QURAN AL AQEEL SILVER TERMURAH", # <-- TAMBAHKAN INI
        "AL-QUR'AN TERJEMAH HC AL ALEEM A5",
        "AL QUR'AN EDISI TAHLILAN 30 Juz + Doa Tahlil | Pengganti Buku Yasin | Al Aqeel A6 Pastel HVS Edisi Tahlilan",
        "AL QUR'AN A6 NON TERJEMAH HVS WARNA PASTEL",
        "Paket Wakaf Murah 50 pcs Alquran Al Aqeel | Alquran 18 Baris"
    ]
    # Kondisi dimana Nama Produk ada dalam daftar produk_khusus
    produk_khusus = [re.sub(r'\s+', ' ', name.replace('\xa0', ' ')).strip() for name in produk_khusus_raw]

    if 'Nama Produk' in rekap_df.columns:
        rekap_df['Nama Produk Clean Temp'] = rekap_df['Nama Produk'].astype(str).str.replace('\xa0', ' ').str.replace(r'\s+', ' ', regex=True).str.strip()
        kondisi = rekap_df['Nama Produk Clean Temp'].isin(produk_khusus)
    else:
        kondisi = pd.Series([False] * len(rekap_df), index=rekap_df.index)
    
    if 'Nama Variasi' in rekap_df.columns:
        new_product_names = rekap_df.loc[kondisi, 'Nama Produk'].copy()
    
        for idx in new_product_names.index:
            nama_produk_asli = rekap_df.loc[idx, 'Nama Produk'] # Ambil nama produk asli (belum bersih)
            nama_produk_clean = rekap_df.loc[idx, 'Nama Produk Clean Temp'] # Ambil nama produk bersih
            nama_variasi_ori = rekap_df.loc[idx, 'Nama Variasi']
    
            if pd.notna(nama_variasi_ori):
                var_str = str(nama_variasi_ori).strip()
                part_to_append = ''
    
                # --- LOGIKA KHUSUS UNTUK PRODUK CUSTOM ---
                produk_yang_ambil_full_variasi = [
                    "CUSTOM AL QURAN MENGENANG", 
                    "AL QUR'AN GOLD TERMURAH",
                    "Alquran Cover Emas Kertas HVS Al Aqeel Gold Murah",
                    "AL-QUR'AN SAKU A7 MAHEER HAFALAN AL QUR'AN",
                    "AL-QURAN AL AQEEL SILVER TERMURAH",
                    "Paket Wakaf Murah 50 pcs Alquran Al Aqeel | Alquran 18 Baris"
                ]
                if any(produk in nama_produk_clean for produk in produk_yang_ambil_full_variasi):
                    # REVISI: Ambil seluruh string variasi, jangan di-split
                    part_to_append = var_str
                # --- AKHIR LOGIKA KHUSUS ---
                # 3. TAHLILAN (Ambil setelah koma)
                elif "AL QUR'AN EDISI TAHLILAN 30 Juz + Doa Tahlil | Pengganti Buku Yasin | Al Aqeel A6 Pastel HVS Edisi Tahlilan" in nama_produk_clean:
                    if ',' in var_str:
                        part_to_append = var_str.split(',', 1)[-1].strip() # Ambil setelah koma
                    else:
                        part_to_append = var_str # Fallback jika tidak ada koma (misal "Tidak custom")

                # 4. AL ALEEM (QPP Only)
                elif "AL-QUR'AN TERJEMAH HC AL ALEEM A5" in nama_produk_clean:
                    if 'QPP' in var_str.upper():
                        part_to_append = 'QPP'
                    elif 'HVS' in var_str.upper():
                        part_to_append = 'HVS'
                    elif 'KORAN' in var_str.upper():
                        part_to_append = 'KORAN'
                    # else: part_to_append tetap ''
                        
                elif "AL QUR'AN NON TERJEMAH Al AQEEL A5 KERTAS KORAN WAKAF" in nama_produk_clean or "AL QUR'AN A6 NON TERJEMAH HVS WARNA PASTEL" in nama_produk_clean:
                    var_upper = var_str.upper()
                    # Cari "PAKET ISI X" atau "SATUAN"
                    paket_match = re.search(r'(PAKET\s*ISI\s*\d+)', var_upper)
                    satuan_match = 'SATUAN' in var_upper
                
                    
                    if paket_match:
                        part_to_append = paket_match.group(1) # Hasilnya 'PAKET ISI 7'
                    elif satuan_match:
                        part_to_append = 'SATUAN'
                    else:
                        # --- LOGIKA FALLBACK TAMBAHAN ---
                        # Jika bukan PAKET/SATUAN, jalankan logika generik
                        if ',' in var_str:
                            parts = [p.strip().upper() for p in var_str.split(',')]
                            size_keywords = {'QPP', 'A5', 'B5', 'A6', 'A7', 'HVS', 'KORAN'}
                            relevant_parts = [p for p in parts if p in size_keywords]
                            if relevant_parts:
                                part_to_append = relevant_parts[0]
                        else:
                            part_to_append = var_str

                # --- Akhir Logika Lama ---
    
                # Gabungkan HANYA jika part_to_append tidak kosong
                if part_to_append:
                    new_product_names.loc[idx] = f"{nama_produk_asli} ({part_to_append})"
    
        rekap_df.loc[kondisi, 'Nama Produk'] = new_product_names
    
    if 'Nama Produk Clean Temp' in rekap_df.columns:
        rekap_df.drop(columns=['Nama Produk Clean Temp'], inplace=True)

    # Gabungkan dengan data seller conversion
    iklan_per_pesanan = seller_conv_df.groupby('Kode Pesanan')['Pengeluaran(Rp)'].sum().reset_index()
    rekap_df = pd.merge(rekap_df, iklan_per_pesanan, left_on='No. Pesanan', right_on='Kode Pesanan', how='left')
    rekap_df['Pengeluaran(Rp)'] = rekap_df['Pengeluaran(Rp)'].fillna(0)

    # 1. Pastikan Total Harga Produk ada dan numerik
    rekap_df['Total Harga Produk'] = rekap_df.get('Total Harga Produk', 0).fillna(0)
    
    # 2. Hitung biaya baru berdasarkan Total Harga Produk (ini berlaku per-baris/per-produk)
    # rekap_df['Biaya Adm 8%'] = rekap_df['Total Harga Produk'] * 0.08
    # rekap_df['Biaya Layanan 2%'] = rekap_df['Total Harga Produk'] * 0.02
    # rekap_df['Biaya Layanan Gratis Ongkir Xtra 4,5%'] = rekap_df['Total Harga Produk'] * 0.045
    
    # 3. Hitung Biaya Proses Pesanan yang dibagi rata
    #    Hitung dulu ada berapa produk dalam satu pesanan
    product_count_per_order = rekap_df.groupby('No. Pesanan')['No. Pesanan'].transform('size')

    # # 3. Siapkan data 'Biaya Layanan Promo XTRA' dari service_fee_df
    # #    (Fungsi clean_and_convert_to_numeric sudah ada di file Anda)
    # service_fee_subset = service_fee_df[['No. Pesanan', 'Biaya Layanan Promo XTRA']].copy()
    # service_fee_subset['No. Pesanan'] = service_fee_subset['No. Pesanan'].astype(str)
    
    # # Gunakan fungsi clean yang ada, lalu .abs() untuk menghilangkan minus
    # service_fee_subset['BiayaLayananPromo_Clean'] = clean_and_convert_to_numeric(service_fee_subset['Biaya Layanan Promo XTRA']).abs()
    
    # # Agregasi (sum) untuk jaga-jaga jika ada duplikat no. pesanan di file service fee
    # service_fee_agg = service_fee_subset.groupby('No. Pesanan')['BiayaLayananPromo_Clean'].sum().reset_index()
    
    # # 4. Gabungkan (merge) data biaya layanan ini ke rekap_df
    # rekap_df = pd.merge(rekap_df, service_fee_agg, on='No. Pesanan', how='left')
    # rekap_df['BiayaLayananPromo_Clean'] = rekap_df['BiayaLayananPromo_Clean'].fillna(0)

    rekap_df['Total Penghasilan Dibagi'] = (rekap_df['Total Penghasilan'] / product_count_per_order).fillna(0)

    # Bersihkan kolom keuangan yang akan kita gunakan (aman jika sudah numerik)
    rekap_df['Voucher dari Penjual'] = clean_and_convert_to_numeric(rekap_df['Voucher disponsor oleh Penjual'])
    rekap_df['Promo Gratis Ongkir dari Penjual'] = clean_and_convert_to_numeric(rekap_df['Promo Gratis Ongkir dari Penjual'])
    # Pastikan kolom ongkir retur dibersihkan TANPA abs()

    # Buat kolom 'Dibagi' untuk alokasi per produk
    rekap_df['Voucher dari Penjual Dibagi'] = (rekap_df['Voucher dari Penjual'] / product_count_per_order).fillna(0).abs()
    rekap_df['Gratis Ongkir dari Penjual Dibagi'] = (rekap_df['Promo Gratis Ongkir dari Penjual'] / product_count_per_order).fillna(0).abs()
    
    #    Bagi 1250 dengan jumlah produk tersebut
    rekap_df['Biaya Proses Pesanan Dibagi'] = 1250 / product_count_per_order

    basis_biaya = rekap_df['Total Harga Produk'] - rekap_df['Voucher dari Penjual Dibagi']
    # rekap_df['Biaya Adm 8%'] = basis_biaya * 0.08
    # Ambil tahun dari kolom Waktu Pesanan Dibuat
    tahun_pesanan = pd.to_datetime(rekap_df['Waktu Pesanan Dibuat']).dt.year
    
    # Rumus dinamis: 2026 (9%), selain itu/2025 (8%)
    rekap_df['Biaya Adm 8%'] = np.where(tahun_pesanan == 2026, basis_biaya * 0.09, basis_biaya * 0.08)
    # rekap_df['Biaya Layanan 2%'] = basis_biaya * 0.02
    rekap_df['Biaya Layanan Gratis Ongkir Xtra 4,5%'] = basis_biaya * 0.045
    rekap_df['Biaya Layanan 2%'] = 0
    # rekap_df = rekap_df.drop(columns=['BiayaLayananPromo_Clean'], errors='ignore')
    
    # 4. Terapkan logika "hanya di baris pertama" HANYA untuk biaya yang benar-benar per-pesanan
    order_level_costs = [
        # 'Voucher dari Penjual', 
        'Pengeluaran(Rp)',
        'Total Penghasilan' 
        # 'Biaya Administrasi', 'Biaya Layanan', dan 'Biaya Proses Pesanan' DIHAPUS dari sini
    ]
    is_first_item_mask = ~rekap_df.duplicated(subset='No. Pesanan', keep='first')
    
    for col in order_level_costs:
        if col in rekap_df.columns:
            rekap_df[col] = rekap_df[col].fillna(0)
            rekap_df[col] = rekap_df[col] * is_first_item_mask

    # 5. Pastikan semua biaya bernilai positif (menghilangkan tanda minus)
    cost_columns_to_abs = [
        'Voucher dari Penjual', 'Pengeluaran(Rp)', 'Biaya Administrasi', 
        'Biaya Layanan 2%', 'Biaya Layanan Gratis Ongkir Xtra 4,5%', 
        'Biaya Proses Pesanan' # <-- Cukup kolom asli
    ]
    for col in cost_columns_to_abs:
        if col in rekap_df.columns:
            rekap_df[col] = rekap_df[col].abs()

    # Kalkulasi Penjualan Netto per baris produk
    rekap_df['Penjualan Netto'] = (
        rekap_df.get('Total Harga Produk', 0) -
        rekap_df.get('Voucher dari Penjual Dibagi', 0) -     # <-- DIUBAH
        rekap_df.get('Pengeluaran(Rp)', 0) -
        rekap_df.get('Biaya Adm 8%', 0) -
        rekap_df.get('Biaya Layanan 2%', 0) -
        rekap_df.get('Biaya Layanan Gratis Ongkir Xtra 4,5%', 0) -
        rekap_df.get('Biaya Proses Pesanan Dibagi', 0) -
        rekap_df.get('Gratis Ongkir dari Penjual Dibagi', 0) # <-- DITAMBAH
    )

    # Urutkan berdasarkan No. Pesanan untuk memastikan produk dalam pesanan yang sama berkelompok
    rekap_df.sort_values(by='No. Pesanan', inplace=True)
    rekap_df.reset_index(drop=True, inplace=True)

    # # Terapkan logika retur: nol-kan semua kolom pendapatan/biaya dan isi Total Penghasilan (Netto)
    # kondisi_retur_final = rekap_df['No. Pesanan'].isin(returned_orders_list)
    
    # if not rekap_df[kondisi_retur_final].empty:
    #     cols_to_zero_out = [
    #         'Jumlah Terjual', 'Harga Setelah Diskon', 'Total Harga Produk',
    #         'Voucher dari Penjual Dibagi', 'Pengeluaran(Rp)', 'Biaya Adm 8%', 
    #         'Biaya Layanan 2%', 'Biaya Layanan Gratis Ongkir Xtra 4,5%', 
    #         'Biaya Proses Pesanan Dibagi', 'Gratis Ongkir dari Penjual Dibagi'
    #         # 'Penjualan Netto' dihapus dari daftar ini
    #     ]
    #     # Pastikan kolom ada sebelum mencoba meng-nol-kan
    #     valid_cols_to_zero = [col for col in cols_to_zero_out if col in rekap_df.columns]
        
    #     # Set semua kolom kalkulasi ke 0 untuk baris retur
    #     rekap_df.loc[kondisi_retur_final, valid_cols_to_zero] = 0
        
    #     # Atur 'Penjualan Netto' ke nilai 'Total Penghasilan Dibagi' (yang negatif)
    #     rekap_df.loc[kondisi_retur_final, 'Penjualan Netto'] = rekap_df.loc[kondisi_retur_final, 'Total Penghasilan Dibagi']

    cols_to_zero_out = [
        # 'Jumlah Terjual', 'Harga Setelah Diskon', 'Total Harga Produk',
        'Voucher dari Penjual Dibagi', 'Pengeluaran(Rp)', 'Biaya Adm 8%', 
        'Biaya Layanan 2%', 'Biaya Layanan Gratis Ongkir Xtra 4,5%', 
        'Biaya Proses Pesanan Dibagi', 'Gratis Ongkir dari Penjual Dibagi'
    ]
    valid_cols_to_zero = [col for col in cols_to_zero_out if col in rekap_df.columns]
    
    # B. Proses FULL RETURN
    if full_return_orders:
        kondisi_full_retur = rekap_df['No. Pesanan'].isin(full_return_orders)
        if kondisi_full_retur.any():
            # 1. Nol-kan kolom kalkulasi
            rekap_df.loc[kondisi_full_retur, valid_cols_to_zero] = 0
            # 2. Set 'Penjualan Netto' ke 'Total Penghasilan Dibagi' (yang sudah negatif)
            rekap_df.loc[kondisi_full_retur, 'Penjualan Netto'] = rekap_df.loc[kondisi_full_retur, 'Total Penghasilan Dibagi']

    # C. Proses PARTIAL RETURN
    if partial_return_orders:
        # 1. Bersihkan 'Jumlah Pengembalian Dana ke Pembeli' dan siapkan pembaginya
        if 'Jumlah Pengembalian Dana ke Pembeli' not in rekap_df.columns:
            rekap_df['Jumlah Pengembalian Dana ke Pembeli'] = 0
        
        # rekap_df['Jumlah Pengembalian Dana ke Pembeli'] = clean_and_convert_to_numeric(rekap_df['Jumlah Pengembalian Dana ke Pembeli'])
        rekap_df['Jumlah Pengembalian Dana ke Pembeli'] = 0
        
        # Buat kolom baru untuk jumlah retur per pesanan
        # Map-kan 'count' dari dict yang kita buat
        rekap_df['__return_count__'] = rekap_df['No. Pesanan'].map(
            lambda x: partial_return_items_map.get(x, {}).get('count', 1) # default 1 utk hindari /0
        )
        
        # Hitung nilai pengembalian per item retur
        rekap_df['Pengembalian Dana Per Item'] = (
            rekap_df['Jumlah Pengembalian Dana ke Pembeli'] / rekap_df['__return_count__']
        ).fillna(0)
        
        # 2. Identifikasi baris-baris yang merupakan item retur parsial
        def is_partial_return_item(row):
            order_id = row['No. Pesanan']
            if order_id not in partial_return_items_map:
                return False
            
            item_key = (row['Nama Produk'], row['Nama Variasi'])
            return item_key in partial_return_items_map[order_id]['keys']

        kondisi_partial_item = rekap_df.apply(is_partial_return_item, axis=1)
        
        # 3. Terapkan logika untuk item-item tersebut
        if kondisi_partial_item.any():
            # 3a. Nol-kan kolom kalkulasi
            rekap_df.loc[kondisi_partial_item, valid_cols_to_zero] = 0
            # 3b. Set 'Penjualan Netto' ke 'Pengembalian Dana Per Item'
            rekap_df.loc[kondisi_partial_item, 'Penjualan Netto'] = rekap_df.loc[kondisi_partial_item, 'Pengembalian Dana Per Item']
            
        # Hapus kolom bantu
        rekap_df = rekap_df.drop(columns=['__return_count__', 'Pengembalian Dana Per Item'], errors='ignore')
    
    # Buat DataFrame Final
    rekap_final = pd.DataFrame({
        'No.': np.arange(1, len(rekap_df) + 1),
        'No. Pesanan': rekap_df['No. Pesanan'],
        'Waktu Pesanan Dibuat': rekap_df['Waktu Pesanan Dibuat'],
        'Waktu Dana Dilepas': rekap_df['Tanggal Dana Dilepaskan'],
        'Nama Produk': rekap_df['Nama Produk'],
        'Jumlah Terjual': rekap_df['Jumlah Terjual'],
        'Harga Satuan': rekap_df['Harga Setelah Diskon'],
        'Total Harga Produk': rekap_df['Total Harga Produk'],
        'Voucher Ditanggung Penjual': rekap_df.get('Voucher dari Penjual Dibagi', 0),
        'Biaya Komisi AMS + PPN Shopee': rekap_df.get('Pengeluaran(Rp)', 0),
        'Biaya Adm 8%': rekap_df.get('Biaya Adm 8%', 0),
        'Biaya Layanan 2%': rekap_df.get('Biaya Layanan 2%', 0),
        'Biaya Layanan Gratis Ongkir Xtra 4,5%': rekap_df.get('Biaya Layanan Gratis Ongkir Xtra 4,5%', 0),
        'Biaya Proses Pesanan': rekap_df.get('Biaya Proses Pesanan Dibagi', 0),
        'Gratis Ongkir dari Penjual': rekap_df.get('Gratis Ongkir dari Penjual Dibagi', 0), # <-- DITAMBAH
        'Total Penghasilan': rekap_df['Penjualan Netto'],
        'Metode Pembayaran': rekap_df.get('Metode pembayaran pembeli', '')
    })

    # --- PERUBAIKAN 4: Mengosongkan sel duplikat untuk pesanan multi-produk ---
    cols_to_blank = ['No. Pesanan', 'Waktu Pesanan Dibuat', 'Waktu Dana Dilepas']
    rekap_final.loc[rekap_final['No. Pesanan'].duplicated(), cols_to_blank] = ''

    return rekap_final.fillna(0)

def process_rekap_pacific(order_df, income_df, seller_conv_df):
    """
    Fungsi untuk memproses sheet 'REKAP' KHUSUS untuk Pacific Bookstore.
    Perbedaan utama: Biaya Layanan dihitung dari Total Harga Produk.
    """
    # Bagian ini sama persis dengan fungsi rekap sebelumnya
    order_agg = order_df.groupby(['No. Pesanan', 'Nama Produk' ,'Nama Variasi']).agg({
        'Jumlah': 'sum',
        'Harga Setelah Diskon': 'first',
        'Total Harga Produk': 'sum'
        #'Nama Variasi': 'first'
    }).reset_index()
    order_agg.rename(columns={'Jumlah': 'Jumlah Terjual'}, inplace=True)

    income_df['No. Pesanan'] = income_df['No. Pesanan'].astype(str)
    order_agg['No. Pesanan'] = order_agg['No. Pesanan'].astype(str)
    seller_conv_df['Kode Pesanan'] = seller_conv_df['Kode Pesanan'].astype(str)
    
    rekap_df = pd.merge(income_df, order_agg, on='No. Pesanan', how='left')

    # # 1. Pastikan 'Total Penghasilan' (dari income_df) adalah numerik
    # rekap_df['Total Penghasilan'] = clean_and_convert_to_numeric(rekap_df['Total Penghasilan'])
    
    # # 2. Tandai baris retur BARU: Dapatkan daftar No. Pesanan yang diretur
    # #    Ini adalah No. Pesanan di mana SETIDAKNYA SATU baris memiliki Total Penghasilan < 0
    # returned_orders_list = rekap_df[rekap_df['Total Penghasilan'] < 0]['No. Pesanan'].unique()
    # 1. Pastikan 'No. Pengajuan' ada dan bersih (di rekap_df, dari income_dilepas)
    if 'No. Pengajuan' not in rekap_df.columns:
        rekap_df['No. Pengajuan'] = np.nan # Buat kolomnya jika tidak ada
    rekap_df['No. Pengajuan'] = rekap_df['No. Pengajuan'].astype(str).str.strip()
    
    # 2. Dapatkan daftar No. Pesanan yang punya 'No. Pengajuan'
    potential_return_orders = rekap_df[
        rekap_df['No. Pengajuan'].notna() & 
        (rekap_df['No. Pengajuan'] != 'nan') & 
        (rekap_df['No. Pengajuan'] != '')
    ]['No. Pesanan'].unique()
    
    # 3. Siapkan list untuk menampung No. Pesanan Full vs Partial
    full_return_orders = set()
    partial_return_orders = set()
    
    # 4. Siapkan dict untuk menyimpan item-item yang diretur sebagian
    #    Format: { 'No. Pesanan': { 'keys': set(...), 'count': X } }
    partial_return_items_map = {}
    
    # 5. Iterasi HANYA pada No. Pesanan yang berpotensi retur
    for order_id in potential_return_orders:
        # 6. Cek di order_all_df (order_df ASLI)
        order_details = order_df[order_df['No. Pesanan'] == order_id]
        
        if order_details.empty:
            continue 
            
        total_items_in_order = len(order_details)
        
        # 7. Cek 'Status Pembatalan/ Pengembalian'
        returned_items = order_details[order_details['Status Pembatalan/ Pengembalian'] == 'Permintaan Disetujui']
        returned_items_count = len(returned_items)
        
        if returned_items_count == 0:
            # Punya No. Pengajuan tapi tidak ada 'Permintaan Disetujui'
            continue 
            
        # 8. Tentukan Tipe Retur
        if returned_items_count > 0 and returned_items_count == total_items_in_order:
            # FULL RETURN
            full_return_orders.add(order_id)
        elif returned_items_count > 0 and returned_items_count < total_items_in_order:
            # PARTIAL RETURN
            partial_return_orders.add(order_id)
            
            # 9. Simpan (Nama Produk, Nama Variasi) dari item yang diretur
            returned_item_keys = [
                (row['Nama Produk'], row['Nama Variasi']) 
                for _, row in returned_items.iterrows()
            ]
            partial_return_items_map[order_id] = {
                'keys': set(returned_item_keys),
                'count': returned_items_count # Simpan jumlah item retur
            }
    
    # REVISI 2: Gabungkan Nama Produk dan Variasi untuk produk spesifik
    produk_khusus_raw = [
        "CUSTOM AL QURAN MENGENANG/WAFAT 40/100/1000 HARI",
        "AL QUR'AN GOLD TERMURAH",
        "Alquran Cover Emas Kertas HVS Al Aqeel Gold Murah",
        "TERBARU Al Quran Edisi Tahlilan Pengganti Buku Yasin Al Aqeel A6 Kertas HVS | SURABAYA | Mushaf Untuk Pengajian Kado Islami Hampers",
        "Al Quran Terjemah Al Aleem A5 HVS 15 Baris | SURABAYA | Alquran Untuk Pengajian Majelis Taklim",
        "Al Quran Saku Resleting Al Quddus A7 QPP Cover Kulit | SURABAYA | Untuk Santri Traveler Muslim",
        "Al Quran Wakaf Ibtida Al Quddus A5 Kertas HVS | Alquran SURABAYA",
        "Al Fikrah Al Quran Terjemah Fitur Lengkap A5 Kertas HVS | Alquran SURABAYA",
        "Al Quddus Al Quran Wakaf Ibtida A5 Kertas HVS | Alquran SURABAYA",
        "Al Quran Terjemah Al Aleem A5 Kertas HVS 15 Baris | SURABAYA | Alquran Untuk Majelis Taklim Kajian",
        "Al Quran Terjemah Per Kata A5 | Tajwid 2 Warna | Alquran Al Fikrah HVS 15 Baris | SURABAYA",
        "Al Quran Saku Resleting Al Quddus A7 Cover Kulit Kertas QPP | Alquran SURABAYA",
        "Al Quran Saku Pastel Al Aqeel A6 Kertas HVS | SURABAYA | Alquran Untuk Wakaf Hadiah Islami Hampers",
        "Al Quran Untuk Wakaf Al Aqeel A5 Kertas Koran 18 Baris | SURABAYA | Alquran Hadiah Islami Hampers",
        "Al Qur'an Untuk Wakaf Al Aqeel A5 Kertas Koran 18 Baris",
        "Alquran Edisi Tahlilan Lebih Mulia Daripada Buku Yasin Biasa | Al Aqeel A6 Kertas HVS | SURABAYA |",
        "PAKET MURAH ALQURAN AL AQEEL MUSHAF NON TERJEMAHAN | SURABAYA | al quran Wakaf/Shodaqoh hadiah hampers islami"
    ]
    # Kondisi dimana Nama Produk ada dalam daftar produk_khusus
    produk_khusus = [re.sub(r'\s+', ' ', name.replace('\xa0', ' ')).strip() for name in produk_khusus_raw]

    if 'Nama Produk' in rekap_df.columns:
        rekap_df['Nama Produk Clean Temp'] = rekap_df['Nama Produk'].astype(str).str.replace('\xa0', ' ').str.replace(r'\s+', ' ', regex=True).str.strip()
        kondisi = rekap_df['Nama Produk Clean Temp'].isin(produk_khusus)
    else:
        kondisi = pd.Series([False] * len(rekap_df), index=rekap_df.index)
    
    if 'Nama Variasi' in rekap_df.columns:
        new_product_names = rekap_df.loc[kondisi, 'Nama Produk'].copy()
    
        for idx in new_product_names.index:
            nama_produk_asli = rekap_df.loc[idx, 'Nama Produk'] # Ambil nama produk asli (belum bersih)
            nama_produk_clean = rekap_df.loc[idx, 'Nama Produk Clean Temp'] # Ambil nama produk bersih
            nama_variasi_ori = rekap_df.loc[idx, 'Nama Variasi']
    
            if pd.notna(nama_variasi_ori):
                var_str = str(nama_variasi_ori).strip()
                part_to_append = ''
                
                val_raw = rekap_df.loc[idx, 'Harga Setelah Diskon']
                
                # 2. Bersihkan dan ubah ke integer agar bisa dibandingkan dengan angka
                try:
                    # Menghilangkan titik/koma jika ada dan konversi ke int
                    harga_satuan = int(float(str(val_raw).replace('.', '').replace(',', '')))
                except:
                    harga_satuan = 0
    
                # --- LOGIKA KHUSUS UNTUK PRODUK CUSTOM ---
                produk_yang_ambil_full_variasi = [
                    "CUSTOM AL QURAN MENGENANG", 
                    "AL QUR'AN GOLD TERMURAH",
                    "Alquran Cover Emas Kertas HVS Al Aqeel Gold Murah",
                    "AL-QUR'AN SAKU A7 MAHEER HAFALAN AL QUR'AN",
                    "AL QUR'AN EDISI TAHLILAN 30 Juz + Doa Tahlil | Pengganti Buku Yasin | Al Aqeel A6 Pastel HVS Edisi Tahlilan" # (Sesuaikan string ini)
                ]
                if any(produk in nama_produk_clean for produk in produk_yang_ambil_full_variasi):
                    # REVISI: Ambil seluruh string variasi, jangan di-split
                    part_to_append = var_str
                # --- AKHIR LOGIKA KHUSUS ---
                elif "PAKET MURAH ALQURAN AL AQEEL MUSHAF NON TERJEMAHAN | SURABAYA | al quran Wakaf/Shodaqoh hadiah hampers islami" in nama_produk_clean:
                    # Menghapus apapun yang ada di dalam kurung ( ) termasuk kurungnya
                    # Contoh: "A5 KORAN (MERAH)" menjadi "A5 KORAN"
                    part_to_append = re.sub(r'\(.*?\)', '', var_str).strip()
                elif "Alquran Edisi Tahlilan Lebih Mulia Daripada Buku Yasin Biasa | Al Aqeel A6 Kertas HVS | SURABAYA |" in nama_produk_clean:
                    # Asumsi format variasi: "WARNA, SPESIFIKASI" (misal: "Merah, sisipan 1 halaman")
                    if ',' in var_str:
                        # Ambil bagian SETELAH koma pertama (Spesifikasi)
                        spesifikasi = var_str.split(',', 1)[-1].strip()
                        part_to_append = spesifikasi
                    else:
                        # Jika tidak ada koma, mungkin formatnya beda atau cuma 1 kata.
                        # Cek apakah ini WARNA (jika ya, abaikan). Jika bukan warna, anggap spesifikasi.
                        warna_keywords = ['MERAH', 'COKLAT', 'BIRU', 'UNGU', 'HIJAU', 'RANDOM', 'HITAM']
                        is_warna = any(w in var_str.upper() for w in warna_keywords)
                        
                        if not is_warna:
                            part_to_append = var_str # Ambil jika bukan warna
                        else:
                            part_to_append = '' # Abaikan jika cuma warna

                # elif "Al Quran Saku Pastel Al Aqeel A6 Kertas HVS | SURABAYA | Alquran Untuk Wakaf Hadiah Islami Hampers" in nama_produk_clean:

                #     harga = str(rekap_df['Harga Setelah Diskon']).replace('.', '').replace(',', '').strip()
                
                #     if harga == "19200":
                #         part_to_append = "GROSIR 3-4"
                #     elif harga == "18900":
                #         part_to_append = "GROSIR 5-6"
                #     elif harga == "18600":
                #         part_to_append = "GROSIR > 7"
                #     else:
                #         part_to_append = ""
                
                # elif "Al Quran Untuk Wakaf Al Aqeel A5 Kertas Koran 18 Baris | SURABAYA | Alquran Hadiah Islami Hampers" in nama_produk_clean:
                
                #     harga = str(rekap_df['Harga Setelah Diskon']).replace('.', '').replace(',', '').strip()
                
                #     if harga == "21550":
                #         part_to_append = "GROSIR 3-4"
                #     elif harga == "21300":
                #         part_to_append = "GROSIR 5-6"
                #     elif harga == "21000":
                #         part_to_append = "GROSIR > 7"
                #     else:
                #         part_to_append = ""
                            
                # elif "Al Quran Saku Pastel Al Aqeel A6 Kertas HVS | SURABAYA | Alquran Untuk Wakaf Hadiah Islami Hampers" in nama_produk_clean or "Al Quran Untuk Wakaf Al Aqeel A5 Kertas Koran 18 Baris | SURABAYA | Alquran Hadiah Islami Hampers" in nama_produk_clean or "Al Qur'an Untuk Wakaf Al Aqeel A5 Kertas Koran 18 Baris" in nama_produk_clean:
                elif "Al Quran Saku Pastel Al Aqeel A6 Kertas HVS | SURABAYA | Alquran Untuk Wakaf Hadiah Islami Hampers" in nama_produk_clean:
                    if harga_satuan == 19500:
                        part_to_append = "GROSIR 1-2"
                    elif harga_satuan == 19200:
                        part_to_append = "GROSIR 3-4"
                    elif harga_satuan == 18900:
                        part_to_append = "GROSIR 5-6"
                    elif harga_satuan == 18600:
                        part_to_append = "GROSIR > 7"
                
                elif "Al Quran Untuk Wakaf Al Aqeel A5 Kertas Koran 18 Baris | SURABAYA | Alquran Hadiah Islami Hampers" in nama_produk_clean:
                    if harga_satuan == 21800:
                        part_to_append = "GROSIR 1-2"
                    elif harga_satuan == 21550:
                        part_to_append = "GROSIR 3-4"
                    elif harga_satuan == 21300:
                        part_to_append = "GROSIR 5-6"
                    elif harga_satuan == 21000:
                        part_to_append = "GROSIR > 7"
    
            
                elif "Al Qur'an Untuk Wakaf Al Aqeel A5 Kertas Koran 18 Baris" in nama_produk_clean:
                    var_upper = var_str.upper()
                    # Cari "PAKET ISI X" atau "SATUAN"
                    paket_match = re.search(r'(PAKET\s*ISI\s*\d+)', var_upper)
                    satuan_match = 'SATUAN' in var_upper
                    
                    if paket_match:
                        part_to_append = paket_match.group(1) # Hasilnya 'PAKET ISI 7'
                    elif satuan_match:
                        part_to_append = 'SATUAN'
                    else:
                        # --- LOGIKA FALLBACK TAMBAHAN ---
                        # Jika bukan PAKET/SATUAN, jalankan logika generik
                        if ',' in var_str:
                            parts = [p.strip().upper() for p in var_str.split(',')]
                            size_keywords = {'QPP', 'A5', 'B5', 'A6', 'A7', 'HVS', 'KORAN'}
                            relevant_parts = [p for p in parts if p in size_keywords]
                            if relevant_parts:
                                part_to_append = relevant_parts[0]
                        else:
                            part_to_append = var_str

                # --- Akhir Logika Lama ---
    
                # Gabungkan HANYA jika part_to_append tidak kosong
                if part_to_append:
                    new_product_names.loc[idx] = f"{nama_produk_asli} ({part_to_append})"
    
        rekap_df.loc[kondisi, 'Nama Produk'] = new_product_names
    
    if 'Nama Produk Clean Temp' in rekap_df.columns:
        rekap_df.drop(columns=['Nama Produk Clean Temp'], inplace=True)

    iklan_per_pesanan = seller_conv_df.groupby('Kode Pesanan')['Pengeluaran(Rp)'].sum().reset_index()
    rekap_df = pd.merge(rekap_df, iklan_per_pesanan, left_on='No. Pesanan', right_on='Kode Pesanan', how='left')
    rekap_df['Pengeluaran(Rp)'] = rekap_df['Pengeluaran(Rp)'].fillna(0)

    # --- LOGIKA BARU UNTUK Pacifik Bookstore ---
    # 1. Pastikan Total Harga Produk ada dan numerik
    rekap_df['Total Harga Produk'] = rekap_df.get('Total Harga Produk', 0).fillna(0)
    
    # 2. Hitung biaya baru berdasarkan Total Harga Produk (ini berlaku per-baris/per-produk)
    # rekap_df['Biaya Adm 8%'] = rekap_df['Total Harga Produk'] * 0.08
    # rekap_df['Biaya Layanan 2%'] = rekap_df['Total Harga Produk'] * 0.02
    # rekap_df['Biaya Layanan Gratis Ongkir Xtra 4,5%'] = rekap_df['Total Harga Produk'] * 0.045
    # rekap_df['Biaya Adm 8%'] = 0
    # Hitung biaya berdasarkan (Total Harga Produk - Voucher Dibagi) 
    
    # 3. Hitung Biaya Proses Pesanan yang dibagi rata
    #    Hitung dulu ada berapa produk dalam satu pesanan
    product_count_per_order = rekap_df.groupby('No. Pesanan')['No. Pesanan'].transform('size')
    rekap_df['Total Penghasilan Dibagi'] = (rekap_df['Total Penghasilan'] / product_count_per_order).fillna(0)

    # Bersihkan kolom keuangan yang akan kita gunakan (aman jika sudah numerik)
    rekap_df['Voucher dari Penjual'] = clean_and_convert_to_numeric(rekap_df['Voucher disponsor oleh Penjual'])
    rekap_df['Promo Gratis Ongkir dari Penjual'] = clean_and_convert_to_numeric(rekap_df['Promo Gratis Ongkir dari Penjual'])

    # Buat kolom 'Dibagi' untuk alokasi per produk
    rekap_df['Voucher dari Penjual Dibagi'] = (rekap_df['Voucher dari Penjual'] / product_count_per_order).fillna(0).abs()
    rekap_df['Gratis Ongkir dari Penjual Dibagi'] = (rekap_df['Promo Gratis Ongkir dari Penjual'] / product_count_per_order).fillna(0).abs()
    
    #    Bagi 1250 dengan jumlah produk tersebut
    rekap_df['Biaya Proses Pesanan Dibagi'] = 1250 / product_count_per_order
    # rekap_df['Biaya Proses Pesanan Dibagi'] = 0

    basis_biaya = rekap_df['Total Harga Produk'] - rekap_df['Voucher dari Penjual Dibagi']
    # rekap_df['Biaya Adm 8%'] = basis_biaya * 0.08
    # Ambil tahun dari kolom Waktu Pesanan Dibuat
    tahun_pesanan = pd.to_datetime(rekap_df['Waktu Pesanan Dibuat']).dt.year
    
    # Rumus dinamis: 2026 (9%), selain itu/2025 (8%)
    rekap_df['Biaya Adm 8%'] = np.where(tahun_pesanan == 2026, basis_biaya * 0.09, basis_biaya * 0.08)
    # rekap_df['Biaya Layanan 2%'] = basis_biaya * 0.02
    # rekap_df['Biaya Layanan Gratis Ongkir Xtra 4,5%'] = basis_biaya * 0.045
    # rekap_df['Biaya Layanan 4,5%'] = basis_biaya * 0.045
    # --- PERUBAHAN: Ambil Biaya Layanan dari Income, dibagi jumlah produk ---
    # 1. Bersihkan kolom 'Biaya Layanan' dari income_df (pastikan ada)
    rekap_df['Biaya Layanan_Clean'] = clean_and_convert_to_numeric(rekap_df.get('Biaya Layanan', 0))
    
    # 2. Bagi per produk dan hilangkan minus (.abs())
    rekap_df['Biaya Layanan 4,5%'] = (rekap_df['Biaya Layanan_Clean'] / product_count_per_order).fillna(0).abs()
    rekap_df['Biaya Layanan Gratis Ongkir Xtra 4,5%'] = 0
    
    # 4. Terapkan logika "hanya di baris pertama" HANYA untuk biaya yang benar-benar per-pesanan
    order_level_costs = [
        # 'Voucher dari Penjual', 
        'Pengeluaran(Rp)',
        'Total Penghasilan'
        # 'Biaya Administrasi' dan 'Biaya Proses Pesanan' DIHAPUS dari sini
    ]
    is_first_item_mask = ~rekap_df.duplicated(subset='No. Pesanan', keep='first')
    
    for col in order_level_costs:
        if col in rekap_df.columns:
            rekap_df[col] = rekap_df[col].fillna(0)
            rekap_df[col] = rekap_df[col] * is_first_item_mask

    # Pastikan semua biaya bernilai positif
    cost_columns_to_abs = [
        'Voucher dari Penjual', 'Pengeluaran(Rp)', 'Biaya Administrasi', 
        'Biaya Layanan 2%', 'Biaya Layanan Gratis Ongkir Xtra 4,5%', 
        'Biaya Proses Pesanan'
    ]
    for col in cost_columns_to_abs:
        if col in rekap_df.columns:
            rekap_df[col] = rekap_df[col].abs()

    # Kalkulasi Penjualan Netto (sama seperti sebelumnya)
    rekap_df['Penjualan Netto'] = (
        rekap_df.get('Total Harga Produk', 0) -
        rekap_df.get('Voucher dari Penjual Dibagi', 0) -     # <-- DIUBAH
        rekap_df.get('Pengeluaran(Rp)', 0) -
        rekap_df.get('Biaya Adm 8%', 0) -
        rekap_df.get('Biaya Layanan 2%', 0) -
        rekap_df.get('Biaya Layanan Gratis Ongkir Xtra 4,5%', 0) -
        rekap_df.get('Biaya Proses Pesanan Dibagi', 0) -
        rekap_df.get('Gratis Ongkir dari Penjual Dibagi', 0) # <-- DITAMBAH
    )

    # Urutkan berdasarkan No. Pesanan untuk memastikan produk dalam pesanan yang sama berkelompok
    rekap_df.sort_values(by='No. Pesanan', inplace=True)
    rekap_df.reset_index(drop=True, inplace=True)

    # # Terapkan logika retur: nol-kan semua kolom pendapatan/biaya dan isi Total Penghasilan (Netto)
    # kondisi_retur_final = rekap_df['No. Pesanan'].isin(returned_orders_list)
    
    # if not rekap_df[kondisi_retur_final].empty:
    #     cols_to_zero_out = [
    #         'Jumlah Terjual', 'Harga Setelah Diskon', 'Total Harga Produk',
    #         'Voucher dari Penjual Dibagi', 'Pengeluaran(Rp)', 'Biaya Adm 8%', 
    #         'Biaya Layanan 2%', 'Biaya Layanan Gratis Ongkir Xtra 4,5%', 
    #         'Biaya Proses Pesanan Dibagi', 'Gratis Ongkir dari Penjual Dibagi'
    #         # 'Penjualan Netto' dihapus dari daftar ini
    #     ]
    #     # Pastikan kolom ada sebelum mencoba meng-nol-kan
    #     valid_cols_to_zero = [col for col in cols_to_zero_out if col in rekap_df.columns]
        
    #     # Set semua kolom kalkulasi ke 0 untuk baris retur
    #     rekap_df.loc[kondisi_retur_final, valid_cols_to_zero] = 0
        
    #     # Atur 'Penjualan Netto' ke nilai 'Total Penghasilan Dibagi' (yang negatif)
    #     rekap_df.loc[kondisi_retur_final, 'Penjualan Netto'] = rekap_df.loc[kondisi_retur_final, 'Total Penghasilan Dibagi']
    cols_to_zero_out = [
        # 'Jumlah Terjual', 'Harga Setelah Diskon', 'Total Harga Produk',
        'Voucher dari Penjual Dibagi', 'Pengeluaran(Rp)', 'Biaya Adm 8%', 
        'Biaya Layanan 2%', 'Biaya Layanan Gratis Ongkir Xtra 4,5%', 
        'Biaya Proses Pesanan Dibagi', 'Gratis Ongkir dari Penjual Dibagi'
    ]
    valid_cols_to_zero = [col for col in cols_to_zero_out if col in rekap_df.columns]
    
    # B. Proses FULL RETURN
    if full_return_orders:
        kondisi_full_retur = rekap_df['No. Pesanan'].isin(full_return_orders)
        if kondisi_full_retur.any():
            # 1. Nol-kan kolom kalkulasi
            rekap_df.loc[kondisi_full_retur, valid_cols_to_zero] = 0
            # 2. Set 'Penjualan Netto' ke 'Total Penghasilan Dibagi' (yang sudah negatif)
            rekap_df.loc[kondisi_full_retur, 'Penjualan Netto'] = rekap_df.loc[kondisi_full_retur, 'Total Penghasilan Dibagi']

    # C. Proses PARTIAL RETURN
    if partial_return_orders:
        # 1. Bersihkan 'Jumlah Pengembalian Dana ke Pembeli' dan siapkan pembaginya
        if 'Jumlah Pengembalian Dana ke Pembeli' not in rekap_df.columns:
            rekap_df['Jumlah Pengembalian Dana ke Pembeli'] = 0
        
        # rekap_df['Jumlah Pengembalian Dana ke Pembeli'] = clean_and_convert_to_numeric(rekap_df['Jumlah Pengembalian Dana ke Pembeli'])
        rekap_df['Jumlah Pengembalian Dana ke Pembeli'] = 0
        
        # Buat kolom baru untuk jumlah retur per pesanan
        # Map-kan 'count' dari dict yang kita buat
        rekap_df['__return_count__'] = rekap_df['No. Pesanan'].map(
            lambda x: partial_return_items_map.get(x, {}).get('count', 1) # default 1 utk hindari /0
        )
        
        # Hitung nilai pengembalian per item retur
        rekap_df['Pengembalian Dana Per Item'] = (
            rekap_df['Jumlah Pengembalian Dana ke Pembeli'] / rekap_df['__return_count__']
        ).fillna(0)
        
        # 2. Identifikasi baris-baris yang merupakan item retur parsial
        def is_partial_return_item(row):
            order_id = row['No. Pesanan']
            if order_id not in partial_return_items_map:
                return False
            
            item_key = (row['Nama Produk'], row['Nama Variasi'])
            return item_key in partial_return_items_map[order_id]['keys']

        kondisi_partial_item = rekap_df.apply(is_partial_return_item, axis=1)
        
        # 3. Terapkan logika untuk item-item tersebut
        if kondisi_partial_item.any():
            # 3a. Nol-kan kolom kalkulasi
            rekap_df.loc[kondisi_partial_item, valid_cols_to_zero] = 0
            # 3b. Set 'Penjualan Netto' ke 'Pengembalian Dana Per Item'
            rekap_df.loc[kondisi_partial_item, 'Penjualan Netto'] = rekap_df.loc[kondisi_partial_item, 'Pengembalian Dana Per Item']
            
        # Hapus kolom bantu
        rekap_df = rekap_df.drop(columns=['__return_count__', 'Pengembalian Dana Per Item'], errors='ignore')
    
    rekap_final = pd.DataFrame({
        'No.': np.arange(1, len(rekap_df) + 1),
        'No. Pesanan': rekap_df['No. Pesanan'],
        'Waktu Pesanan Dibuat': rekap_df['Waktu Pesanan Dibuat'],
        'Waktu Dana Dilepas': rekap_df['Tanggal Dana Dilepaskan'],
        'Nama Produk': rekap_df['Nama Produk'],
        'Jumlah Terjual': rekap_df['Jumlah Terjual'],
        'Harga Satuan': rekap_df['Harga Setelah Diskon'],
        'Total Harga Produk': rekap_df['Total Harga Produk'],
        'Voucher Ditanggung Penjual': rekap_df.get('Voucher dari Penjual Dibagi', 0),
        'Biaya Komisi AMS + PPN Shopee': rekap_df.get('Pengeluaran(Rp)', 0),
        'Biaya Adm 8%': rekap_df.get('Biaya Adm 8%', 0),
        'Biaya Layanan 4,5%': rekap_df.get('Biaya Layanan 4,5%', 0),
        'Biaya Layanan Gratis Ongkir Xtra 4,5%': rekap_df.get('Biaya Layanan Gratis Ongkir Xtra 4,5%', 0),
        'Biaya Proses Pesanan': rekap_df.get('Biaya Proses Pesanan Dibagi', 0), # <-- Gunakan kolom yang sudah dibagi
        'Gratis Ongkir dari Penjual': rekap_df.get('Gratis Ongkir dari Penjual Dibagi', 0), # <-- DITAMBAH
        'Total Penghasilan': rekap_df['Penjualan Netto'],
        'Metode Pembayaran': rekap_df.get('Metode pembayaran pembeli', '')
    })

    cols_to_blank = ['No. Pesanan', 'Waktu Pesanan Dibuat', 'Waktu Dana Dilepas']
    rekap_final.loc[rekap_final['No. Pesanan'].duplicated(), cols_to_blank] = ''

    return rekap_final.fillna(0)

def process_rekap_dama(order_df, income_df, seller_conv_df):
    """
    Fungsi untuk memproses sheet 'REKAP' KHUSUS untuk DAMA.ID STORE (Shopee).
    Biaya Adm, Layanan, dan Proses dihitung berdasarkan Total Harga Produk.
    """
    if 'Nama Variasi' in order_df.columns:
        order_df['Nama Variasi'] = order_df['Nama Variasi'].fillna('')
    else:
        order_df['Nama Variasi'] = ''
        
    # Bagian ini sama persis dengan fungsi rekap pacific/human
    order_agg = order_df.groupby(['No. Pesanan', 'Nama Produk', 'Nama Variasi']).agg({
        'Jumlah': 'sum',
        'Harga Setelah Diskon': 'first',
        'Total Harga Produk': 'sum'
        #'Nama Variasi': 'first'
    }).reset_index()
    order_agg.rename(columns={'Jumlah': 'Jumlah Terjual'}, inplace=True)

    income_df['No. Pesanan'] = income_df['No. Pesanan'].astype(str)
    order_agg['No. Pesanan'] = order_agg['No. Pesanan'].astype(str)
    # seller_conv_df['Kode Pesanan'] = seller_conv_df['Kode Pesanan'].astype(str)
    
    rekap_df = pd.merge(income_df, order_agg, on='No. Pesanan', how='left')

    # # 1. Pastikan 'Total Penghasilan' (dari income_df) adalah numerik
    # rekap_df['Total Penghasilan'] = clean_and_convert_to_numeric(rekap_df['Total Penghasilan'])
    
    # # 2. Tandai baris retur BARU: Dapatkan daftar No. Pesanan yang diretur
    # #    Ini adalah No. Pesanan di mana SETIDAKNYA SATU baris memiliki Total Penghasilan < 0
    # returned_orders_list = rekap_df[rekap_df['Total Penghasilan'] < 0]['No. Pesanan'].unique()
    # 1. Pastikan 'No. Pengajuan' ada dan bersih (di rekap_df, dari income_dilepas)
    if 'No. Pengajuan' not in rekap_df.columns:
        rekap_df['No. Pengajuan'] = np.nan # Buat kolomnya jika tidak ada
    rekap_df['No. Pengajuan'] = rekap_df['No. Pengajuan'].astype(str).str.strip()
    
    # 2. Dapatkan daftar No. Pesanan yang punya 'No. Pengajuan'
    potential_return_orders = rekap_df[
        rekap_df['No. Pengajuan'].notna() & 
        (rekap_df['No. Pengajuan'] != 'nan') & 
        (rekap_df['No. Pengajuan'] != '')
    ]['No. Pesanan'].unique()
    
    # 3. Siapkan list untuk menampung No. Pesanan Full vs Partial
    full_return_orders = set()
    partial_return_orders = set()
    
    # 4. Siapkan dict untuk menyimpan item-item yang diretur sebagian
    #    Format: { 'No. Pesanan': { 'keys': set(...), 'count': X } }
    partial_return_items_map = {}
    
    # 5. Iterasi HANYA pada No. Pesanan yang berpotensi retur
    for order_id in potential_return_orders:
        # 6. Cek di order_all_df (order_df ASLI)
        order_details = order_df[order_df['No. Pesanan'] == order_id]
        
        if order_details.empty:
            continue 
            
        total_items_in_order = len(order_details)
        
        # 7. Cek 'Status Pembatalan/ Pengembalian'
        returned_items = order_details[order_details['Status Pembatalan/ Pengembalian'] == 'Permintaan Disetujui']
        returned_items_count = len(returned_items)
        
        if returned_items_count == 0:
            # Punya No. Pengajuan tapi tidak ada 'Permintaan Disetujui'
            continue 
            
        # 8. Tentukan Tipe Retur
        if returned_items_count > 0 and returned_items_count == total_items_in_order:
            # FULL RETURN
            full_return_orders.add(order_id)
        elif returned_items_count > 0 and returned_items_count < total_items_in_order:
            # PARTIAL RETURN
            partial_return_orders.add(order_id)
            
            # 9. Simpan (Nama Produk, Nama Variasi) dari item yang diretur
            returned_item_keys = [
                (row['Nama Produk'], row['Nama Variasi']) 
                for _, row in returned_items.iterrows()
            ]
            partial_return_items_map[order_id] = {
                'keys': set(returned_item_keys),
                'count': returned_items_count # Simpan jumlah item retur
            }
    
    if not seller_conv_df.empty:
        seller_conv_df['Kode Pesanan'] = seller_conv_df['Kode Pesanan'].astype(str)
        iklan_per_pesanan = seller_conv_df.groupby('Kode Pesanan')['Pengeluaran(Rp)'].sum().reset_index()
        rekap_df = pd.merge(rekap_df, iklan_per_pesanan, left_on='No. Pesanan', right_on='Kode Pesanan', how='left')
        rekap_df['Pengeluaran(Rp)'] = rekap_df['Pengeluaran(Rp)'].fillna(0)
    else:
        # Jika file tidak ada (kosong), buat kolom 'Pengeluaran(Rp)' dan isi dengan 0
        rekap_df['Pengeluaran(Rp)'] = 0
    # --- AKHIR BLOK KONDISIONAL ---

    # produk_khusus = ["CUSTOM AL QURAN MENGENANG/WAFAT 40/100/1000 HARI", "AL QUR'AN GOLD TERMURAH"]
    # kondisi = rekap_df['Nama Produk'].isin(produk_khusus)
    # if 'Nama Variasi' in rekap_df.columns:
    #     rekap_df.loc[kondisi, 'Nama Produk'] = rekap_df['Nama Produk'] + ' ' + rekap_df['Nama Variasi'].fillna('').str.strip()
    # if 'Nama Variasi' in rekap_df.columns:
    #     # Ambil variasi, ganti NaN dengan string kosong
    #     variasi_clean = rekap_df['Nama Variasi'].fillna('').astype(str).str.strip()
        
    #     # Kondisi untuk menggabungkan: Variasi tidak kosong dan tidak '0'
    #     # (Dan jika Anda hanya ingin produk tertentu, tambahkan kondisi nama produk di sini)
    #     kondisi_gabung = (variasi_clean != '') & (variasi_clean != '0') & (variasi_clean != 'nan')
        
    #     # Gabungkan Nama Produk + Variasi hanya untuk baris yang memenuhi syarat
    #     # Gunakan .loc untuk memastikan kita tidak menimpa baris yang tidak punya variasi
    #     rekap_df.loc[kondisi_gabung, 'Nama Produk'] = (
    #         rekap_df.loc[kondisi_gabung, 'Nama Produk'] + ' (' + variasi_clean.loc[kondisi_gabung] + ')'
    #     )

    # --- LOGIKA PERHITUNGAN BIAYA UNTUK DAMA.ID STORE ---
    rekap_df['Total Harga Produk'] = rekap_df.get('Total Harga Produk', 0).fillna(0) 
    
    # Hitung Biaya Proses Pesanan yang dibagi rata
    product_count_per_order = rekap_df.groupby('No. Pesanan')['No. Pesanan'].transform('size')
    rekap_df['Total Penghasilan Dibagi'] = (rekap_df['Total Penghasilan'] / product_count_per_order).fillna(0)

    # Bersihkan kolom keuangan yang akan kita gunakan (aman jika sudah numerik)
    rekap_df['Voucher dari Penjual'] = clean_and_convert_to_numeric(rekap_df['Voucher disponsor oleh Penjual'])
    rekap_df['Promo Gratis Ongkir dari Penjual'] = clean_and_convert_to_numeric(rekap_df['Promo Gratis Ongkir dari Penjual'])

    # Buat kolom 'Dibagi' untuk alokasi per produk
    rekap_df['Voucher dari Penjual Dibagi'] = (rekap_df['Voucher dari Penjual'] / product_count_per_order).fillna(0).abs()
    rekap_df['Gratis Ongkir dari Penjual Dibagi'] = (rekap_df['Promo Gratis Ongkir dari Penjual'] / product_count_per_order).fillna(0).abs()
    
    rekap_df['Biaya Proses Pesanan Dibagi'] = 1250 / product_count_per_order

    # Hitung biaya berdasarkan Total Harga Produk
    # rekap_df['Biaya Adm 8%'] = rekap_df['Total Harga Produk'] * 0.08
    # rekap_df['Biaya Layanan 2%'] = rekap_df['Total Harga Produk'] * 0.02
    rekap_df['Biaya Layanan 2%'] = 0
    # rekap_df['Biaya Layanan Gratis Ongkir Xtra 4,5%'] = rekap_df['Total Harga Produk'] * 0.045

    # Hitung biaya berdasarkan (Total Harga Produk - Voucher Dibagi)
    basis_biaya = rekap_df['Total Harga Produk'] - rekap_df['Voucher dari Penjual Dibagi']
    # rekap_df['Biaya Adm 8%'] = basis_biaya * 0.08
    # Ambil tahun dari kolom Waktu Pesanan Dibuat
    tahun_pesanan = pd.to_datetime(rekap_df['Waktu Pesanan Dibuat']).dt.year
    
    # Rumus dinamis: 2026 (9%), selain itu/2025 (8%)
    rekap_df['Biaya Adm 8%'] = np.where(tahun_pesanan == 2026, basis_biaya * 0.09, basis_biaya * 0.08)
    # rekap_df['Biaya Layanan 2%'] = basis_biaya * 0.02
    rekap_df['Biaya Layanan Gratis Ongkir Xtra 4,5%'] = basis_biaya * 0.045
    # --- AKHIR LOGIKA DAMA.ID STORE ---
    
    # Terapkan logika "hanya di baris pertama" untuk biaya per-pesanan
    order_level_costs = [
        # 'Voucher dari Penjual', 
        'Pengeluaran(Rp)',
        'Total Penghasilan' 
        # Biaya Adm, Layanan, dan Proses Pesanan Dihapus karena dihitung per produk/dibagi
    ]
    is_first_item_mask = ~rekap_df.duplicated(subset='No. Pesanan', keep='first')
    
    for col in order_level_costs:
        if col in rekap_df.columns:
            rekap_df[col] = rekap_df[col].fillna(0)
            rekap_df[col] = rekap_df[col] * is_first_item_mask

    # Pastikan semua biaya bernilai positif
    cost_columns_to_abs = [
        'Voucher dari Penjual', 'Pengeluaran(Rp)', 'Biaya Adm 8%', 
        'Biaya Layanan 2%', 'Biaya Layanan Gratis Ongkir Xtra 4,5%', 
        # 'Biaya Proses Pesanan' tidak perlu di-abs karena sudah dibagi
    ]
    for col in cost_columns_to_abs:
        if col in rekap_df.columns:
             # Cek dulu apakah kolomnya numerik sebelum .abs()
             if pd.api.types.is_numeric_dtype(rekap_df[col]):
                  rekap_df[col] = rekap_df[col].abs()

    # Kalkulasi Penjualan Netto
    rekap_df['Penjualan Netto'] = (
        rekap_df.get('Total Harga Produk', 0) -
        rekap_df.get('Voucher dari Penjual Dibagi', 0) -     # <-- DIUBAH
        rekap_df.get('Pengeluaran(Rp)', 0) -
        rekap_df.get('Biaya Adm 8%', 0) -
        rekap_df.get('Biaya Layanan 2%', 0) -
        rekap_df.get('Biaya Layanan Gratis Ongkir Xtra 4,5%', 0) -
        rekap_df.get('Biaya Proses Pesanan Dibagi', 0) -
        rekap_df.get('Gratis Ongkir dari Penjual Dibagi', 0) # <-- DITAMBAH
    )

    # Urutkan berdasarkan No. Pesanan untuk memastikan produk dalam pesanan yang sama berkelompok
    rekap_df.sort_values(by='No. Pesanan', inplace=True)
    rekap_df.reset_index(drop=True, inplace=True)

    # # Terapkan logika retur: nol-kan semua kolom pendapatan/biaya dan isi Total Penghasilan (Netto)
    # kondisi_retur_final = rekap_df['No. Pesanan'].isin(returned_orders_list)
    
    # if not rekap_df[kondisi_retur_final].empty:
    #     cols_to_zero_out = [
    #         'Jumlah Terjual', 'Harga Setelah Diskon', 'Total Harga Produk',
    #         'Voucher dari Penjual Dibagi', 'Pengeluaran(Rp)', 'Biaya Adm 8%', 
    #         'Biaya Layanan 2%', 'Biaya Layanan Gratis Ongkir Xtra 4,5%', 
    #         'Biaya Proses Pesanan Dibagi', 'Gratis Ongkir dari Penjual Dibagi'
    #         # 'Penjualan Netto' dihapus dari daftar ini
    #     ]
    #     # Pastikan kolom ada sebelum mencoba meng-nol-kan
    #     valid_cols_to_zero = [col for col in cols_to_zero_out if col in rekap_df.columns]
        
    #     # Set semua kolom kalkulasi ke 0 untuk baris retur
    #     rekap_df.loc[kondisi_retur_final, valid_cols_to_zero] = 0
        
    #     # Atur 'Penjualan Netto' ke nilai 'Total Penghasilan Dibagi' (yang negatif)
    #     rekap_df.loc[kondisi_retur_final, 'Penjualan Netto'] = rekap_df.loc[kondisi_retur_final, 'Total Penghasilan Dibagi']
    cols_to_zero_out = [
        # 'Jumlah Terjual', 'Harga Setelah Diskon', 'Total Harga Produk',
        'Voucher dari Penjual Dibagi', 'Pengeluaran(Rp)', 'Biaya Adm 8%', 
        'Biaya Layanan 2%', 'Biaya Layanan Gratis Ongkir Xtra 4,5%', 
        'Biaya Proses Pesanan Dibagi', 'Gratis Ongkir dari Penjual Dibagi'
    ]
    valid_cols_to_zero = [col for col in cols_to_zero_out if col in rekap_df.columns]
    
    # B. Proses FULL RETURN
    if full_return_orders:
        kondisi_full_retur = rekap_df['No. Pesanan'].isin(full_return_orders)
        if kondisi_full_retur.any():
            # 1. Nol-kan kolom kalkulasi
            rekap_df.loc[kondisi_full_retur, valid_cols_to_zero] = 0
            # 2. Set 'Penjualan Netto' ke 'Total Penghasilan Dibagi' (yang sudah negatif)
            rekap_df.loc[kondisi_full_retur, 'Penjualan Netto'] = rekap_df.loc[kondisi_full_retur, 'Total Penghasilan Dibagi']

    # C. Proses PARTIAL RETURN
    if partial_return_orders:
        # 1. Bersihkan 'Jumlah Pengembalian Dana ke Pembeli' dan siapkan pembaginya
        if 'Jumlah Pengembalian Dana ke Pembeli' not in rekap_df.columns:
            rekap_df['Jumlah Pengembalian Dana ke Pembeli'] = 0
        
        # rekap_df['Jumlah Pengembalian Dana ke Pembeli'] = clean_and_convert_to_numeric(rekap_df['Jumlah Pengembalian Dana ke Pembeli'])
        rekap_df['Jumlah Pengembalian Dana ke Pembeli'] = 0
        
        # Buat kolom baru untuk jumlah retur per pesanan
        # Map-kan 'count' dari dict yang kita buat
        rekap_df['__return_count__'] = rekap_df['No. Pesanan'].map(
            lambda x: partial_return_items_map.get(x, {}).get('count', 1) # default 1 utk hindari /0
        )
        
        # Hitung nilai pengembalian per item retur
        rekap_df['Pengembalian Dana Per Item'] = (
            rekap_df['Jumlah Pengembalian Dana ke Pembeli'] / rekap_df['__return_count__']
        ).fillna(0)
        
        # 2. Identifikasi baris-baris yang merupakan item retur parsial
        def is_partial_return_item(row):
            order_id = row['No. Pesanan']
            if order_id not in partial_return_items_map:
                return False
            
            item_key = (row['Nama Produk'], row['Nama Variasi'])
            return item_key in partial_return_items_map[order_id]['keys']

        kondisi_partial_item = rekap_df.apply(is_partial_return_item, axis=1)
        
        # 3. Terapkan logika untuk item-item tersebut
        if kondisi_partial_item.any():
            # 3a. Nol-kan kolom kalkulasi
            rekap_df.loc[kondisi_partial_item, valid_cols_to_zero] = 0
            # 3b. Set 'Penjualan Netto' ke 'Pengembalian Dana Per Item'
            rekap_df.loc[kondisi_partial_item, 'Penjualan Netto'] = rekap_df.loc[kondisi_partial_item, 'Pengembalian Dana Per Item']
            
        # Hapus kolom bantu
        rekap_df = rekap_df.drop(columns=['__return_count__', 'Pengembalian Dana Per Item'], errors='ignore')
    
    rekap_final = pd.DataFrame({
        'No.': np.arange(1, len(rekap_df) + 1),
        'No. Pesanan': rekap_df['No. Pesanan'],
        'Waktu Pesanan Dibuat': rekap_df['Waktu Pesanan Dibuat'],
        'Waktu Dana Dilepas': rekap_df['Tanggal Dana Dilepaskan'],
        'Nama Produk': rekap_df['Nama Produk'],
        'Nama Variasi': rekap_df['Nama Variasi'],
        'Jumlah Terjual': rekap_df['Jumlah Terjual'],
        'Harga Satuan': rekap_df['Harga Setelah Diskon'],
        'Total Harga Produk': rekap_df['Total Harga Produk'],
        'Voucher Ditanggung Penjual': rekap_df.get('Voucher dari Penjual Dibagi', 0),
        'Biaya Komisi AMS + PPN Shopee': rekap_df.get('Pengeluaran(Rp)', 0),
        'Biaya Adm 8%': rekap_df.get('Biaya Adm 8%', 0),
        'Biaya Layanan 2%': rekap_df.get('Biaya Layanan 2%', 0),
        'Biaya Layanan Gratis Ongkir Xtra 4,5%': rekap_df.get('Biaya Layanan Gratis Ongkir Xtra 4,5%', 0),
        'Biaya Proses Pesanan': rekap_df.get('Biaya Proses Pesanan Dibagi', 0),
        'Gratis Ongkir dari Penjual': rekap_df.get('Gratis Ongkir dari Penjual Dibagi', 0), # <-- DITAMBAH
        'Total Penghasilan': rekap_df['Penjualan Netto'],
        'Metode Pembayaran': rekap_df.get('Metode pembayaran pembeli', '')
    })

    cols_to_blank = ['No. Pesanan', 'Waktu Pesanan Dibuat', 'Waktu Dana Dilepas']
    rekap_final.loc[rekap_final['No. Pesanan'].duplicated(), cols_to_blank] = ''

    return rekap_final.fillna(0)
    
def process_iklan(iklan_df):
    """Fungsi untuk memproses dan membuat sheet 'IKLAN'."""
    iklan_df['Nama Iklan Clean'] = iklan_df['Nama Iklan'].str.replace(r'\s*baris\s*\[\d+\]$', '', regex=True).str.strip()
    iklan_df['Nama Iklan Clean'] = iklan_df['Nama Iklan Clean'].str.replace(r'\s*\[\d+\]$', '', regex=True).str.strip()
    
    iklan_agg = iklan_df.groupby('Nama Iklan Clean').agg({
        'Dilihat': 'sum',
        'Jumlah Klik': 'sum',
        'Biaya': 'sum',
        'Produk Terjual': 'sum',
        'Omzet Penjualan': 'sum'
    }).reset_index()
    iklan_agg.rename(columns={'Nama Iklan Clean': 'Nama Iklan'}, inplace=True)

    total_row = pd.DataFrame({
        'Nama Iklan': ['TOTAL'],
        'Dilihat': [iklan_agg['Dilihat'].sum()],
        'Jumlah Klik': [iklan_agg['Jumlah Klik'].sum()],
        'Biaya': [iklan_agg['Biaya'].sum()],
        'Produk Terjual': [iklan_agg['Produk Terjual'].sum()],
        'Omzet Penjualan': [iklan_agg['Omzet Penjualan'].sum()]
    })
    
    iklan_final = pd.concat([iklan_agg, total_row], ignore_index=True)
    return iklan_final

def get_harga_beli_fuzzy(nama_produk, katalog_df, score_threshold_primary=80, score_threshold_fallback=75):
    """
    REVISI 3: Mencari harga beli dari satu dataframe katalog saja.
    """
    try:
        search_name = str(nama_produk).strip()
        if not search_name:
            return 0

        # Logika fuzzy matching langsung ke katalog_df
        s = search_name.upper()
        s_clean = re.sub(r'[^A-Z0-9\s×xX\-]', ' ', s)
        s_clean = re.sub(r'\s+', ' ', s_clean).strip()

        # 1) Deteksi ukuran
        ukuran_found = None
        ukuran_patterns = [
            r'\bA[0-9]\b', r'\bB[0-9]\b', r'\b\d{1,3}\s*[x×X]\s*\d{1,3}\b', r'\b\d{1,3}\s*CM\b'
        ]
        for pat in ukuran_patterns:
            m = re.search(pat, s_clean)
            if m:
                ukuran_found = m.group(0).replace(' ', '').upper()
                break

        # 2) Deteksi jenis kertas
        jenis_kertas_map = {
            'HVS': 'HVS', 'QPP': 'QPP', 'KORAN': 'KORAN', 'KK': 'KORAN', # Map KK ke KORAN
            'GLOSSY':'GLOSSY','DUPLEX':'DUPLEX','ART':'ART','COVER':'COVER',
            'MATT':'MATT','MATTE':'MATTE','CTP':'CTP','BOOK PAPER':'BOOK PAPER',
            'ART PAPER': 'ART PAPER', 'ART PAPER': 'Art Paper'
        }
        jenis_kertas_tokens_to_search = list(jenis_kertas_map.keys()) # Cari semua keys (termasuk KK)
        
        jenis_found = None
        s_clean_words = set(s_clean.split()) # Pisah kata-kata di nama produk
        
        for token_to_find in jenis_kertas_tokens_to_search:
            if token_to_find in s_clean_words: # Cek jika token ada sebagai kata utuh
                jenis_found = jenis_kertas_map[token_to_find] # Ambil nilai dari map (misal KORAN jika KK ditemukan)
                break # Ambil yang pertama ditemukan

        # 3) Filter kandidat
        candidates = katalog_df.copy()
        if ukuran_found:
            candidates = candidates[candidates['UKURAN_NORM'].str.contains(re.escape(ukuran_found), na=False)]
        if jenis_found and not candidates.empty:
            candidates = candidates[candidates['JENIS_KERTAS_NORM'].str.contains(jenis_found, na=False)]

        if candidates.empty:
            candidates = katalog_df.copy()

        # 4) Fuzzy matching
        best_score, best_price, best_title = 0, 0, ""
        for _, row in candidates.iterrows():
            title = str(row['JUDUL_NORM'])
            score = fuzz.token_set_ratio(s_clean, title)
            if score > best_score or (score == best_score and len(title) > len(best_title)):
                best_score, best_price, best_title = score, row.get('KATALOG_HARGA_NUM', 0), title

        if best_score >= score_threshold_primary and best_price > 0:
            return float(best_price)

        # 5) Fallback ke seluruh katalog jika perlu
        best_score2, best_price2 = best_score, best_price
        for _, row in katalog_df.iterrows():
            title = str(row['JUDUL_NORM'])
            score = fuzz.token_set_ratio(s_clean, title)
            if score > best_score2 or (score == best_score2 and len(title) > len(best_title)):
                best_score2, best_price2, best_title = score, row.get('KATALOG_HARGA_NUM', 0), title

        if best_score2 >= score_threshold_fallback and best_price2 > 0:
            return float(best_price2)

        return 0
    except Exception:
        return 0

def calculate_eksemplar(nama_produk, jumlah_terjual):
    """Menghitung jumlah eksemplar berdasarkan 'PAKET ISI X' atau 'SATUAN'."""
    try:
        nama_produk_upper = str(nama_produk).upper()
        
        # Cari "PAKET ISI [ANGKA]"
        paket_match = re.search(r'PAKET\s*ISI\s*(\d+)', nama_produk_upper)
        # Cari "SATUAN"
        satuan_match = 'SATUAN' in nama_produk_upper
        paket_khusus = re.search(r'PAKET WAKAF MURAH 50 PCS', nama_produk_upper)
        
        faktor = 1 # Default adalah 1
        
        if paket_match:
            # Jika ketemu "PAKET ISI 7", ambil angka 7
            faktor = int(paket_match.group(1))
        elif satuan_match:
            # Jika ketemu "SATUAN", faktornya 1
            faktor = 1
        elif paket_khusus:
            faktor = 50
        # else:
            # Jika tidak ada keduanya, faktor tetap 1 (dihitung satuan)
            
        return jumlah_terjual * faktor
    except Exception:
        return jumlah_terjual # Fallback jika ada error

def get_eksemplar_multiplier(nama_produk):
    if pd.isna(nama_produk): return 1
    nama_produk = str(nama_produk).upper()
        
    # Deteksi PAKET ISI X atau PAKET X atau ISI X
    match = re.search(r'(?:PAKET\s*ISI|PAKET|ISI)\s*(\d+)', nama_produk)
    if match:
        return int(match.group(1))
    # Jika ada kata SATUAN, anggap 1
    if 'SATUAN' in nama_produk:
        return 1
    return 1
    
def process_summary(rekap_df, iklan_final_df, katalog_df, harga_custom_tlj_df, store_type):
    """
    Fungsi untuk memproses sheet 'SUMMARY'.
    - Menggabungkan produk dari REKAP dan IKLAN.
    - Menggunakan logika harga beli baru.
    """
    rekap_copy = rekap_df.copy()
    rekap_copy['No. Pesanan'] = rekap_copy['No. Pesanan'].replace('', np.nan).ffill()

    # summary_df = rekap_copy.groupby(['Nama Produk', 'Harga Satuan'], as_index=False).agg({
    #     'Jumlah Terjual': 'sum', 
    #     # 'Harga Satuan': 'first', <-- Dihapus karena sudah jadi bagian key
    #     'Total Harga Produk': 'sum',
    #     'Voucher Ditanggung Penjual': 'sum', 'Biaya Komisi AMS + PPN Shopee': 'sum',
    #     'Biaya Adm 8%': 'sum', 'Biaya Layanan 2%': 'sum',
    #     'Biaya Layanan Gratis Ongkir Xtra 4,5%': 'sum', 'Biaya Proses Pesanan': 'sum',
    #     'Total Penghasilan': 'sum'
    # })
    # 1. Identifikasi baris retur (Harga Satuan == 0 DAN Total Penghasilan != 0)
    #    Kita gunakan Harga Satuan 0 sebagai penanda, karena REKAP sudah meng-nol-kannya.
    # kondisi_retur = (rekap_copy['Harga Satuan'] == 0) & (rekap_copy['Total Penghasilan'] != 0)
    
    # # 2. Buat Peta (Map) dari Nama Produk ke Harga Satuan Asli (non-nol)
    # #    Kita ambil baris non-retur, drop duplikat nama produk, dan buat dict
    # harga_asli_map = rekap_copy[~kondisi_retur].drop_duplicates(subset=['Nama Produk']) \
    #                                           .set_index('Nama Produk')['Harga Satuan']
    
    # # 3. Terapkan (map) harga asli ini ke kolom 'Harga Satuan' PADA BARIS RETUR
    # #    Ini "memaksa" baris retur agar memiliki Harga Satuan yang sama dengan baris aslinya
    # rekap_copy.loc[kondisi_retur, 'Harga Satuan'] = rekap_copy['Nama Produk'].map(harga_asli_map)
    
    # # 4. Ganti NaN (jika retur tapi tidak ada penjualan normal) kembali ke 0 agar groupby-nya tetap
    # rekap_copy['Harga Satuan'] = rekap_copy['Harga Satuan'].fillna(0)
    kondisi_retur_summary = rekap_copy['Total Penghasilan'] <= 0
    
    # Set 'Jumlah Terjual' ke 0 HANYA untuk baris retur
    # Ini terjadi di 'rekap_copy', jadi 'REKAP' asli tetap utuh
    rekap_copy.loc[kondisi_retur_summary, 'Jumlah Terjual'] = 0
    rekap_copy.loc[kondisi_retur_summary, 'Total Harga Produk'] = 0
    
    # --- ▲▲▲ AKHIR BLOK PERBAIKAN ▲▲▲ ---

    # Agregasi data utama dari REKAP
    # Sekarang groupby ini akan menggabungkan retur (yang Harga Satuannya sudah "diperbaiki")
    # dengan penjualan normal.
    biaya_layanan_col = 'Biaya Layanan 4,5%' if store_type == 'Pacific Bookstore' else 'Biaya Layanan 2%'
    summary_df = rekap_copy.groupby(['Nama Produk', 'Harga Satuan'], as_index=False).agg({
        'Jumlah Terjual': 'sum', 
        # 'Harga Satuan': 'first', <-- Dihapus karena sudah jadi bagian key
        'Total Harga Produk': 'sum',
        'Voucher Ditanggung Penjual': 'sum', 'Biaya Komisi AMS + PPN Shopee': 'sum',
        'Biaya Adm 8%': 'sum', biaya_layanan_col: 'sum',
        'Biaya Layanan Gratis Ongkir Xtra 4,5%': 'sum', 'Biaya Proses Pesanan': 'sum',
        'Total Penghasilan': 'sum' # Ini akan menjumlahkan (Penjualan Positif + Penjualan Negatif)
    })

    summary_df = summary_df[summary_df['Total Penghasilan'] != 0].copy()

    # --- LOGIKA BARU: Tambahkan Produk dari IKLAN yang tidak ada di REKAP ---
    # Siapkan kolom 'Iklan Klik' dengan nilai default 0
    summary_df['Iklan Klik'] = 0.0
    
    # Daftar produk khusus yang biaya iklannya perlu didistribusikan
    produk_khusus = [
        "CUSTOM AL QURAN MENGENANG/WAFAT 40/100/1000 HARI",
        "AL QUR'AN GOLD TERMURAH",
        "AL QUR'AN A6 NON TERJEMAH HVS WARNA PASTEL",
        "Alquran Cover Emas Kertas HVS Al Aqeel Gold Murah",
        "Al Qur'an Untuk Wakaf Al Aqeel A5 Kertas Koran 18 Baris",
        "AL-QUR'AN SAKU A7 MAHEER HAFALAN AL QUR'AN",
        "AL-QUR'AN TERJEMAH HC AL ALEEM A5",
        "AL-QURAN AL AQEEL SILVER TERMURAH",
        "AL QUR'AN NON TERJEMAH Al AQEEL A5 KERTAS KORAN WAKAF",
        "AL QUR'AN EDISI TAHLILAN 30 Juz + Doa Tahlil | Pengganti Buku Yasin | Al Aqeel A6 Pastel HVS Edisi Tahlilan",
        "TERBARU Al Quran Edisi Tahlilan Pengganti Buku Yasin Al Aqeel A6 Kertas HVS | SURABAYA | Mushaf Untuk Pengajian Kado Islami Hampers",
        "Al Quran Terjemah Al Aleem A5 HVS 15 Baris | SURABAYA | Alquran Untuk Pengajian Majelis Taklim",
        "Al Quran Saku Resleting Al Quddus A7 QPP Cover Kulit | SURABAYA | Untuk Santri Traveler Muslim",
        "Al Quran Wakaf Ibtida Al Quddus A5 Kertas HVS | Alquran SURABAYA",
        "Al Fikrah Al Quran Terjemah Fitur Lengkap A5 Kertas HVS | Alquran SURABAYA",
        "Al Quddus Al Quran Wakaf Ibtida A5 Kertas HVS | Alquran SURABAYA",
        "Al Quran Terjemah Al Aleem A5 Kertas HVS 15 Baris | SURABAYA | Alquran Untuk Majelis Taklim Kajian",
        "Al Quran Terjemah Per Kata A5 | Tajwid 2 Warna | Alquran Al Fikrah HVS 15 Baris | SURABAYA",
        "Al Quran Saku Resleting Al Quddus A7 Cover Kulit Kertas QPP | Alquran SURABAYA",
        "Al Quran Saku Pastel Al Aqeel A6 Kertas HVS | SURABAYA | Alquran Untuk Wakaf Hadiah Islami Hampers",
        "Al Quran Untuk Wakaf Al Aqeel A5 Kertas Koran 18 Baris | SURABAYA | Alquran Hadiah Islami Hampers",
        "Alquran Edisi Tahlilan Lebih Mulia Daripada Buku Yasin Biasa | Al Aqeel A6 Kertas HVS | SURABAYA |",
        "Paket Wakaf Murah 50 pcs Alquran Al Aqeel | Alquran 18 Baris",
        "PAKET MURAH ALQURAN AL AQEEL MUSHAF NON TERJEMAHAN | SURABAYA | al quran Wakaf/Shodaqoh hadiah hampers islami"
    ]
    produk_khusus = [re.sub(r'\s+', ' ', name.replace('\xa0', ' ')).strip() for name in produk_khusus]
    
    # # Ambil data iklan yang relevan
    iklan_data = iklan_final_df[iklan_final_df['Nama Iklan'] != 'TOTAL'][['Nama Iklan', 'Biaya']].copy()

    # # 1. Definisikan Nama Iklan dan target Nama Produk
    # nama_iklan_kustom = "Al Quran Saku Pastel Al Aqeel A6 Kertas HVS | SURABAYA | Alquran Untuk Wakaf Hadiah Islami Hampers"
    # # nama_iklan_kustom = "INDEX"
    # target_produk_kustom = [
    #     "Al Qur'an Saku Pastel Al Aqeel A6 Kertas HVS | Hadiah Islami, Cover Cantik",
    #     "Al Qur'an Pastel Al Aqeel A6 Kertas HVS | Wakaf, Hadiah Islami, Cover Cantik",
    #     "Alquran Edisi Tahlilan Lebih Mulia Daripada Buku Yasin Biasa | Al Aqeel A6 Kertas HVS | SURABAYA |"
    #     # Tambahkan nama produk target lainnya di sini jika ada
    # ]
    
    # # 2. Cek hanya jika ini Pacific Bookstore
    # if store_type == 'Pacific Bookstore':
    #     # 3. Cari biaya iklan kustom
    #     iklan_cost_row_kustom = iklan_data[iklan_data['Nama Iklan'] == nama_iklan_kustom]
        
    #     if not iklan_cost_row_kustom.empty:
    #         total_iklan_cost_kustom = iklan_cost_row_kustom['Biaya'].iloc[0]
            
    #         # 4. Cari baris summary yang cocok (gunakan .isin() untuk list)
    #         matching_summary_rows_kustom = summary_df['Nama Produk'].isin(target_produk_kustom)
            
    #         # 5. Hitung jumlah yang cocok
    #         num_variations_kustom = matching_summary_rows_kustom.sum()
            
    #         if num_variations_kustom > 0:
    #             # 6. Bagi dan alokasikan biaya
    #             distributed_cost_kustom = total_iklan_cost_kustom / num_variations_kustom
    #             summary_df.loc[matching_summary_rows_kustom, 'Iklan Klik'] = distributed_cost_kustom
                
    #             # 7. Hapus iklan ini dari 'iklan_data' agar tidak diproses lagi oleh loop di bawah
    #             iklan_data = iklan_data[iklan_data['Nama Iklan'] != nama_iklan_kustom]
    
    # Konfigurasi Produk Khusus dengan Variasi Wajib & Denominator
    force_config = {}
    if store_type == "Human Store":
        force_config = {
            "Alquran Cover Emas Kertas HVS Al Aqeel Gold Murah": {
                "variasi": ["A7 SATUAN", "A7 PAKET ISI 3", "A7 PAKET ISI 5", "A7 PAKET ISI 7", "A5 SATUAN", "A5 PAKET ISI 3"],
                "denom": 20
            },
            "AL QUR'AN NON TERJEMAH Al AQEEL A5 KERTAS KORAN WAKAF": {
                "variasi": ["SATUAN", "PAKET ISI 3", "PAKET ISI 5", "PAKET ISI 7"],
                "denom": 16
            }
        }
    elif store_type == "Pacific Bookstore":
        force_config = {
            # "Al Quran Saku Pastel Al Aqeel A6 Kertas HVS | SURABAYA | Alquran Untuk Wakaf Hadiah Islami Hampers": {
            #     "variasi": ["SATUAN", "PAKET ISI 3", "PAKET ISI 5", "PAKET ISI 7"],
            #     "denom": 16
            # },
            # "Al Quran Untuk Wakaf Al Aqeel A5 Kertas Koran 18 Baris | SURABAYA | Alquran Hadiah Islami Hampers": {
            #     "variasi": ["SATUAN", "PAKET ISI 3", "PAKET ISI 5", "PAKET ISI 7"],
            #     "denom": 16
            # },
            "Alquran GOLD Hard Cover Al Aqeel Kertas HVS | SURABAYA | Alquran untuk Pengajian Wakaf Hadiah Islami Hampers": {
                "variasi": ["A5 Gold Satuan", "A5 Gold Paket isi 3", "A7 Gold Satuan", "A7 Gold Paket isi 3", "A7 Gold Paket isi 5", "A7 Gold Paket isi 7"],
                "denom": 20
            }
        }

    # PROSES GENERASI BARIS & HITUNG IKLAN KHUSUS
    for produk_base, config in force_config.items():
        # Bersihkan nama produk di summary_df untuk matching yang akurat
        summary_df['Nama Produk Clean'] = summary_df['Nama Produk'].astype(str).str.replace(r'\s+', ' ', regex=True).str.strip()
        
        # Cari biaya di iklan_data
        matching_ads = iklan_data[iklan_data['Nama Iklan'].str.contains(produk_base, case=False, na=False, regex=False)]
        
        if not matching_ads.empty:
            total_biaya_iklan = matching_ads['Biaya'].sum()
            denom = config['denom']
            
            # 1. Pastikan SEMUA variasi wajib ada
            for var in config['variasi']:
                # Format pencarian: "Nama Produk (Variasi)"
                nama_lengkap_search = f"{produk_base} ({var})".replace('  ', ' ').strip()
                
                # Cek apakah sudah ada (case-insensitive & space-insensitive)
                exists = summary_df['Nama Produk Clean'].str.contains(re.escape(nama_lengkap_search), case=False, na=False).any()
                
                if not exists:
                    # Buat baris baru jika tidak ada
                    new_row = pd.DataFrame([{col: 0 for col in summary_df.columns}])
                    new_row['Nama Produk'] = f"{produk_base} ({var})"
                    summary_df = pd.concat([summary_df, new_row], ignore_index=True)
                    # Update Nama Produk Clean untuk iterasi selanjutnya
                    summary_df['Nama Produk Clean'] = summary_df['Nama Produk'].astype(str).str.replace(r'\s+', ' ', regex=True).str.strip()

            # 2. Hitung Iklan Klik untuk semua baris yang mengandung produk_base ini
            mask_summary = summary_df['Nama Produk'].str.contains(produk_base, case=False, na=False, regex=False)
            indices = summary_df[mask_summary].index
            
            for idx in indices:
                p_name = summary_df.at[idx, 'Nama Produk']
                # Hitung jumlah baris yang memiliki Nama Produk yang SAMA PERSIS (untuk pembagi)
                count_same = (summary_df['Nama Produk'] == p_name).sum()
                mult = get_eksemplar_multiplier(p_name)
                
                # Rumus: (Multiplier * Biaya) / Denom / Count
                summary_df.at[idx, 'Iklan Klik'] = (mult * total_biaya_iklan) / denom / count_same
            
            # Hapus dari iklan_data agar tidak terproses logika standar di bawah
            iklan_data = iklan_data[~iklan_data['Nama Iklan'].str.contains(produk_base, case=False, na=False, regex=False)]
    summary_df.drop(columns=['Nama Produk Clean'], inplace=True, errors='ignore')

    # LOGIKA STANDAR UNTUK PRODUK KHUSUS LAINNYA (TANPA GENERATE VARIASI)
    produk_khusus_biasa = [
        "Paket Alquran Khusus Wakaf Al Aqeel A5 Kertas Koran",
        "AL QUR'AN A6 NON TERJEMAH HVS WARNA PASTEL",
        "Alquran Edisi Tahlilan Lebih Mulia Daripada Buku Yasin Biasa",
        "Al Quran Saku Pastel Al Aqeel A6 Kertas HVS | SURABAYA | Alquran Untuk Wakaf Hadiah Islami Hampers",
        "Al Quran Untuk Wakaf Al Aqeel A5 Kertas Koran 18 Baris | SURABAYA | Alquran Hadiah Islami Hampers",
        "Paket Wakaf Murah 50 pcs Alquran Al Aqeel | Alquran 18 Baris",
        "PAKET MURAH ALQURAN AL AQEEL MUSHAF NON TERJEMAHAN | SURABAYA | al quran Wakaf/Shodaqoh hadiah hampers islami",
        "Alquran Edisi Tahlilan Lebih Mulia Daripada Buku Yasin Biasa | Al Aqeel A6 Kertas HVS | SURABAYA |"
    ]
    
    for p_biasa in produk_khusus_biasa:
        matching_ads = iklan_data[iklan_data['Nama Iklan'].str.contains(p_biasa, case=False, na=False, regex=False)]
        if not matching_ads.empty:
            total_biaya = matching_ads['Biaya'].sum()
            mask_summary = summary_df['Nama Produk'].str.contains(p_biasa, case=False, na=False, regex=False)
            num_rows = mask_summary.sum()
            if num_rows > 0:
                summary_df.loc[mask_summary, 'Iklan Klik'] = total_biaya / num_rows
            else:
                # --- PERBAIKAN DI SINI ---
                # Jika 0 penjualan, buat baris baru agar biaya iklan tetap muncul di Summary
                new_row_ads = pd.DataFrame([{col: 0 for col in summary_df.columns}])
                new_row_ads['Nama Produk'] = p_biasa
                new_row_ads['Iklan Klik'] = total_biaya
                summary_df = pd.concat([summary_df, new_row_ads], ignore_index=True)
            iklan_data = iklan_data[~iklan_data['Nama Iklan'].str.contains(p_biasa, case=False, na=False, regex=False)]
    
    # 2. Proses Produk Normal (yang tersisa di iklan_data)
    # Gunakan merge untuk produk yang namanya cocok persis
    summary_df = pd.merge(summary_df, iklan_data, left_on='Nama Produk', right_on='Nama Iklan', how='left')
    
    # Gabungkan hasil merge dengan kolom 'Iklan Klik' yang sudah ada
    # `summary_df['Biaya']` akan berisi biaya untuk produk normal
    summary_df['Iklan Klik'] = summary_df['Iklan Klik'] + summary_df['Biaya'].fillna(0)
    summary_df.drop(columns=['Nama Iklan', 'Biaya'], inplace=True, errors='ignore')
    
    # 3. Tambahkan Produk yang Hanya Ada di IKLAN (dan bukan produk khusus)
    iklan_only_names = set(iklan_data['Nama Iklan']) - set(summary_df['Nama Produk'])
    if iklan_only_names:
        iklan_only_df = iklan_data[iklan_data['Nama Iklan'].isin(iklan_only_names)].copy()
        iklan_only_df.rename(columns={'Nama Iklan': 'Nama Produk', 'Biaya': 'Iklan Klik'}, inplace=True)
        summary_df = pd.concat([summary_df, iklan_only_df], ignore_index=True)
    
    # Pastikan semua nilai NaN di kolom numerik utama menjadi 0
    summary_df.fillna(0, inplace=True)
    # --- AKHIR LOGIKA BARU ---

    # Sisa fungsi sama seperti sebelumnya, dengan penyesuaian pada pemanggilan `get_harga_beli_fuzzy`
    # summary_df['Penjualan Netto'] = (
    #     summary_df['Total Harga Produk'] - summary_df['Voucher Ditanggung Penjual'] -
    #     summary_df['Biaya Komisi AMS + PPN Shopee'] - summary_df['Biaya Adm 8%'] -
    #     summary_df['Biaya Layanan 2%'] - summary_df['Biaya Layanan Gratis Ongkir Xtra 4,5%'] -
    #     summary_df['Biaya Proses Pesanan']
    # )
    # summary_df['Penjualan Netto'] = summary_df['Total Penghasilan']

    if store_type in ['Pacific Bookstore']:
        summary_df['Penjualan Netto'] = (
            summary_df['Total Harga Produk'] - summary_df['Voucher Ditanggung Penjual'] -
            summary_df['Biaya Komisi AMS + PPN Shopee'] - summary_df['Biaya Adm 8%'] -
            summary_df['Biaya Layanan 4,5%'] - summary_df['Biaya Layanan Gratis Ongkir Xtra 4,5%'] -
            summary_df['Biaya Proses Pesanan']
        )
    else:
        summary_df['Penjualan Netto'] = summary_df['Total Penghasilan']
        
    summary_df['Biaya Packing'] = summary_df['Jumlah Terjual'] * 200

    summary_df['Jumlah Eksemplar'] = summary_df.apply(
        lambda row: calculate_eksemplar(row['Nama Produk'], row['Jumlah Terjual']), 
        axis=1
    )

    if store_type in ['Pacific Bookstore']:
        # summary_df['Biaya Kirim ke Sby'] = summary_df['Jumlah Terjual'] * 733
        summary_df['Biaya Kirim ke Sby'] = 0
        biaya_ekspedisi_final = summary_df['Biaya Kirim ke Sby']
    else:
        summary_df['Biaya Ekspedisi'] = 0
        biaya_ekspedisi_final = summary_df['Biaya Ekspedisi']

    # --- PERUBAHAN PADA PEMANGGILAN FUNGSI ---
    # Pastikan rekap_df (rekap_copy) yang belum diagregasi digunakan untuk lookup variasi
    summary_df['Harga Beli'] = summary_df['Nama Produk'].apply(
        lambda x: get_harga_beli_fuzzy(x, katalog_df)
    )

    # --- LOGIKA BARU UNTUK HARGA CUSTOM TLJ ---
    # 1. Buat 'temp_lookup_key' yang formatnya SAMA DENGAN 'LOOKUP_KEY' di file Excel
    #    Caranya: ganti ' (' menjadi ' ' dan hapus ')'
    summary_df['temp_lookup_key'] = summary_df['Nama Produk'].astype(str).str.replace(' (', ' ', regex=False).str.replace(')', '', regex=False).str.strip()
    
    # 2. Gabungkan dengan data harga custom menggunakan 'temp_lookup_key'
    summary_df = pd.merge(
        summary_df,
        harga_custom_tlj_df[['LOOKUP_KEY', 'HARGA CUSTOM TLJ']],
        left_on='temp_lookup_key', # <-- Mencocokkan dengan 'CUSTOM... AL AQEEL A6 HVS...'
        right_on='LOOKUP_KEY',
        how='left'
    )
    
    # 3. Ganti nama kolom dan isi nilai kosong dengan 0
    summary_df.rename(columns={'HARGA CUSTOM TLJ': 'Harga Custom TLJ'}, inplace=True)
    summary_df['Harga Custom TLJ'] = summary_df['Harga Custom TLJ'].fillna(0)
    
    # 4. Hapus kolom-kolom sementara
    summary_df.drop(columns=['LOOKUP_KEY', 'temp_lookup_key'], inplace=True, errors='ignore')

    # --- LOGIKA BARU UNTUK TOTAL PEMBELIAN ---
    produk_custom_list = ["CUSTOM AL QURAN MENGENANG/WAFAT 40/100/1000 HARI", "AL QUR'AN EDISI TAHLILAN 30 Juz + Doa Tahlil | Pengganti Buku Yasin | Al Aqeel A6 Pastel HVS Edisi Tahlilan (Custom sisipan 1 hal)", 
                         "AL QUR'AN EDISI TAHLILAN 30 Juz + Doa Tahlil | Pengganti Buku Yasin | Al Aqeel A6 Pastel HVS Edisi Tahlilan (Custom sisipan 2 hal)", "AL QUR'AN EDISI TAHLILAN 30 Juz + Doa Tahlil | Pengganti Buku Yasin | Al Aqeel A6 Pastel HVS Edisi Tahlilan (Custom jacket)", 
                         "AL QUR'AN EDISI TAHLILAN 30 Juz + Doa Tahlil | Pengganti Buku Yasin | Al Aqeel A6 Pastel HVS Edisi Tahlilan (Custom case)", "AL QUR'AN EDISI TAHLILAN 30 Juz + Doa Tahlil | Pengganti Buku Yasin | Al Aqeel A6 Pastel HVS Edisi Tahlilan (Sisipan 1hal+jaket)"]
    
    # Ubah list menjadi satu string regex, pisahkan dengan '|' (OR)
    # Kita gunakan re.escape() untuk memastikan karakter '|' di dalam string tahlilan tidak merusak regex
    produk_custom_regex = '|'.join(re.escape(s) for s in produk_custom_list)

    # Kondisi: jika Nama Produk mengandung string produk custom
    kondisi_custom = summary_df['Nama Produk'].str.contains(produk_custom_regex, na=False)
    
    # Hitung Total Pembelian dengan rumus berbeda jika kondisi terpenuhi
    summary_df['Total Pembelian'] = np.where(
        kondisi_custom,
        (summary_df['Jumlah Terjual'] * summary_df['Harga Beli']) + (summary_df['Jumlah Terjual'] * summary_df['Harga Custom TLJ']), # Rumus untuk produk custom
        summary_df['Jumlah Terjual'] * summary_df['Harga Beli']  # Rumus untuk produk normal
    )
    
    summary_df['Margin'] = (
        summary_df['Penjualan Netto'] - summary_df['Iklan Klik'] - summary_df['Biaya Packing'] - 
        biaya_ekspedisi_final - summary_df['Total Pembelian']
    )
    
    # ... (sisa fungsi `process_summary` Anda tetap sama persis dari sini sampai akhir) ...
    summary_df['Persentase'] = (summary_df.apply(lambda row: row['Margin'] / row['Total Harga Produk'] if row['Total Harga Produk'] != 0 else 0, axis=1))
    summary_df['Jumlah Pesanan'] = summary_df.apply(lambda row: row['Biaya Proses Pesanan'] / 1250 if 1250 != 0 else 0, axis=1)
    summary_df['Penjualan Per Hari'] = round(summary_df['Total Harga Produk'] / 7, 1)
    summary_df['Jumlah buku per pesanan'] = round(summary_df.apply(lambda row: row['Jumlah Eksemplar'] / row['Jumlah Pesanan'] if row.get('Jumlah Pesanan', 0) != 0 else 0, axis=1), 1)
    
    summary_final_data = {
        'No': np.arange(1, len(summary_df) + 1), 'Nama Produk': summary_df['Nama Produk'],
        'Jumlah Terjual': summary_df['Jumlah Terjual'], 'Jumlah Eksemplar': summary_df['Jumlah Eksemplar'], 
        'Jumlah Pesanan': summary_df['Jumlah Pesanan'], 'Harga Satuan': summary_df['Harga Satuan'],
        'Total Penjualan': summary_df['Total Harga Produk'], 'Voucher Ditanggung Penjual': summary_df['Voucher Ditanggung Penjual'],
        'Biaya Komisi AMS + PPN Shopee': summary_df['Biaya Komisi AMS + PPN Shopee'], 'Biaya Adm 8%': summary_df['Biaya Adm 8%'],
        biaya_layanan_col: summary_df[biaya_layanan_col], 'Biaya Layanan Gratis Ongkir Xtra 4,5%': summary_df['Biaya Layanan Gratis Ongkir Xtra 4,5%'],
        'Biaya Proses Pesanan': summary_df['Biaya Proses Pesanan'],
        'Penjualan Netto': summary_df['Penjualan Netto'], 'Iklan Klik': summary_df['Iklan Klik'], 'Biaya Packing': summary_df['Biaya Packing'],
    }
    if store_type in ['Pacific Bookstore']:
        # summary_final_data['Biaya Kirim ke Sby'] = biaya_ekspedisi_final
        summary_final_data['Biaya Ekspedisi'] = biaya_ekspedisi_final
    else:
        summary_final_data['Biaya Ekspedisi'] = biaya_ekspedisi_final
    summary_final_data.update({
        'Harga Beli': summary_df['Harga Beli'], 'Harga Custom TLJ': summary_df['Harga Custom TLJ'],
        'Total Pembelian': summary_df['Total Pembelian'], 'Margin': summary_df['Margin'],
        'Persentase': summary_df['Persentase'],
        'Penjualan Per Hari': summary_df['Penjualan Per Hari'], 'Jumlah buku per pesanan': summary_df['Jumlah buku per pesanan']
    })
    summary_final = pd.DataFrame(summary_final_data)

    # --- LOGIKA PERSINGKAT NAMA PRODUK (KHUSUS HUMAN STORE) ---
    mapping_singkatan = {}
    if store_type == "Human Store":
        mapping_singkatan = {
            "AL-QUR'AN TERJEMAH HC AL ALEEM QPP A6": "Al Aleem A6 QPP",
            "AL-QUR'AN TERJEMAH  HC AL ALEEM QPP A6": "Al Aleem A6 QPP",
            "AL-QURAN AL AQEEL SILVER TERMURAH": "Al Aqeel Silver",
            "Paket Wakaf Murah 50 pcs Alquran Al Aqeel | Alquran 18 Baris": "Paket Wakaf Murah Al Aqeel 50 pcs",
            "AL QUR'AN WAQF IBTIDA | AL QUDDUS A5 KERTAS HVS": "Al Quddus A5 HVS",
            "AL QUR'AN AL AQEEL B5 KERTAS HVS": "Al Aqeel B5 HVS",
            "KAMUS BERGAMBAR 3 BAHASA - INDONESIA INGGRIS ARAB": "Kamus Bergambar 3 Bahasa",
            "AL QUR'AN NON TERJEMAH Al AQEEL A5 KERTAS KORAN WAKAF": "AL AQEEL A5 KORAN",
            "Paket Alquran Khusus Wakaf Al Aqeel A5 Kertas Koran | Alquran Murah Kualitas Terbaik Harga Ekonomis | Jakarta": "Al Aqeel A5 Koran",
            "Al QUR'AN NON TERJEMAH AL AQEEL KERTAS KORAN B5 WAKAF": "Al Aqeel B5 Koran",
            "Alquran Cover Emas Kertas HVS Al Aqeel Gold Murah": "Al Aqeel Gold",
            "AL-QUR'AN TERJEMAH HC AL ALEEM A5": "Al Aleem A5",
            "Komik Pahlawan, Pendidikan Sejarah Untuk Anak": "Komik Pahlawan",
            "AL QUR'AN AL FIKRAH TERJEMAH PER AYAT PER KATA A4 KERTAS HVS": "Al Fikrah A4 HVS",
            "AL QUR'AN HAFALAN SAKU A7 MAHEER KERTAS QPP": "A7 Maheer QPP",
            "AL QUR'AN B5 NON TERJEMAH HVS WARNA PASTEL": "Al Aqeel B5 Pastel",
            "AL QURAN SAKU RESLETING A7 AL QUDDUS KERTAS QPP": "Al Quddus A7 Saku QPP",
            "BUKU CERITA ANAK FABEL SERI DONGENG BINATANG DUA BAHASA": "Fabel Binatang",
            "BUKU CERITA KISAH TELADAN NABI SERI VOL 1-6": "Kisah Teladan Nabi",
            "AL- QUR'AN TAJWID WARNA WAQF IBTIDA | SUBHAAN A5 KERTAS QPP": "Subhaan A5 QPP",
            "BUKU LAGU HARMONI NUSANTARA LAGU NASIONAL & DAERAH": "Buku Lagu Harmoni Nusantara",
            "[KOLEKSI TERBARU] SERI CERITA RAKYAT": "Seri Cerita Rakyat",
            "[KOLEKSI TERBARU] BUKU CERITA ANAK SERI BUDI PEKERTI": "Seri Budi Pekerti",
            "AL- QUR'AN TERJEMAH TAJWID MUMTAAZ A5 KERTAS QPP": "Mumtaaz A5 QPP",
            "AL QUR'AN A6 NON TERJEMAH HVS WARNA PASTEL": "Al Aqeel 6 Pastel",
            "Custom Al Quran Mengenang/Wafat 40/100/1000 Hari": "Alquran Custom",
            "AL QUR'AN EDISI TAHLILAN 30 Juz + Doa Tahlil | Pengganti Buku Yasin | Al Aqeel A6 Pastel HVS Edisi Tahlilan": "A6 edisi Tahlilan",
            "Al-Qur'an Non Terjemah Al Aqeel HVS A5": "Al Aqeel A5 HVS",
            "Al Qur'an Terjemah Per Kata | Tajwid 2 Warna | Al Fikrah A5 Kertas HVS": "Al Fikrah A5 HVS"
        }
    elif store_type == "Pacific Bookstore":
        mapping_singkatan = {
            "Alquran Custom Nama Foto | SURABAYA | Al-Quran untuk Wakaf Tasyakuran Tahlil Yasin Hadiah Hampers Islami": "Alquran Custom Al Aqeel",
            "PAKET MURAH ALQURAN AL AQEEL MUSHAF NON TERJEMAHAN | SURABAYA | al quran Wakaf/Shodaqoh hadiah hampers islami": "PAKET MURAH AL AQEEL MIN 10 EKS",
            "Al Quran Terjemah Per Kata A5 | Tajwid 2 Warna | Alquran Al Fikrah HVS 15 Baris | SURABAYA": "Al Fikrah A5 HVS",
            "Alquran GOLD Hard Cover Al Aqeel Kertas HVS | SURABAYA | Alquran untuk Pengajian Wakaf Hadiah Islami Hampers": "Al Aqeel Gold Kertas HVS",
            "Al Quran Untuk Wakaf Al Aqeel A5 Kertas Koran 18 Baris | SURABAYA | Alquran Hadiah Islami Hampers": "Al Aqeel A5 Kertas Koran",
            "Al Quran Saku Pastel Al Aqeel A6 Kertas HVS | SURABAYA | Alquran Untuk Wakaf Hadiah Islami Hampers": "Al Aqeel A6 Kertas HVS",
            "Alquran Edisi Tahlilan Lebih Mulia Daripada Buku Yasin Biasa | Al Aqeel A6 Kertas HVS | SURABAYA |": "Al Aqeel A6 Edisi Tahlilan Kertas HVS",
            "Alquran Edisi Tahlilan Lebih Mulia Daripada Buku Yasin Biasa": "Al Aqeel A6 Edisi Tahlilan Kertas HVS",
            "Al Quran Saku Resleting Al Quddus A7 Cover Kulit Kertas QPP | Alquran SURABAYA": "Al Quddus A7 Cover Kulit Kertas QPP",
            "Al Quran Saku Resleting Al Quddus A7 QPP Cover Kulit | SURABAYA | Untuk Santri Traveler Muslim": "Al Quddus A7 Cover Kulit Kertas QPP",
            "Al Quran Terjemah Al Aleem A5 Kertas HVS 15 Baris | SURABAYA | Alquran Untuk Majelis Taklim Kajian": "Al Aleem A5 Kertas HVS",
            "Al Quran Wakaf Ibtida Al Quddus A5 Kertas HVS | Alquran SURABAYA": "Al Quddus Ibtida A5 Kertas HVS"
        }

        # def apply_shorten(nama_full):
        #     if pd.isna(nama_full): return nama_full
            
        #     # Pisahkan nama produk dan variasi (teks di dalam kurung)
        #     # Regex ini mencari bagian dalam kurung terakhir
        #     match_variasi = re.search(r'(\s*\(.*\))$', nama_full)
        #     variasi_part = match_variasi.group(1) if match_variasi else ""
        #     nama_produk_saja = nama_full.replace(variasi_part, "").strip()

        #     # Cek apakah nama produk mengandung salah satu keyword di mapping
        #     for original_name, short_name in mapping_singkatan.items():
        #         if original_name.lower() in nama_produk_saja.lower():
        #             # Gabungkan Nama Singkat dengan Variasi aslinya
        #             return f"{short_name}{variasi_part}"
            
        #     return nama_full
    # Jika ada mapping yang terisi (Human/Pacific), jalankan fungsinya
    if mapping_singkatan:
        def apply_shorten(nama_full):
            if pd.isna(nama_full): return nama_full
            # Deteksi variasi di dalam kurung terakhir
            match_variasi = re.search(r'(\s*\(.*\))$', nama_full)
            variasi_part = match_variasi.group(1) if match_variasi else ""
            nama_produk_saja = nama_full.replace(variasi_part, "").strip()

            for original_name, short_name in mapping_singkatan.items():
                if original_name.lower() in nama_produk_saja.lower():
                    return f"{short_name}{variasi_part}"
            return nama_full

        summary_final['Nama Produk'] = summary_final['Nama Produk'].apply(apply_shorten)
    # Terapkan ke kolom Nama Produk
    # summary_final['Nama Produk'] = summary_final['Nama Produk'].apply(apply_shorten)
        
    summary_final = summary_final.sort_values(by='Nama Produk', ascending=True).reset_index(drop=True)
    summary_final['No'] = range(1, len(summary_final) + 1)
    
    total_row = pd.DataFrame(summary_final.sum(numeric_only=True)).T
    total_row['Nama Produk'] = 'Total'
    total_penjualan_netto = total_row['Penjualan Netto'].iloc[0]
    total_iklan_klik = total_row['Iklan Klik'].iloc[0]
    total_biaya_packing = total_row['Biaya Packing'].iloc[0]
    total_pembelian = total_row['Total Pembelian'].iloc[0]
    total_harga_produk = total_row['Total Penjualan'].iloc[0]
    total_biaya_proses_pesanan = total_row['Biaya Proses Pesanan'].iloc[0]
    total_jumlah_terjual = total_row['Jumlah Terjual'].iloc[0]
    total_jumlah_eksemplar = total_row['Jumlah Eksemplar'].iloc[0] # <-- DITAMBAH
    biaya_ekspedisi_col_name = 'Biaya Ekspedisi' if store_type == 'Pacific Bookstore' else 'Biaya Ekspedisi'
    total_biaya_ekspedisi = total_row[biaya_ekspedisi_col_name].iloc[0]
    total_margin = total_penjualan_netto - total_biaya_packing - total_biaya_ekspedisi - total_pembelian - total_iklan_klik
    total_row['Margin'] = total_margin
    total_row['Persentase'] = (total_margin / total_harga_produk) if total_harga_produk != 0 else 0
    total_jumlah_pesanan = (total_biaya_proses_pesanan / 1250) if 1250 != 0 else 0
    total_row['Jumlah Pesanan'] = total_jumlah_pesanan
    total_row['Penjualan Per Hari'] = round(total_harga_produk / 7, 1)
    total_row['Jumlah buku per pesanan'] = round(total_jumlah_eksemplar / total_jumlah_pesanan if total_jumlah_pesanan != 0 else 0, 1) # <-- DIUBAH
    for col in ['Harga Satuan', 'Harga Beli', 'No', 'Harga Custom TLJ']:
        if col in total_row.columns: total_row[col] = None
    summary_with_total = pd.concat([summary_final, total_row], ignore_index=True)
    
    return summary_with_total

def format_variation_dama(variation, product_name):
    """
    Format variasi untuk DAMA.ID STORE SUMMARY.
    Hanya mempertahankan warna JIKA produk adalah Hijab/Pashmina.
    Mengabaikan variasi '0'.
    Mempertahankan jenis kertas, ukuran, paket.
    """
    if pd.isna(variation):
        return ''

    var_str = str(variation).strip()
    # Abaikan jika variasi hanya '0'
    if var_str == '0':
        return ''

    product_name_upper = str(product_name).upper()

    # Keywords warna (lowercase)
    color_keywords = {'merah', 'biru', 'hijau', 'kuning', 'hitam', 'putih', 'ungu', 'coklat', 'cokelat', # Tambah 'cokelat'
                      'abu', 'pink', 'gold', 'silver', 'cream', 'navy', 'maroon', 'random',
                      'army', 'olive', 'mocca', 'dusty', 'sage'}
    # Keywords produk yang warnanya dipertahankan
    hijab_keywords = {'HIJAB', 'PASHMINA', 'PASMINA'}
    # Keywords/patterns lain yang selalu dipertahankan
    keep_keywords = {'HVS', 'QPP', 'KORAN', 'KK', 'KWARTO', 'BIGBOS', 'ART PAPER'}
    keep_patterns = [r'\b(PAKET\s*\d+)\b', r'\b((A|B)\d{1,2})\b']

    # Cek apakah warna perlu dipertahankan
    keep_color = any(keyword in product_name_upper for keyword in hijab_keywords)

    parts = re.split(r'[\s,]+', var_str) # Pisahkan berdasarkan spasi atau koma
    final_parts = []

    for part in parts:
        part_upper = part.upper()
        part_lower = part.lower()

        # Lewati jika kosong atau hanya '0'
        if not part or part == '0':
            continue

        # Cek apakah bagian ini adalah warna
        is_color = part_lower in color_keywords

        # Logika: Pertahankan bagian jika...
        # 1. BUKAN warna, ATAU
        # 2. ADALAH warna DAN keep_color=True
        if not is_color or (is_color and keep_color):
            # Cek juga apakah cocok dengan pola/keyword yang selalu disimpan
            is_kept_keyword = part_upper in keep_keywords
            is_kept_pattern = any(re.fullmatch(pattern, part_upper) for pattern in keep_patterns)

            # Jika bukan warna, atau warna yang dipertahankan, atau keyword/pola lain
            if not is_color or keep_color or is_kept_keyword or is_kept_pattern:
                 # Map KK ke KORAN
                 final_parts.append('KORAN' if part_upper == 'KK' else part)

    # Hilangkan duplikat sambil mempertahankan urutan (jika perlu, tapi set lebih mudah)
    # Urutkan untuk konsistensi
    unique_parts_ordered = list(dict.fromkeys(final_parts))

    return ' '.join(unique_parts_ordered)

def get_harga_beli_dama(summary_product_name, katalog_dama_df, score_threshold_primary=80, score_threshold_fallback=75):
    """
    Mencari harga beli dari KATALOG_DAMA dengan logika 2-pass (ketat lalu longgar).
    Pass 1: Fuzzy match nama (>=80) DAN atribut (jenis, ukuran, paket) harus cocok.
    Pass 2 (Fallback): Jika Pass 1 gagal, cari fuzzy match nama (>=75) saja.
    """
    try:
        if pd.isna(summary_product_name) or not summary_product_name.strip():
            return 0

        # 1. Parse Nama Produk Summary
        base_name = summary_product_name.strip()
        variasi_part = ''
        match = re.match(r'^(.*?)\s*\((.*?)\)$', summary_product_name.strip())
        if match:
            base_name = match.group(1).strip()
            variasi_part = match.group(2).strip().upper()

        base_name_upper_clean = re.sub(r'\s+', ' ', base_name.upper()).strip()

        # 2. Ekstrak Atribut dari Variasi Part
        ukuran_in_var = ''
        jenis_in_var = ''
        paket_in_var = ''

        size_match = re.search(r'\b((A|B)\d{1,2})\b', variasi_part)
        if size_match: ukuran_in_var = size_match.group(1)

        paper_keywords = {'HVS', 'QPP', 'KORAN', 'KK', 'KWARTO', 'BIGBOS', 'ART PAPER'}
        variasi_words = set(re.split(r'\s+', variasi_part))
        for paper in paper_keywords:
            if paper in variasi_words:
                jenis_in_var = 'KORAN' if paper == 'KK' else paper
                break

        package_match = re.search(r'\b(PAKET\s*\d+)\b', variasi_part)
        if package_match: 
            # Bersihkan spasi agar "PAKET 10" menjadi "PAKET10" untuk pencocokan
            # Normalisasi spasi, misal "PAKET  10" atau "PAKET 10" menjadi "PAKET 10"
            paket_in_var = re.sub(r'\s+', ' ', package_match.group(1)).strip()
        
        # --- ▼▼▼ TAMBAHKAN BLOK INI ▼▼▼ ---
        warna_in_var = ''
        color_keywords_set = {'MERAH', 'BIRU', 'HIJAU', 'KUNING', 'HITAM', 'PUTIH', 'UNGU', 'COKLAT', 'COKELAT',
                              'ABU', 'PINK', 'GOLD', 'SILVER', 'CREAM', 'NAVY', 'MAROON', 'RANDOM',
                              'ARMY', 'OLIVE', 'MOCCA', 'DUSTY', 'SAGE'}
        # variasi_words sudah didefinisikan di atas (saat cek paper_keywords)
        found_colors = variasi_words.intersection(color_keywords_set)
        if found_colors:
            warna_in_var = list(found_colors)[0] # Ambil warna pertama yang ditemukan
        
        # Tentukan apakah pencocokan warna diperlukan
        hijab_keywords = {'PASHMINA', 'HIJAB', 'PASMINA'}
        match_warna_required = any(keyword in base_name_upper_clean for keyword in hijab_keywords)

        
        # --- Inisialisasi untuk 2-Pass ---
        best_strict_score = -1
        best_strict_price = 0
        
        best_fallback_score = -1
        best_fallback_price = 0

        # 3. Iterasi Katalog Dama
        for index, row in katalog_dama_df.iterrows():
            katalog_name = row['NAMA PRODUK']
            katalog_jenis = row['JENIS AL QUR\'AN']
            katalog_ukuran = row['UKURAN']
            # Bersihkan spasi di data katalog juga untuk pencocokan yang adil
            katalog_paket = row['PAKET']
            katalog_warna = row['WARNA']
            
            # Hitung Skor Nama (Fuzzy Match)
            name_score = fuzz.token_set_ratio(base_name_upper_clean, katalog_name)

            # --- Pass 1: Cek Logika Ketat (Primary Threshold) ---
            if name_score >= score_threshold_primary:
                match_ok = True # Asumsikan cocok

                # Cek Atribut: Hanya filter JIKA atribut ada di variasi
                if jenis_in_var and katalog_jenis != jenis_in_var:
                    match_ok = False
                if ukuran_in_var and katalog_ukuran != ukuran_in_var:
                    match_ok = False
                
                # Logika Paket: Harus sama persis, atau keduanya kosong
                if paket_in_var != katalog_paket:
                    match_ok = False

                if match_warna_required:
                    # Ini adalah produk HIJAB/PASHMINA.
                    # Pengecekan ketat: Warna di variasi (warna_in_var) HARUS SAMA
                    # dengan warna di katalog (katalog_warna).
                    if katalog_warna != warna_in_var:
                        match_ok = False

                # Jika semua cek atribut lolos
                if match_ok:
                    if name_score > best_strict_score:
                        best_strict_score = name_score
                        best_strict_price = row['HARGA']

            # --- Pass 2: Simpan Skor Fallback ---
            if name_score >= score_threshold_fallback:
                if name_score > best_fallback_score:
                    best_fallback_score = name_score
                    best_fallback_price = row['HARGA']

        # 4. Kembalikan Hasil
        # Prioritaskan hasil dari Pass 1 (strict)
        if best_strict_score != -1: # Jika ada kecocokan ketat
            return best_strict_price
        
        # Jika tidak ada kecocokan ketat, gunakan hasil Pass 2 (fallback)
        if best_fallback_score != -1:
            return best_fallback_price

        # Jika tidak ada yang cocok sama sekali
        return 0

    except Exception as e:
        # st.error(f"Error di get_harga_beli_dama for '{summary_product_name}': {e}")
        return 0

def get_eksemplar_multiplier_dama(nama_produk):
    if pd.isna(nama_produk): return 1
    nama_produk = str(nama_produk).upper()
    # Khusus Dama: B5 (Bigbos) dihitung 1
    if 'BIGBOS' in nama_produk:
        return 1
    match = re.search(r'(?:PAKET\s*ISI|PAKET|ISI)\s*(\d+)', nama_produk)
    if match:
        return int(match.group(1))
    if 'SATUAN' in nama_produk:
        return 1
    return 1
    
# --- TAMBAHKAN FUNGSI BARU INI ---
def process_summary_dama(rekap_df, iklan_final_df, katalog_dama_df, harga_custom_tlj_df): # Tambah katalog_dama_df
    """
    Fungsi untuk memproses sheet 'SUMMARY' KHUSUS untuk DAMA.ID STORE (Shopee).
    Menggabungkan Nama Produk + Variasi Relevan (tanpa warna kecuali Hijab).
    Menggunakan KATALOG_DAMA untuk Harga Beli.
    """
    rekap_copy = rekap_df.copy()
    rekap_copy['No. Pesanan'] = rekap_copy['No. Pesanan'].replace('', np.nan).ffill()

    # --- ▼▼▼ BLOK PERBAIKAN RETUR SUMMARY (DAMA) ▼▼▼ ---
    # 1. Identifikasi baris retur (Harga Satuan == 0 DAN Total Penghasilan != 0)
    #    Logika retur Dama/Pacific juga meng-nol-kan Harga Satuan, jadi ini aman.
    # kondisi_retur = (rekap_copy['Harga Satuan'] == 0) & (rekap_copy['Total Penghasilan'] != 0)
    
    # # 2. Buat Peta (Map) dari Nama Produk ke Harga Satuan Asli (non-nol)
    # #    PENTING: Gunakan 'Nama Produk' (yang masih original) untuk membuat peta
    # harga_asli_map = rekap_copy[~kondisi_retur].drop_duplicates(subset=['Nama Produk']) \
    #                                           .set_index('Nama Produk')['Harga Satuan']
    
    # # 3. Terapkan (map) harga asli ini ke kolom 'Harga Satuan' PADA BARIS RETUR
    # rekap_copy.loc[kondisi_retur, 'Harga Satuan'] = rekap_copy['Nama Produk'].map(harga_asli_map)
    # rekap_copy['Harga Satuan'] = rekap_copy['Harga Satuan'].fillna(0)
    kondisi_retur_summary = rekap_copy['Total Penghasilan'] <= 0
    
    # Set 'Jumlah Terjual' ke 0 HANYA untuk baris retur
    # Ini terjadi di 'rekap_copy', jadi 'REKAP' asli tetap utuh
    rekap_copy.loc[kondisi_retur_summary, 'Jumlah Terjual'] = 0
    rekap_copy.loc[kondisi_retur_summary, 'Total Harga Produk'] = 0
    # --- ▲▲▲ AKHIR BLOK PERBAIKAN ▲▲▲ ---

    # --- LOGIKA BARU PEMBUATAN NAMA PRODUK DISPLAY ---
    rekap_copy['Nama Produk Original'] = rekap_copy['Nama Produk']
    if 'Nama Variasi' in rekap_copy.columns:
        # Terapkan fungsi format variasi baru
        rekap_copy['Formatted Variation'] = rekap_copy.apply(
            lambda row: format_variation_dama(row['Nama Variasi'], row['Nama Produk Original']),
            axis=1
        )
        # Buat Nama Produk Display
        rekap_copy['Nama Produk Display'] = rekap_copy.apply(
            lambda row: f"{row['Nama Produk Original']} ({row['Formatted Variation']})"
                        if row['Formatted Variation'] else row['Nama Produk Original'],
            axis=1
        )
    else:
         rekap_copy['Nama Produk Display'] = rekap_copy['Nama Produk Original']
         rekap_copy['Formatted Variation'] = ''

    # grouping_key = 'Nama Produk Display'
    # # --- AKHIR LOGIKA BARU ---

    # # Agregasi data utama dari REKAP
    # agg_dict = {
    #     'Nama Produk Original': 'first',
    #     'Nama Produk Display': 'first',
    #     'Cleaned Variation': 'first', # <-- Tambahkan ini jika perlu
    #     'Jumlah Terjual': 'sum', 'Harga Satuan': 'first', 'Total Harga Produk': 'sum',
    #     'Voucher Ditanggung Penjual': 'sum', 'Biaya Komisi AMS + PPN Shopee': 'sum',
    #     'Biaya Adm 8%': 'sum', 'Biaya Layanan 2%': 'sum',
    #     'Biaya Layanan Gratis Ongkir Xtra 4,5%': 'sum', 'Biaya Proses Pesanan': 'sum',
    #     'Total Penghasilan': 'sum'
    # }
    # summary_df = rekap_copy.groupby(grouping_key, as_index=False).agg(agg_dict)
    # summary_df.rename(columns={'Nama Produk Display': 'Nama Produk'}, inplace=True)
    
    grouping_key_list = ['Nama Produk Display', 'Harga Satuan']
    # --- ▲▲▲ AKHIR MODIFIKASI ▲▲▲ ---
    # --- AKHIR LOGIKA KHUSUS DAMA.ID STORE ---
    # summary_df = summary_df[summary_df['Total Penghasilan'] != 0].copy()

    # Agregasi data utama dari REKAP
    agg_dict = {
        'Nama Produk Original': 'first',
        'Nama Produk Display': 'first',
        # 'Cleaned Variation': 'first', 
        'Jumlah Terjual': 'sum', 
        'Total Harga Produk': 'sum',
        'Voucher Ditanggung Penjual': 'sum', 'Biaya Komisi AMS + PPN Shopee': 'sum',
        'Biaya Adm 8%': 'sum', 'Biaya Layanan 2%': 'sum',
        'Biaya Layanan Gratis Ongkir Xtra 4,5%': 'sum', 'Biaya Proses Pesanan': 'sum',
        'Total Penghasilan': 'sum'
    }
    # --- ▼▼▼ MODIFIKASI: Gunakan grouping_key_list ▼▼▼ ---
    summary_df = rekap_copy.groupby(grouping_key_list, as_index=False).agg(agg_dict)
    # --- ▲▲▲ AKHIR MODIFIKASI ▲▲▲ ---
    summary_df.rename(columns={'Nama Produk Display': 'Nama Produk'}, inplace=True)
    
    summary_df = summary_df[summary_df['Total Penghasilan'] != 0].copy()

    # --- LOGIKA IKLAN (Tetap sama) ---
    summary_df['Iklan Klik'] = 0.0
    produk_khusus_raw = ["CUSTOM AL QURAN MENGENANG/WAFAT 40/100/1000 HARI", "Paket Hemat Paket Al Quran | AQ Al Aqeel Wakaf Kerta koran Non Terjemah", "Alquran Al Aqeel A5 Kertas Koran Tanpa Terjemahan Wakaf Ibtida"]
    produk_khusus = [re.sub(r'\s+', ' ', name.replace('\xa0', ' ')).strip() for name in produk_khusus_raw]
    iklan_data = iklan_final_df[iklan_final_df['Nama Iklan'] != 'TOTAL'][['Nama Iklan', 'Biaya']].copy()
    # Konfigurasi Produk Khusus Dama
    force_config_dama = {
        "Alquran Al Aqeel A5 Kertas Koran Tanpa Terjemahan Wakaf Ibtida": {
            "variasi": ["A5 SATUAN", "B5 (Bigbos)", "A5 PAKET3", "A5 PAKET 5", "A5 PAKET 7"],
            "denom": 17
        },
        "Al Quran Wakaf Saku A6 Al Aqeel HVS Paket Wakaf": {
            "variasi": ["SATUAN", "PAKET ISI 3", "PAKET ISI 5", "PAKET ISI 7"],
            "denom": 16
        },
        "Al Quran Gold Silver Al Aqeel Besar Sedang Kecil": {
            "variasi": ["A4 Satuan", "B5 Satuan", "A7 Satuan", "A6 Satuan", "A5 Satuan", "A7 Paket isi 3", "A7 Paket isi 5", "A7 Paket isi 7", "A5 Paket isi 3"],
            "denom": 23
        }
    }

    for produk_base, config in force_config_dama.items():
        summary_df['Nama Produk Clean'] = summary_df['Nama Produk'].astype(str).str.replace(r'\s+', ' ', regex=True).str.strip()
        
        matching_ads = iklan_data[iklan_data['Nama Iklan'].str.contains(produk_base, case=False, na=False, regex=False)]
        if not matching_ads.empty:
            total_biaya_iklan = matching_ads['Biaya'].sum()
            denom = config['denom']
            
            for var in config['variasi']:
                nama_lengkap_search = f"{produk_base} ({var})".replace('  ', ' ').strip()
                exists = summary_df['Nama Produk Clean'].str.contains(re.escape(nama_lengkap_search), case=False, na=False).any()
                
                if not exists:
                    new_row = pd.DataFrame([{col: 0 for col in summary_df.columns}])
                    new_row['Nama Produk'] = f"{produk_base} ({var})"
                    summary_df = pd.concat([summary_df, new_row], ignore_index=True)
                    summary_df['Nama Produk Clean'] = summary_df['Nama Produk'].astype(str).str.replace(r'\s+', ' ', regex=True).str.strip()
            
            mask_summary = summary_df['Nama Produk'].str.contains(produk_base, case=False, na=False, regex=False)
            indices = summary_df[mask_summary].index
            
            for idx in indices:
                p_name = summary_df.at[idx, 'Nama Produk']
                count_same = (summary_df['Nama Produk'] == p_name).sum()
                mult = get_eksemplar_multiplier_dama(p_name)
                summary_df.at[idx, 'Iklan Klik'] = (mult * total_biaya_iklan) / denom / count_same
            
            iklan_data = iklan_data[~iklan_data['Nama Iklan'].str.contains(produk_base, case=False, na=False, regex=False)]

    summary_df.drop(columns=['Nama Produk Clean'], inplace=True, errors='ignore')

    # Logika Standar Dama untuk Tahlil
    if not iklan_data.empty:
        p_tahlil = "ALQURAN SAKU A6 EDISI TAHLIL TERBARU"
        matching_ads = iklan_data[iklan_data['Nama Iklan'].str.contains(p_tahlil, case=False, na=False, regex=False)]
        if not matching_ads.empty:
            total_biaya = matching_ads['Biaya'].sum()
            mask_summary = summary_df['Nama Produk'].str.contains(p_tahlil, case=False, na=False, regex=False)
            num_rows = mask_summary.sum()
            if num_rows > 0:
                summary_df.loc[mask_summary, 'Iklan Klik'] = total_biaya / num_rows
            else:
                # --- PERBAIKAN DI SINI ---
                # Jika 0 penjualan, buat baris baru agar biaya iklan tetap muncul di Summary
                new_row_ads = pd.DataFrame([{col: 0 for col in summary_df.columns}])
                new_row_ads['Nama Produk'] = p_biasa
                new_row_ads['Iklan Klik'] = total_biaya
                summary_df = pd.concat([summary_df, new_row_ads], ignore_index=True)
            iklan_data = iklan_data[~iklan_data['Nama Iklan'].str.contains(p_tahlil, case=False, na=False, regex=False)]
                
    summary_df = pd.merge(summary_df, iklan_data, left_on='Nama Produk Original', right_on='Nama Iklan', how='left')
    summary_df['Iklan Klik'] = summary_df['Iklan Klik'] + summary_df['Biaya'].fillna(0)
    summary_df.drop(columns=['Nama Iklan', 'Biaya'], inplace=True, errors='ignore')

    iklan_only_names = set(iklan_data['Nama Iklan']) - set(summary_df['Nama Produk Original'])
    if iklan_only_names:
        iklan_only_df = iklan_data[iklan_data['Nama Iklan'].isin(iklan_only_names)].copy()
        iklan_only_df.rename(columns={'Nama Iklan': 'Nama Produk', 'Biaya': 'Iklan Klik'}, inplace=True)
        iklan_only_df['Nama Produk Original'] = iklan_only_df['Nama Produk']
        summary_df = pd.concat([summary_df, iklan_only_df], ignore_index=True)
    summary_df.fillna(0, inplace=True)
    # --- AKHIR LOGIKA IKLAN ---

    # Hitungan selanjutnya
    summary_df['Penjualan Netto'] = (
        summary_df['Total Harga Produk'] - summary_df['Voucher Ditanggung Penjual'] -
        summary_df['Biaya Komisi AMS + PPN Shopee'] - summary_df['Biaya Adm 8%'] -
        summary_df['Biaya Layanan 2%'] - summary_df['Biaya Layanan Gratis Ongkir Xtra 4,5%'] -
        summary_df['Biaya Proses Pesanan']
    )
    # summary_df['Penjualan Netto'] = summary_df['Total Penghasilan']
    summary_df['Biaya Packing'] = summary_df['Jumlah Terjual'] * 200

    summary_df['Jumlah Eksemplar'] = summary_df.apply(
        lambda row: row['Jumlah Terjual'] * get_eksemplar_multiplier_dama(row['Nama Produk']),
        axis=1
    )
    
    # Terapkan Pengecualian DAMA.ID STORE
    # hijab_keywords_dama = {'PASHMINA', 'HIJAB', 'PASMINA'}
    hijab_keywords_dama = {'PIRING', 'BAJU', 'MOBIL'}
    # Gunakan 'Nama Produk Original' untuk pengecekan yang andal
    kondisi_hijab = summary_df['Nama Produk Original'].str.upper().str.contains('|'.join(hijab_keywords_dama), na=False)
    summary_df.loc[kondisi_hijab, 'Jumlah Eksemplar'] = 0
    
    summary_df['Biaya Ekspedisi'] = 0 # DAMA.ID STORE pakai Biaya Ekspedisi = 0
    biaya_ekspedisi_final = summary_df['Biaya Ekspedisi']

    # --- PANGGIL FUNGSI HARGA BELI BARU ---
    summary_df['Harga Beli'] = summary_df['Nama Produk'].apply(
        lambda x: get_harga_beli_dama(x, katalog_dama_df) # Panggil fungsi dama
    )
    # --- AKHIR PERUBAHAN ---

    # Harga Custom & Total Pembelian
    summary_df = pd.merge(
        summary_df,
        harga_custom_tlj_df[['LOOKUP_KEY', 'HARGA CUSTOM TLJ']],
        left_on='Nama Produk',
        right_on='LOOKUP_KEY', how='left'
    )
    summary_df.rename(columns={'HARGA CUSTOM TLJ': 'Harga Custom TLJ'}, inplace=True)
    summary_df['Harga Custom TLJ'] = summary_df['Harga Custom TLJ'].fillna(0)
    summary_df.drop(columns=['LOOKUP_KEY'], inplace=True, errors='ignore')

    produk_custom_str = "CUSTOM AL QURAN MENGENANG/WAFAT 40/100/1000 HARI"
    kondisi_custom = summary_df['Nama Produk Original'].str.contains(produk_custom_str, na=False)
    summary_df['Total Pembelian'] = np.where(
        kondisi_custom,
        (summary_df['Jumlah Terjual'] * summary_df['Harga Beli']) + (summary_df['Jumlah Terjual'] * summary_df['Harga Custom TLJ']),
        summary_df['Jumlah Terjual'] * summary_df['Harga Beli']
    )

    summary_df['Margin'] = (
        summary_df['Penjualan Netto'] - summary_df['Iklan Klik'] - summary_df['Biaya Packing'] -
        biaya_ekspedisi_final - summary_df['Total Pembelian']
    )

    # ... (Sisa fungsi, termasuk pembuatan DataFrame Final dan baris Total, tetap sama) ...
    # Pastikan kolom output 'Nama Produk' menggunakan `summary_df['Nama Produk']` (hasil display)
    # Hapus kolom 'Nama Produk Original' sebelum membuat baris total
    summary_df['Persentase'] = (summary_df.apply(lambda row: row['Margin'] / row['Total Harga Produk'] if row['Total Harga Produk'] != 0 else 0, axis=1))
    summary_df['Jumlah Pesanan'] = summary_df.apply(lambda row: row['Biaya Proses Pesanan'] / 1250 if 1250 != 0 else 0, axis=1)
    summary_df['Penjualan Per Hari'] = round(summary_df['Total Harga Produk'] / 7, 1)
    summary_df['Jumlah buku per pesanan'] = round(summary_df.apply(lambda row: row['Jumlah Eksemplar'] / row['Jumlah Pesanan'] if row.get('Jumlah Pesanan', 0) != 0 else 0, axis=1), 1)

    summary_final_data = {
        'No': np.arange(1, len(summary_df) + 1),
        'Nama Produk': summary_df['Nama Produk'], # Nama produk display
        'Jumlah Terjual': summary_df['Jumlah Terjual'], 'Jumlah Eksemplar': summary_df['Jumlah Eksemplar'], 
        'Jumlah Pesanan': summary_df['Jumlah Pesanan'], 'Harga Satuan': summary_df['Harga Satuan'],
        'Total Penjualan': summary_df['Total Harga Produk'], 'Voucher Ditanggung Penjual': summary_df['Voucher Ditanggung Penjual'],
        'Biaya Komisi AMS + PPN Shopee': summary_df['Biaya Komisi AMS + PPN Shopee'], 'Biaya Adm 8%': summary_df['Biaya Adm 8%'],
        'Biaya Layanan 2%': summary_df['Biaya Layanan 2%'], 'Biaya Layanan Gratis Ongkir Xtra 4,5%': summary_df['Biaya Layanan Gratis Ongkir Xtra 4,5%'],
        'Biaya Proses Pesanan': summary_df['Biaya Proses Pesanan'],
        'Penjualan Netto': summary_df['Penjualan Netto'], 'Iklan Klik': summary_df['Iklan Klik'], 'Biaya Packing': summary_df['Biaya Packing'],
        'Biaya Ekspedisi': biaya_ekspedisi_final, # Kolom Biaya Ekspedisi
        'Harga Beli': summary_df['Harga Beli'], 'Harga Custom TLJ': summary_df['Harga Custom TLJ'],
        'Total Pembelian': summary_df['Total Pembelian'], 'Margin': summary_df['Margin'],
        'Persentase': summary_df['Persentase'],
        'Penjualan Per Hari': summary_df['Penjualan Per Hari'], 'Jumlah buku per pesanan': summary_df['Jumlah buku per pesanan']
    }
    summary_final = pd.DataFrame(summary_final_data)

    # --- LOGIKA PERSINGKAT NAMA PRODUK DAMA.ID STORE ---
    mapping_dama = {
        "Alquran Al Aqeel A5 Kertas Koran Tanpa Terjemahan Wakaf Ibtida": "Al Aqeel A5 Kertas Koran",
        "AL QUR'AN CUSTOM NAMA FOTO DI COVER SISIPAN ACARA TASYAKUR TAHLIL YASIN": "AL QUR'AN CUSTOM COVER SISIPAN",
        "PAKET MURAH Alquran Al-Aqeel Tanpa Terjemahan | BANDUNG | Alquran Wakaf Hadiah Hampers Islami": "PAKET MURAH Al-Aqeel Tanpa Terjemahan",
        "Al Quran Gold Silver Al Aqeel Besar Sedang Kecil": "Al Aqeel Gold Silver",
        "ALQURAN A6 HVS EDISI TAHLIL TERBARU": "al aqeel A6 edisi tahlilan",
        "Al Quran Wakaf Saku A6 Al Aqeel HVS Paket Wakaf": "Al Aqeel A6 HVS",
        "AL QURAN LATIN TERJEMAHAN DAN TADJWID MUSHAF AL FIKRAH KERTAS HVS": "AL FIKRAH A5 HVS",
        "Al Quran Mushaf Al Aqeel Full Color A5 HVS": "Al Aqeel A5 HVS",
        "AL QURAN AL QUDDUS SAKU A7 KULIT RESLETING": "AL QUDDUS SAKU A7 KULIT",
        "BELLA SQUARE PREMIUM | HIJAB SEGIEMPAT | VARIASI WARNA | MURAH FASHION MUSLIM": "HIJAB SEGIEMPAT BELLA SQUARE",
        "Mushaf Al-Qur'an Al Quddus Tanpa terjemahan uk A5 DAN A4": "Al Quddus Tanpa terjemahan uk A5 DAN A4",
        "Juz'amma Edisi Terbaru Lebih Lengkap Terjemahan Tadjwid Asmaul Husna Soft Cover Kertas Koran": "Juz'amma Kertas Koran",
        "HIJAB PASMINA KAOS RAYON COOL TECH BY DAMA": "PASMINA KAOS RAYON",
        "PASHMINA OVAL CERUTY BABYDOLL PREMIUM": "PASHMINA OVAL CERUTY BABYDOLL",
        "BUKU CERITA ANAK SERI BUDI PEKERTI KOBER TK SD": "BUKU CERITA SERI BUDI PEKERTI TK SD",
        "AL QUR'AN TERJEMAHAN AL ALEEM WAQAF IBTIDA": "AL ALEEM WAQAF IBTIDA",
        "AlQuran Mushaf Al Aqeel B5": "Al Aqeel B5 HVS",
        "SERI DONGENG BINATANG | DONGENG FABEL | DONGENG BINATANG MENARIK": "SERI DONGENG BINATANG",
        "Buku Cerita Seri Terladan Nabi Seri 6 Untuk Anak Anak": "Buku Cerita Seri Teladan Nabi",
        "BUKU CERITA SERI CERITA RAKYAT | NUSANTARA": "BUKU CERITA SERI CERITA RAKYAT",
        "AL QUR'AN TADJWID DAN TERJEMAHAN TAFSIR ASBABUNNUZUL WAQAF IBTIDA MUSHAF MUMTAAZ": "AL QUR'AN TADJWID DAN TERJEMAHAN MUMTAAZ WAQAF IBTIDA",
        "Juz'amma Edisi Terbaru Lebih Lengkap Terjemahan Tajwid Asmaul Husnah kertas HVS": "Juz'amma kertas HVS",
        "Kamus Bergambar Bilingual TK SD PAUD": "Kamus Bergambar TK SD PAUD",
        "AL QURAN MUSHAF AL ALEEM A6 SAKU": "AL ALEEM A6 SAKU",
        "HIJAB PAYET CANTIK | PARIS JEPANG | hijab kekinian": "HIJAB PAYET PARIS JEPANG",
        "TERBARU KOMIK SERI PAHLAWAN INDONESIA | BUKU PAHLAWAN": "KOMIK SERI PAHLAWAN INDONESIA",
        "HARMONI NUSANTARA | LAGU NASIONAL DAN LAGU DAERAH INDONESIA": "LAGU NASIONAL DAN LAGU DAERAH INDONESIA",
        "HIJAB BERGO JERSEY BY DAMA | KERUDUNG INSTAN": "HIJAB BERGO JERSEY",
        "HIJAB VOAL MOTIF LASER CUT PREMIUM": "HIJAB VOAL MOTIF LASER CUT",
        "Al QURAN TADJWID TANPA TERJEMAHAN MUSHAF SUBHAAN": "SUBHAAN TADJWID TANPA TERJEMAHAN"
    }

    def apply_shorten_dama(nama_full):
        if pd.isna(nama_full): return nama_full
        nama_full_str = str(nama_full)
        # Ambil variasi dalam kurung jika ada
        match_variasi = re.search(r'(\s*\(.*\))$', nama_full_str)
        variasi_part = match_variasi.group(1) if match_variasi else ""
        nama_produk_saja = nama_full_str.replace(variasi_part, "").strip()

        for original_name, short_name in mapping_dama.items():
            if original_name.lower() in nama_produk_saja.lower():
                return f"{short_name}{variasi_part}"
        return nama_full_str

    summary_final['Nama Produk'] = summary_final['Nama Produk'].apply(apply_shorten_dama)
    # --- AKHIR LOGIKA PERSINGKAT ---
    
    # Pastikan semua data di kolom Nama Produk menjadi teks agar bisa diurutkan
    summary_final['Nama Produk'] = summary_final['Nama Produk'].astype(str)
    
    # Baru lakukan pengurutan
    summary_final = summary_final.sort_values(by='Nama Produk', ascending=True).reset_index(drop=True)
    summary_final['No'] = range(1, len(summary_final) + 1)

    if 'Nama Produk Original' in summary_final.columns:
         summary_final = summary_final.drop(columns=['Nama Produk Original'])

    total_row = pd.DataFrame(summary_final.sum(numeric_only=True)).T
    total_row['Nama Produk'] = 'Total'
    total_penjualan_netto = total_row['Penjualan Netto'].iloc[0]
    total_iklan_klik = total_row['Iklan Klik'].iloc[0]
    total_biaya_packing = total_row['Biaya Packing'].iloc[0]
    total_pembelian = total_row['Total Pembelian'].iloc[0]
    total_harga_produk = total_row['Total Penjualan'].iloc[0]
    total_biaya_proses_pesanan = total_row['Biaya Proses Pesanan'].iloc[0]
    total_jumlah_terjual = total_row['Jumlah Terjual'].iloc[0]
    total_jumlah_eksemplar = total_row['Jumlah Eksemplar'].iloc[0] # <-- DITAMBAH
    total_biaya_ekspedisi = total_row['Biaya Ekspedisi'].iloc[0]
    total_margin = total_penjualan_netto - total_biaya_packing - total_biaya_ekspedisi - total_pembelian - total_iklan_klik
    total_row['Margin'] = total_margin
    total_row['Persentase'] = (total_margin / total_harga_produk) if total_harga_produk != 0 else 0
    total_jumlah_pesanan = (total_biaya_proses_pesanan / 1250) if 1250 != 0 else 0
    total_row['Jumlah Pesanan'] = total_jumlah_pesanan
    total_row['Penjualan Per Hari'] = round(total_harga_produk / 7, 1)
    total_row['Jumlah buku per pesanan'] = round(total_jumlah_eksemplar / total_jumlah_pesanan if total_jumlah_pesanan != 0 else 0, 1)
    for col in ['Harga Satuan', 'Harga Beli', 'No', 'Harga Custom TLJ']:
        if col in total_row.columns: total_row[col] = None
    summary_with_total = pd.concat([summary_final, total_row], ignore_index=True)

    return summary_with_total

def get_harga_beli_fuzzy_tiktok(nama_produk, variasi, katalog_df, score_threshold_primary=80, score_threshold_fallback=75):
    """
    Mencari harga beli khusus untuk TikTok dengan logika baru:
    - Jika ada variasi, hapus semua ukuran (A5, B5, dll.) dari nama produk, lalu gabungkan.
    - Jika tidak ada variasi, gunakan nama produk asli.
    """
    nama_produk_clean = str(nama_produk).strip()
    variasi_clean = str(variasi).strip()

    # Pola regex untuk menemukan dan menghapus ukuran seperti A5, B5, dll.
    size_pattern = r'\s*\b(A|B)\d{1,2}\b\s*'

    # Jika ada variasi yang valid (bukan string kosong)
    if pd.notna(variasi) and variasi_clean:
        # Hapus semua pola ukuran dari string nama produk
        nama_produk_tanpa_ukuran = re.sub(size_pattern, ' ', nama_produk_clean, flags=re.IGNORECASE).strip()
        # Gabungkan dengan Variasi di depan untuk prioritas pencarian
        search_term = f"{variasi_clean} {nama_produk_tanpa_ukuran}"
    else:
        # Jika tidak ada variasi, gunakan nama produk apa adanya
        search_term = nama_produk_clean

    # Panggil fungsi fuzzy matching yang sudah ada dengan search_term yang baru dan lebih bersih
    return get_harga_beli_fuzzy(search_term, katalog_df, score_threshold_primary=score_threshold_primary, score_threshold_fallback=score_threshold_fallback)
    
def parse_pdf_receipt(pdf_file):
    """Mengekstrak tanggal dan total nominal dari satu file PDF nota Lalamove."""
    try:
        full_text = ""
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                full_text += page.extract_text() + "\n"

        # Pola untuk tanggal (misal: 02 Okt 2025)
        date_match = re.search(r'(\d{2})\s+(\w+)\s+(\d{4})', full_text)
        # Pola untuk total harga (misal: Rp9.402)
        total_match = re.search(r'Total Harga\s*\(IDR\)\s*Rp([\d\.]+)', full_text)

        if not total_match: # Fallback jika format sedikit berbeda
             total_match = re.search(r'Total Harga\s*Rp([\d\.]+)', full_text)

        tanggal = None
        if date_match:
            day, month_str, year = date_match.groups()
            month_map = {
                'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'Mei': '05', 'Jun': '06',
                'Jul': '07', 'Agu': '08', 'Sep': '09', 'Okt': '10', 'Nov': '11', 'Des': '12'
            }
            month = month_map.get(month_str[:3], '00')
            tanggal = f"{day}-{month}-{year}"

        nominal = float(total_match.group(1).replace('.', '')) if total_match else 0
        
        return {'Tanggal Kirim Paket': tanggal, 'Nominal': nominal}
    except Exception as e:
        st.warning(f"Gagal memproses PDF: {pdf_file.name}. Error: {e}")
        return None

# KODE BARU (Ganti seluruh fungsi ini)
def process_rekap_tiktok(order_details_df, semua_pesanan_df, creator_order_all_df, store_choice):
    """Fungsi untuk memproses dan membuat sheet 'REKAP' untuk TikTok dengan logika baru."""
    # 1. PREPARASI DATA & MERGE AWAL
    order_details_df['ORDER/ADJUSTMENT ID'] = order_details_df['ORDER/ADJUSTMENT ID'].astype(str)
    semua_pesanan_df['ORDER ID'] = semua_pesanan_df['ORDER ID'].astype(str)
    creator_order_all_df['ID PESANAN'] = creator_order_all_df['ID PESANAN'].astype(str)

    # Pastikan nama kolom konsisten (misal: UPPERCASE seperti di kode Anda selanjutnya)
    order_details_df.columns = [col.upper().strip() for col in order_details_df.columns]
    semua_pesanan_df.columns = [col.upper().strip() for col in semua_pesanan_df.columns]
    creator_order_all_df.columns = [col.upper().strip() for col in creator_order_all_df.columns]


    # 2. MERGE AWAL (Kode Anda yang sudah ada)
    rekap_df = pd.merge(
        order_details_df,
        semua_pesanan_df,
        left_on='ORDER/ADJUSTMENT ID',
        right_on='ORDER ID',
        how='left'
    )

    key_cols = ['ORDER ID', 'PRODUCT NAME', 'VARIATION', 'QUANTITY', 'SKU SUBTOTAL BEFORE DISCOUNT', 'SKU SELLER DISCOUNT']
    # Pastikan semua kolom kunci ada sebelum mencoba drop_duplicates
    if all(col in rekap_df.columns for col in key_cols):
        rows_before_dedup = len(rekap_df)
        rekap_df.drop_duplicates(subset=key_cols, keep='first', inplace=True)
        rows_after_dedup = len(rekap_df)
        if rows_before_dedup > rows_after_dedup:
            st.info(f"Menghapus {rows_before_dedup - rows_after_dedup} baris duplikat setelah merge.")
    else:
        st.warning(f"Tidak dapat melakukan de-duplikasi setelah merge: Kolom kunci {key_cols} tidak lengkap.")

    # 3. FILTER PESANAN BATAL/REFUND & SETTLEMENT NOL (Kode Anda yang sudah ada)
    # ... (Blok filter Cancel/Return Anda) ...
    if 'CANCELLATION/RETURN TYPE' in rekap_df.columns:
        cancelled_orders = rekap_df[rekap_df['CANCELLATION/RETURN TYPE'].fillna('').isin(['Cancel', 'Return/Refund'])]['ORDER ID'].unique()
        if len(cancelled_orders) > 0:
            st.warning(f"Menghapus {len(cancelled_orders)} pesanan karena status Cancel/Return...")
            rekap_df = rekap_df[~rekap_df['ORDER ID'].isin(cancelled_orders)].copy()

    # ... (Blok filter Total Settlement Amount Anda) ...
    if 'TOTAL SETTLEMENT AMOUNT' in order_details_df.columns and 'ORDER/ADJUSTMENT ID' in order_details_df.columns:
         zero_settlement_ids = order_details_df[pd.to_numeric(order_details_df['TOTAL SETTLEMENT AMOUNT'], errors='coerce').fillna(0) == 0]['ORDER/ADJUSTMENT ID'].astype(str).unique()
         if len(zero_settlement_ids) > 0:
             orders_before_filter = len(rekap_df['ORDER ID'].unique())
             rekap_df = rekap_df[~rekap_df['ORDER ID'].astype(str).isin(zero_settlement_ids)].copy()
             orders_after_filter = len(rekap_df['ORDER ID'].unique())
             removed_count = orders_before_filter - orders_after_filter
             if removed_count > 0:
                 st.warning(f"Menghapus {removed_count} pesanan tambahan karena Total Settlement Amount = 0.")


    # 4. EKSTRAKSI VARIASI & PEMBERSIHAN DATA SEBELUM GROUPBY (Kode Anda yang sudah ada)
    rekap_df['Variasi'] = rekap_df['VARIATION'].str.extract(r'\b(A\d{1,2}|B\d{1,2})\b', expand=False).fillna('')
    if 'PRODUCT NAME' in rekap_df.columns:
        rekap_df['PRODUCT NAME'] = rekap_df['PRODUCT NAME'].astype(str).str.strip()
    if 'Variasi' in rekap_df.columns:
        rekap_df['Variasi'] = rekap_df['Variasi'].astype(str).str.strip()

    cols_to_clean = [
        'SKU SUBTOTAL BEFORE DISCOUNT', 'SKU SELLER DISCOUNT', 'QUANTITY', 
        'BONUS CASHBACK SERVICE FEE', 'VOUCHER XTRA SERVICE FEE', 'TOTAL SETTLEMENT AMOUNT',
        'SKU UNIT ORIGINAL PRICE', 'PRE-ORDER SERVICE FEE', 'AFFILIATE SHOP ADS COMMISSION' # Penting untuk Harga Satuan nanti
    ]
    for col in cols_to_clean:
        if col in rekap_df.columns:
            # Menggunakan regex yang lebih sederhana untuk angka (termasuk desimal jika ada)
            rekap_df[col] = (rekap_df[col].astype(str)
                             .str.replace(r'[^\d\.\-]', '', regex=True)) # Izinkan titik dan minus
            rekap_df[col] = pd.to_numeric(rekap_df[col], errors='coerce').fillna(0).abs() # .abs() sebaiknya di akhir

    if 'ORDER CREATED TIME(UTC)' in rekap_df.columns:
        created_time_col = 'ORDER CREATED TIME(UTC)'
    elif 'ORDER CREATED TIME' in rekap_df.columns:
        created_time_col = 'ORDER CREATED TIME'
    else:
        # Pengaman jika kolom tidak ada
        st.warning("Kolom 'ORDER CREATED TIME(UTC)' atau 'ORDER CREATED TIME' tidak ditemukan. 'Waktu Pesanan Dibuat' akan kosong.")
        rekap_df['ORDER CREATED TIME_MISSING'] = pd.NaT # Buat kolom dummy
        created_time_col = 'ORDER CREATED TIME_MISSING' # Gunakan kolom dummy
        
    if 'ORDER SETTLED TIME(UTC)' in rekap_df.columns:
        settled_time_col = 'ORDER SETTLED TIME(UTC)'
    elif 'ORDER SETTLED TIME' in rekap_df.columns:
        settled_time_col = 'ORDER SETTLED TIME'
    else:
        # Pengaman jika kolom tidak ada
        st.warning("Kolom 'ORDER SETTLED TIME(UTC)' atau 'ORDER SETTLED TIME' tidak ditemukan. 'Waktu Dana Dilepas' akan kosong.")
        rekap_df['ORDER SETTLED TIME_MISSING'] = pd.NaT # Buat kolom dummy
        settled_time_col = 'ORDER SETTLED TIME_MISSING' # Gunakan kolom dummy

    rekap_df['Harga Satuan Temp'] = rekap_df['SKU UNIT ORIGINAL PRICE'] - (rekap_df['SKU SELLER DISCOUNT'] / rekap_df['QUANTITY'].replace(0, 1))

    product_count = rekap_df.groupby('ORDER ID')['ORDER ID'].transform('size')
    rekap_df['Biaya Pre-order'] = rekap_df['PRE-ORDER SERVICE FEE'] / product_count
    rekap_df['Komisi Iklan Affiliate'] = rekap_df['AFFILIATE SHOP ADS COMMISSION'] / product_count

    # 2. LOGIKA AGREGASI PRODUK (Sekarang akan bekerja dengan benar)
    agg_rules = {
        'QUANTITY': 'sum', # <-- Penjumlahan Kuantitas terjadi di sini
        'SKU SUBTOTAL BEFORE DISCOUNT': 'sum',
        'SKU SELLER DISCOUNT': 'sum',
        'PRE-ORDER SERVICE FEE': 'sum', # Ambil salah satu saja karena akan dibagi
        'Biaya Pre-order': 'sum',
        'Komisi Iklan Affiliate': 'first',
        'Harga Satuan Temp': 'first', # Ambil harga satuan pertama
        'BONUS CASHBACK SERVICE FEE': 'first', # Jumlahkan biaya ini
        'VOUCHER XTRA SERVICE FEE': 'first',   # Jumlahkan biaya ini
        'TOTAL SETTLEMENT AMOUNT': 'first' # Ambil settlement amount pertama (biasanya sama per pesanan)
    }

    # Tambahkan kolom waktu secara dinamis menggunakan variabel yang kita buat
    agg_rules[created_time_col] = 'first'
    agg_rules[settled_time_col] = 'first'
    
    # Grouping berdasarkan ID Pesanan, Nama Produk, dan Variasi
    rekap_df = rekap_df.groupby(['ORDER ID', 'PRODUCT NAME', 'Variasi'], as_index=False).agg(agg_rules)
    rekap_df.rename(columns={'QUANTITY': 'Jumlah Terjual'}, inplace=True) # Ganti nama setelah agregasi
    
    # 3. MENGHITUNG BIAYA-BIAYA BARU (setelah agregasi)
    rekap_df['Total Penjualan'] = rekap_df['SKU SUBTOTAL BEFORE DISCOUNT'] - rekap_df['SKU SELLER DISCOUNT']
    rekap_df['Biaya Komisi Platform 8%'] = rekap_df['Total Penjualan'] * 0.08
    rekap_df['Komisi Dinamis 5%'] = rekap_df['Total Penjualan'] * 0.05
    
    product_count = rekap_df.groupby('ORDER ID')['ORDER ID'].transform('size')
    rekap_df['Biaya Layanan Cashback Bonus 1,5%'] = rekap_df['BONUS CASHBACK SERVICE FEE'] / product_count
    rekap_df['Biaya Layanan Voucher Xtra'] = rekap_df['VOUCHER XTRA SERVICE FEE'] / product_count
    rekap_df['Biaya Proses Pesanan'] = 1250 / product_count

    # 4. MENGAMBIL KOMISI AFFILIATE
    creator_order_all_df['Variasi_Clean'] = creator_order_all_df['SKU'].str.extract(r'\b(A\d{1,2}|B\d{1,2})\b', expand=False).fillna('')
    # Merge affiliate HANYA jika bukan DAMA.ID STORE
    # if store_choice != "DAMA.ID STORE":
    #     rekap_df = pd.merge(
    #         rekap_df,
    #         creator_order_all_df[['ID PESANAN', 'PRODUK', 'Variasi_Clean', 'PEMBAYARAN KOMISI AKTUAL']],
    #         left_on=['ORDER ID', 'PRODUCT NAME', 'Variasi'],
    #         right_on=['ID PESANAN', 'PRODUK', 'Variasi_Clean'],
    #         how='left'
    #     )
    #     rekap_df.rename(columns={'PEMBAYARAN KOMISI AKTUAL': 'Komisi Affiliate'}, inplace=True)
    #     rekap_df['Komisi Affiliate'] = pd.to_numeric(rekap_df['Komisi Affiliate'], errors='coerce').fillna(0).abs() # Pastikan numerik dan positif
    #     rekap_df.drop(columns=['ID PESANAN', 'PRODUK', 'Variasi_Clean'], inplace=True, errors='ignore')
    # else:
    #     # Jika DAMA.ID STORE, buat kolom Komisi Affiliate berisi 0
    #     rekap_df['Komisi Affiliate'] = 0
    if not creator_order_all_df.empty:
        # Pastikan kolom SKU ada sebelum extract
        if 'SKU' in creator_order_all_df.columns:
            creator_order_all_df['Variasi_Clean'] = creator_order_all_df['SKU'].str.extract(r'\b(A\d{1,2}|B\d{1,2})\b', expand=False).fillna('')
        else:
            creator_order_all_df['Variasi_Clean'] = ''

        rekap_df = pd.merge(
            rekap_df,
            creator_order_all_df[['ID PESANAN', 'PRODUK', 'Variasi_Clean', 'PERKIRAAN PEMBAYARAN KOMISI STANDAR']],
            left_on=['ORDER ID', 'PRODUCT NAME', 'Variasi'],
            right_on=['ID PESANAN', 'PRODUK', 'Variasi_Clean'],
            how='left'
        )
        rekap_df.rename(columns={'PERKIRAAN PEMBAYARAN KOMISI STANDAR': 'Komisi Affiliate'}, inplace=True)
        rekap_df['Komisi Affiliate'] = pd.to_numeric(rekap_df['Komisi Affiliate'], errors='coerce').fillna(0).abs()
        rekap_df.drop(columns=['ID PESANAN', 'PRODUK', 'Variasi_Clean'], inplace=True, errors='ignore')
    else:
        # Jika file tidak diupload (DataFrame kosong), isi 0
        rekap_df['Komisi Affiliate'] = 0

    # 5. RUMUS BARU UNTUK TOTAL PENGHASILAN
    rekap_df['Total Penghasilan'] = (
        rekap_df['Total Penjualan'] -
        rekap_df['Komisi Affiliate'] -
        rekap_df['Biaya Komisi Platform 8%'] -
        rekap_df['Komisi Dinamis 5%'] -
        rekap_df['Biaya Layanan Cashback Bonus 1,5%'] -
        rekap_df['Biaya Layanan Voucher Xtra'] -
        rekap_df['Biaya Proses Pesanan'] -
        rekap_df['Biaya Pre-order'] -
        rekap_df['Komisi Iklan Affiliate']
    )

    # 6. MEMBUAT FINAL DATAFRAME
    rekap_final = pd.DataFrame({
        'No.': np.arange(1, len(rekap_df) + 1),
        'No. Pesanan': rekap_df['ORDER ID'],
        'Waktu Pesanan Dibuat': rekap_df[created_time_col],
        'Waktu Dana Dilepas': rekap_df[settled_time_col],
        'Nama Produk': rekap_df['PRODUCT NAME'],
        'Variasi': rekap_df['Variasi'],
        'Jumlah Terjual': rekap_df['Jumlah Terjual'],
        # 'Harga Satuan': rekap_df['SKU UNIT ORIGINAL PRICE'],
        'Harga Satuan': rekap_df['Harga Satuan Temp'],
        'Total Harga Sebelum Diskon': rekap_df['SKU SUBTOTAL BEFORE DISCOUNT'],
        'Diskon Penjual': rekap_df['SKU SELLER DISCOUNT'],
        'Total Penjualan': rekap_df['Total Penjualan'],
        'Komisi Affiliate': rekap_df['Komisi Affiliate'],
        'Biaya Komisi Platform 8%': rekap_df['Biaya Komisi Platform 8%'],
        'Komisi Dinamis 5%': rekap_df['Komisi Dinamis 5%'],
        'Komisi Iklan Affiliate': rekap_df['Komisi Iklan Affiliate'],
        'Biaya Pre-order':rekap_df['Biaya Pre-order'],
        'Biaya Layanan Cashback Bonus 1,5%': rekap_df['Biaya Layanan Cashback Bonus 1,5%'],
        'Biaya Layanan Voucher Xtra': rekap_df['Biaya Layanan Voucher Xtra'],
        'Biaya Proses Pesanan': rekap_df['Biaya Proses Pesanan'],
        'Total Penghasilan': rekap_df['Total Penghasilan']
    })

    # 1. Tentukan kolom mana yang akan dijumlahkan dan mana yang akan diambil nilai pertamanya
    cols_to_sum = [
        'Komisi Affiliate'        
    ]
    
    # Kolom yang nilainya sama untuk semua duplikat, jadi kita ambil yang pertama
    cols_to_first = [
        'Waktu Pesanan Dibuat',
        'Waktu Dana Dilepas',
        'Jumlah Terjual',
        'Total Harga Sebelum Diskon',
        'Diskon Penjual',
        'Komisi Iklan Affiliate',
        'Biaya Pre-order',
        'Biaya Layanan Cashback Bonus 1,5%',
        'Biaya Layanan Voucher Xtra',
        'Harga Satuan',
        'Biaya Proses Pesanan'
    ]
    
    # Buat dictionary aturan agregasi
    agg_rules_final = {col: 'sum' for col in cols_to_sum}
    agg_rules_final.update({col: 'first' for col in cols_to_first})
    
    # 2. Lakukan grouping berdasarkan No. Pesanan, Nama Produk, dan Variasi
    #    'as_index=False' penting agar kolom grouping tidak menjadi index
    rekap_final = rekap_final.groupby(['No. Pesanan', 'Nama Produk', 'Variasi'], as_index=False).agg(agg_rules_final)
    
    # 3. Hitung ulang kolom-kolom yang bergantung pada hasil agregasi
    
    # Hitung ulang Total Penjualan dari komponen yang sudah dijumlahkan
    rekap_final['Total Penjualan'] = rekap_final['Total Harga Sebelum Diskon'] - rekap_final['Diskon Penjual']
    
    # Hitung ulang biaya berbasis persentase
    rekap_final['Biaya Komisi Platform 8%'] = rekap_final['Total Penjualan'] * 0.08
    rekap_final['Komisi Dinamis 5%'] = rekap_final['Total Penjualan'] * 0.05
    
    # Hitung ulang Total Penghasilan
    rekap_final['Total Penghasilan'] = (
        rekap_final['Total Penjualan'] -
        rekap_final['Komisi Affiliate'] -
        rekap_final['Biaya Komisi Platform 8%'] -
        rekap_final['Komisi Dinamis 5%'] -
        rekap_final['Biaya Layanan Cashback Bonus 1,5%'] -
        rekap_final['Biaya Layanan Voucher Xtra'] -
        rekap_final['Biaya Proses Pesanan'] -
        rekap_final['Biaya Pre-order'] -
        rekap_final['Komisi Iklan Affiliate']
    )
    
    # 4. Susun ulang kolom dan perbarui nomor baris 'No.'
    final_columns_order = [
        'No.', 'No. Pesanan', 'Waktu Pesanan Dibuat', 'Waktu Dana Dilepas', 'Nama Produk',
        'Variasi', 'Jumlah Terjual', 'Harga Satuan', 'Total Harga Sebelum Diskon',
        'Diskon Penjual', 'Total Penjualan', 'Komisi Affiliate',
        'Biaya Komisi Platform 8%', 'Komisi Dinamis 5%', 'Komisi Iklan Affiliate', 'Biaya Pre-order', 'Biaya Layanan Cashback Bonus 1,5%',
        'Biaya Layanan Voucher Xtra', 'Biaya Proses Pesanan', 'Total Penghasilan'
    ]
    rekap_final = rekap_final.reindex(columns=final_columns_order)
    rekap_final['No.'] = np.arange(1, len(rekap_final) + 1)

    cols_to_blank = ['No. Pesanan', 'Waktu Pesanan Dibuat', 'Waktu Dana Dilepas']
    rekap_final.loc[rekap_final['No. Pesanan'].duplicated(), cols_to_blank] = ''

    return rekap_final.fillna(0)

def process_summary_tiktok(rekap_df, katalog_df, harga_custom_tlj_df, ekspedisi_df, product_data_df, store_choice):
    """Fungsi untuk memproses dan membuat sheet 'SUMMARY' untuk TikTok."""
    # Agregasi data dari REKAP berdasarkan Nama Produk dan Variasi (ini sudah mencegah duplikasi)
    summary_df = rekap_df.groupby(['Nama Produk', 'Variasi']).agg({
        'Jumlah Terjual': 'sum',
        'Harga Satuan': 'first',
        'Diskon Penjual': 'sum',
        'Total Penjualan': 'sum',
        'Komisi Affiliate': 'sum',
        'Biaya Komisi Platform 8%': 'sum',
        'Komisi Dinamis 5%': 'sum',
        'Komisi Iklan Affiliate': 'sum',
        'Biaya Pre-order': 'sum',
        'Biaya Layanan Cashback Bonus 1,5%': 'sum',
        'Biaya Layanan Voucher Xtra': 'sum',
        'Biaya Proses Pesanan': 'sum',
    }).reset_index()

    summary_df['Komisi Affiliate'] = summary_df['Komisi Affiliate'] + summary_df['Komisi Iklan Affiliate']

    # Hitung Penjualan Netto
    summary_df['Penjualan Netto'] = (
        summary_df['Total Penjualan'] -
        summary_df['Komisi Affiliate'] -
        summary_df['Biaya Komisi Platform 8%'] -
        summary_df['Komisi Dinamis 5%'] -
        summary_df['Biaya Layanan Cashback Bonus 1,5%'] -
        summary_df['Biaya Layanan Voucher Xtra'] -
        summary_df['Biaya Proses Pesanan'] -
        summary_df['Biaya Pre-order']
    )
    
    # # 1. Ambil 'Nama Produk', 'Variasi', dan 'Jumlah' dari sheet EKSPEDISI
    # ekspedisi_cost = ekspedisi_df[['Nama Produk', 'Variasi', 'Jumlah']].rename(columns={'Jumlah': 'Biaya Ekspedisi'})
    
    # # 2. Gabungkan (merge) ke summary_df menggunakan KEDUA kolom sebagai kunci
    # summary_df = pd.merge(
    #     summary_df, 
    #     ekspedisi_cost, 
    #     on=['Nama Produk', 'Variasi'],  # <-- Kunci perubahannya di sini
    #     how='left'
    # )
    
    # # 3. Isi nilai yang tidak cocok (NaN) dengan 0
    # summary_df['Biaya Ekspedisi'] = summary_df['Biaya Ekspedisi'].fillna(0)
    if store_choice == "DAMA.ID STORE":
        # 1. Untuk DAMA.ID STORE, Biaya Ekspedisi selalu 0
        summary_df['Biaya Ekspedisi'] = 0
    else:
        # 1. Ambil 'Nama Produk', 'Variasi', dan 'Jumlah' dari sheet EKSPEDISI
        ekspedisi_cost = ekspedisi_df[['Nama Produk', 'Variasi', 'Jumlah']].rename(columns={'Jumlah': 'Biaya Ekspedisi'})
        
        # 2. Gabungkan (merge) ke summary_df menggunakan KEDUA kolom sebagai kunci
        summary_df = pd.merge(
            summary_df, 
            ekspedisi_cost, 
            on=['Nama Produk', 'Variasi'],
            how='left'
        )
        
        # 3. Isi nilai yang tidak cocok (NaN) dengan 0 dan pastikan numerik
        summary_df['Biaya Ekspedisi'] = pd.to_numeric(summary_df['Biaya Ekspedisi'], errors='coerce').fillna(0)

    summary_df['Biaya Packing'] = summary_df['Jumlah Terjual'] * 200

    # --- PERUBAHAN DI SINI: Gunakan logika harga beli yang sama dengan Shopee ---
    # Untuk TikTok, kita tidak memiliki 'Nama Variasi' dari file income,
    # jadi kita tidak perlu memberikan rekap_lookup_df. Logika custom akan dilewati.
    summary_df['Harga Beli'] = summary_df.apply(
        lambda row: get_harga_beli_fuzzy_tiktok(row['Nama Produk'], row['Variasi'], katalog_df),
        axis=1
    )
    # --- AKHIR PERUBAHAN ---
    
    # --- LOGIKA BARU UNTUK HARGA CUSTOM TLJ (TIKTOK) ---
    # 1. Buat kolom kunci di summary_df untuk pencocokan
    summary_df['LOOKUP_KEY'] = summary_df['Nama Produk'].astype(str).str.strip() + ' ' + summary_df['Variasi'].astype(str).str.strip()
    
    # 2. Gabungkan dengan data harga custom
    summary_df = pd.merge(
        summary_df,
        harga_custom_tlj_df[['LOOKUP_KEY', 'HARGA CUSTOM TLJ']],
        on='LOOKUP_KEY',
        how='left'
    )
    summary_df.rename(columns={'HARGA CUSTOM TLJ': 'Harga Custom TLJ'}, inplace=True)
    summary_df['Harga Custom TLJ'] = summary_df['Harga Custom TLJ'].fillna(0)
    summary_df.drop(columns=['LOOKUP_KEY'], inplace=True, errors='ignore')

    # --- LOGIKA BARU UNTUK TOTAL PEMBELIAN (TIKTOK) ---
    produk_custom_str = "CUSTOM AL QURAN MENGENANG/WAFAT 40/100/1000 HARI"
    kondisi_custom = summary_df['Nama Produk'].str.contains(produk_custom_str, na=False)
    
    summary_df['Total Pembelian'] = np.where(
        kondisi_custom,
        (summary_df['Jumlah Terjual'] * summary_df['Harga Beli']) + (summary_df['Jumlah Terjual'] * summary_df['Harga Custom TLJ']),
        summary_df['Jumlah Terjual'] * summary_df['Harga Beli']
    )

    # 2. Logika Distribusi Iklan (MODIFIKASI: Menambahkan produk iklan tanpa penjualan)
    if not product_data_df.empty:
        # Ambil kolom Biaya dan Nama Produk
        ads_df = product_data_df[['NAMA PRODUK', 'BIAYA']].copy()
        
        # Hitung berapa banyak variasi untuk setiap Nama Produk yang ADA di penjualan
        var_count_per_product = summary_df.groupby('Nama Produk')['Variasi'].transform('count')
        summary_df['var_count'] = var_count_per_product

        # Merge dengan 'outer' agar produk di ads_df yang tidak ada di summary_df tetap masuk
        summary_df = pd.merge(
            summary_df, 
            ads_df, 
            left_on='Nama Produk', 
            right_on='NAMA PRODUK', 
            how='outer'
        )
        
        # Jika Nama Produk kosong (hasil outer merge dari iklan saja), isi dari NAMA PRODUK iklan
        summary_df['Nama Produk'] = summary_df['Nama Produk'].fillna(summary_df['NAMA PRODUK'])
        
        # Hitung Iklan: 
        # Jika ada penjualannya (var_count > 0), bagi biayanya. 
        # Jika tidak ada penjualan (iklan saja), tampilkan biaya penuh.
        summary_df['Iklan'] = np.where(
            summary_df['var_count'] > 0,
            summary_df['BIAYA'].fillna(0) / summary_df['var_count'],
            summary_df['BIAYA'].fillna(0)
        )
        
        # Hapus kolom pembantu
        summary_df.drop(columns=['NAMA PRODUK', 'BIAYA', 'var_count'], inplace=True, errors='ignore')
    else:
        summary_df['Iklan'] = 0
    
    summary_df['Margin'] = (
        summary_df['Penjualan Netto'] -
        summary_df['Biaya Packing'] -
        summary_df['Biaya Ekspedisi'] -
        summary_df['Total Pembelian'] -
        summary_df['Biaya Pre-order']
    )
    
    # ... (Sisa fungsi Anda dari sini sampai akhir tetap sama persis) ...
    summary_df['Persentase'] = summary_df.apply(lambda row: row['Margin'] / row['Total Penjualan'] if row['Total Penjualan'] != 0 else 0, axis=1)
    summary_df['Jumlah Pesanan'] = summary_df['Biaya Proses Pesanan'] / 1250
    summary_df['Total Pemasukan'] = summary_df['Jumlah Terjual'] * summary_df['Harga Satuan']
    summary_df['Penjualan Per Hari'] = round(summary_df['Penjualan Netto'] / 7, 1)
    summary_df['Jumlah buku per pesanan'] = summary_df.apply(lambda row: row['Jumlah Terjual'] / row['Jumlah Pesanan'] if row.get('Jumlah Pesanan', 0) != 0 else 0, axis=1)

    summary_final = pd.DataFrame({
        'No': np.arange(1, len(summary_df) + 1), 'Nama Produk': summary_df['Nama Produk'], 'Variasi': summary_df['Variasi'],
        'Jumlah Terjual': summary_df['Jumlah Terjual'], 'Jumlah Pesanan': summary_df['Jumlah Pesanan'], 'Harga Satuan': summary_df['Harga Satuan'],
        'Total Penjualan': summary_df['Total Pemasukan'],
        # 'Total Diskon Penjual': summary_df['Diskon Penjual'], 'Total Harga Sesudah Diskon': summary_df['Total Penjualan'], 
        'Komisi Affiliate': summary_df['Komisi Affiliate'], 'Biaya Komisi Platform 8%': summary_df['Biaya Komisi Platform 8%'],
        'Komisi Dinamis 5%': summary_df['Komisi Dinamis 5%'], 'Biaya Pre-order': summary_df['Biaya Pre-order'], 'Biaya Layanan Cashback Bonus 1,5%': summary_df['Biaya Layanan Cashback Bonus 1,5%'],
        'Biaya Layanan Voucher Xtra': summary_df['Biaya Layanan Voucher Xtra'], 'Biaya Proses Pesanan': summary_df['Biaya Proses Pesanan'],
        'Penjualan Netto': summary_df['Penjualan Netto'], 'Iklan': summary_df['Iklan'],'Biaya Packing': summary_df['Biaya Packing'],
        'Biaya Ekspedisi': summary_df['Biaya Ekspedisi'], 'Harga Beli': summary_df['Harga Beli'],
        'Harga Custom TLJ': summary_df['Harga Custom TLJ'], 'Total Pembelian': summary_df['Total Pembelian'],
        'Margin': summary_df['Margin'], 'Persentase': summary_df['Persentase'],
        'Penjualan Per Hari': summary_df['Penjualan Per Hari'], 'Jumlah buku per pesanan': summary_df['Jumlah buku per pesanan']
    })

    summary_final = summary_final.drop_duplicates(subset=['Nama Produk', 'Variasi'], keep='first').reset_index(drop=True)
    summary_final = summary_final.sort_values(by='Nama Produk', ascending=True).reset_index(drop=True)
    summary_final['No'] = range(1, len(summary_final) + 1)

    total_row = pd.DataFrame(summary_final.sum(numeric_only=True)).T
    total_row['Nama Produk'] = 'Total'
    total_margin = total_row['Penjualan Netto'].iloc[0] - total_row['Biaya Packing'].iloc[0] - total_row['Biaya Ekspedisi'].iloc[0] - total_row['Total Pembelian'].iloc[0]
    total_row['Margin'] = total_margin
    total_penjualan = total_row['Total Penjualan'].iloc[0]
    total_iklan = total_row['Biaya Pre-order'].iloc[0]
    total_row['Persentase'] = (total_margin / total_penjualan) if total_penjualan != 0 else 0
    total_row['Penjualan Per Hari'] = round(total_penjualan / 7, 1)
    total_jumlah_pesanan = total_row['Jumlah Pesanan'].iloc[0]
    total_jumlah_terjual = total_row['Jumlah Terjual'].iloc[0]
    total_row['Jumlah buku per pesanan'] = round(total_jumlah_terjual / total_jumlah_pesanan if total_jumlah_pesanan != 0 else 0, 1)
    for col in ['Harga Satuan', 'Harga Beli', 'No', 'Harga Custom TLJ', 'Variasi']:
        if col in total_row.columns: total_row[col] = None

    # Pastikan semua kolom numerik yang kosong diisi 0 (terutama untuk produk yang cuma ada di iklan)
    cols_to_fix = [
        'Jumlah Terjual', 'Total Penjualan', 'Penjualan Netto', 'Margin', 
        'Persentase', 'Total Pemasukan', 'Harga Satuan', 'Biaya Proses Pesanan', 'Jumlah Pesanan'
    ]
    for col in cols_to_fix:
        if col in summary_df.columns:
            summary_df[col] = summary_df[col].fillna(0)

    # Variasi untuk produk iklan saja bisa diisi strip atau kosong
    summary_df['Variasi'] = summary_df['Variasi'].fillna('')
    
    summary_with_total = pd.concat([summary_final, total_row], ignore_index=True)
    return summary_with_total.fillna(0)

def process_ekspedisi_tiktok(summary_df, pdf_data_list):
    """Membuat sheet EKSPEDISI berdasarkan data summary dan nota PDF."""
    
    # Ambil data relevan dari summary_df (yang berasal dari rekap_processed)
    # Pastikan Nama Produk dan Variasi bersih dari spasi
    kiri_base = summary_df[summary_df['Nama Produk'] != 'Total'].copy()
    kiri_base['Nama Produk'] = kiri_base['Nama Produk'].astype(str).str.strip()
    kiri_base['Variasi'] = kiri_base['Variasi'].astype(str).str.strip()

    # --- PERUBAIKAN: Agregasi QTY berdasarkan Nama Produk DAN Variasi ---
    kiri_df = kiri_base.groupby(['Nama Produk', 'Variasi'], as_index=False).agg(
        QTY=('Jumlah Terjual', 'sum')
    )
    
    # Bagian Kanan: Data dari PDF (tetap sama)
    kanan_df = pd.DataFrame(pdf_data_list)
    
    # Hitung biaya ekspedisi per produk (berdasarkan total QTY semua produk)
    total_qty = kiri_df['QTY'].sum()
    total_nominal = kanan_df['Nominal'].sum() if not kanan_df.empty else 0
    biaya_per_produk = total_nominal / total_qty if total_qty > 0 else 0
    
    # Tambahkan kolom biaya ke kiri_df
    kiri_df['Biaya Ekspedisi per produk'] = biaya_per_produk
    kiri_df['Jumlah'] = kiri_df['QTY'] * biaya_per_produk

    # Buat baris total untuk bagian kiri
    # Pastikan kolom Variasi ada saat membuat total
    kiri_total = pd.DataFrame([{
        'Nama Produk': 'Total', 
        'Variasi': '', # Kosongkan variasi untuk baris total
        'QTY': total_qty, 
        'Biaya Ekspedisi per produk': None, 
        'Jumlah': kiri_df['Jumlah'].sum()
    }])
    kiri_df = pd.concat([kiri_df, kiri_total], ignore_index=True)

    # Susun ulang kolom agar Variasi setelah Nama Produk
    kiri_df = kiri_df[['Nama Produk', 'Variasi', 'QTY', 'Biaya Ekspedisi per produk', 'Jumlah']]

    # Buat baris total untuk bagian kanan (tetap sama)
    kanan_total = pd.DataFrame([{'Tanggal Kirim Paket': 'Total', 'Nominal': total_nominal}])
    if not kanan_df.empty:
        kanan_df = pd.concat([kanan_df, kanan_total], ignore_index=True)
    else:
        kanan_df = kanan_total
    
    # Gabungkan bagian kiri dan kanan
    final_df = pd.concat([kiri_df, pd.DataFrame(columns=[' ']), kanan_df], axis=1)
    return final_df.fillna('')
    
# --- TAMPILAN STREAMLIT ---

st.set_page_config(layout="wide")
st.title("📊 Rekapanku - Sistem Otomatisasi Laporan")

# --- UI PILIHAN JENIS REKAPAN ---
st.header("1. Konfigurasi Rekapan")
jenis_rekapan = st.radio("Pilih Jenis Rekapan:", ["Mingguan", "Bulanan"], horizontal=True)

if jenis_rekapan == "Bulanan":
    st.info("Mode Bulanan: Gabungkan 3-4 file SUMMARY mingguan menjadi satu file.")
    toko_bulanan = st.selectbox("Pilih Toko untuk Rekapan Bulanan:", [
        "Human Store Shopee", "Pacific Bookstore Shopee", "Dama.id Store Shopee",
        "Human Store Tiktok", "Pacific Bookstore Tiktok", "Dama.id Store Tiktok"
    ])
    
    files_mingguan = []
    col1, col2 = st.columns(2)
    with col1:
        f1 = st.file_uploader("Impor Rekapan Minggu 1 (Wajib)", type=["xlsx"])
        f2 = st.file_uploader("Impor Rekapan Minggu 2 (Wajib)", type=["xlsx"])
    with col2:
        f3 = st.file_uploader("Impor Rekapan Minggu 3 (Wajib)", type=["xlsx"])
        f4 = st.file_uploader("Impor Rekapan Minggu 4 (Opsional)", type=["xlsx"])
    
    if st.button("🚀 Proses Rekapan Bulanan"):
        uploaded_files = [f for f in [f1, f2, f3, f4] if f is not None]
        if len(uploaded_files) < 3:
            st.error("Minimal 3 file (Minggu 1, 2, dan 3) harus diunggah!")
        else:
            try:
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    for i, file in enumerate(uploaded_files):
                        # Baca sheet SUMMARY
                        df_summary = pd.read_excel(file, sheet_name='SUMMARY')
                        
                        # Ambil tanggal dari metadata file (jika ada) atau properti Excel
                        try:
                            from openpyxl import load_workbook
                            wb_meta = load_workbook(file)
                            created_dt = wb_meta.properties.created
                            tgl_str = created_dt.strftime("%d/%m/%Y") if created_dt else datetime.now().strftime("%d/%m/%Y")
                        except:
                            tgl_str = datetime.now().strftime("%d/%m/%Y")
                        
                        # Set Header dengan Tanggal
                        sheet_name = f"SUMMARY {i+1}"
                        df_summary.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Tambahkan Tanggal di baris atas atau sel tertentu (Opsional)
                        worksheet = writer.sheets[sheet_name]
                        worksheet.write(0, df_summary.shape[1], f"Tanggal: {tgl_str}")
                
                output.seek(0)
                st.success("✅ Rekapan Bulanan Berhasil!")
                st.download_button(
                    label=f"📥 Download Rekapan Bulanan {toko_bulanan}.xlsx",
                    data=output,
                    file_name=f"REKAPAN_BULANAN_{toko_bulanan.upper().replace(' ', '_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Error Bulanan: {e}")
    st.stop() # Hentikan eksekusi di sini agar tidak masuk ke logika mingguan di bawah
    
marketplace_choice = st.selectbox(
    "Pilih Marketplace:",
    ("", "Shopee", "TikTok")
)

store_choice = ""
if marketplace_choice == "Shopee":
    store_choice = st.selectbox(
        "Pilih Toko Shopee:",
        ("Human Store", "Pacific Bookstore", "DAMA.ID STORE"),
        key='shopee_store'
    )
elif marketplace_choice == "TikTok":
    # Untuk sekarang, TikTok hanya untuk Human Store
    store_choice = st.selectbox(
        "Pilih Toko TikTok:",
        ("Human Store", "DAMA.ID STORE", "Pacific Bookstore"), # Hanya toko yang relevan untuk TikTok
        key='tiktok_store'
    )
    st.info("Marketplace TikTok saat ini hanya tersedia untuk Human Store, DAMA.ID STORE dan Pacific Bookstore.")

# Hanya tampilkan uploader jika marketplace sudah dipilih
if marketplace_choice:
    try:
        # ... (kode untuk membaca HARGA ONLINE.xlsx tetap sama) ...
        katalog_df = pd.read_excel('HARGA ONLINE.xlsx')
    
        # Lakukan preprocessing langsung ke dataframe tunggal
        katalog_df.columns = [str(c).strip().upper() for c in katalog_df.columns]
        for col in ["JUDUL AL QUR'AN", "JENIS KERTAS", "UKURAN", "KATALOG HARGA"]:
            if col not in katalog_df.columns:
                katalog_df[col] = ""
        katalog_df['JUDUL_NORM'] = katalog_df["JUDUL AL QUR'AN"].astype(str).str.upper().str.replace(r'[^A-Z0-9\s]', ' ', regex=True)
        katalog_df['JENIS_KERTAS_NORM'] = katalog_df['JENIS KERTAS'].astype(str).str.upper().str.replace(r'[^A-Z0-9\s]', ' ', regex=True)
        katalog_df['UKURAN_NORM'] = katalog_df['UKURAN'].astype(str).str.upper().str.replace(r'\s+', '', regex=True)
        katalog_df['KATALOG_HARGA_NUM'] = pd.to_numeric(katalog_df['KATALOG HARGA'].astype(str).str.replace(r'[^0-9\.]', '', regex=True), errors='coerce').fillna(0)
    except FileNotFoundError:
        st.error("Error: File 'HARGA ONLINE.xlsx' tidak ditemukan.")
        st.stop()

    try:
        harga_custom_tlj_df = pd.read_excel('Harga Custom TLJ.xlsx')
        
        # Lakukan preprocessing
        harga_custom_tlj_df.columns = [str(c).strip().upper() for c in harga_custom_tlj_df.columns]
        
        # Pastikan kolom yang dibutuhkan ada
        required_cols = ['NAMA PRODUK', 'VARIASI', 'HARGA CUSTOM TLJ']
        if not all(col in harga_custom_tlj_df.columns for col in required_cols):
            st.error(f"File 'Harga Custom TLJ.xlsx' harus memiliki kolom: {', '.join(required_cols)}")
            st.stop()

        # Buat kolom kunci untuk pencocokan yang mudah (Nama Produk + Variasi)
        harga_custom_tlj_df['LOOKUP_KEY'] = harga_custom_tlj_df['NAMA PRODUK'].astype(str).str.strip() + ' ' + harga_custom_tlj_df['VARIASI'].astype(str).str.strip()
        
        # Pastikan kolom harga adalah numerik
        harga_custom_tlj_df['HARGA CUSTOM TLJ'] = pd.to_numeric(harga_custom_tlj_df['HARGA CUSTOM TLJ'], errors='coerce').fillna(0)

    except FileNotFoundError:
        st.error("Error: File 'Harga Custom TLJ.xlsx' tidak ditemukan.")
        st.stop()
    except Exception as e:
        st.error(f"Error saat membaca file 'Harga Custom TLJ.xlsx': {e}")
        st.stop()

    # --- TAMBAHKAN BLOK BARU INI UNTUK MEMBACA KATALOG DAMA ---
    try:
        katalog_dama_df = pd.read_excel('KATALOG_DAMA.xlsx') # Pastikan nama file benar

        # Lakukan preprocessing
        katalog_dama_df.columns = [str(c).strip().upper() for c in katalog_dama_df.columns]

        # Pastikan kolom yang dibutuhkan ada
        required_dama_cols = ['NAMA PRODUK', 'JENIS AL QUR\'AN', 'WARNA', 'UKURAN', 'PAKET', 'HARGA']
        if not all(col in katalog_dama_df.columns for col in required_dama_cols):
            st.error(f"File 'KATALOG_DAMA.xlsx' harus memiliki kolom: {', '.join(required_dama_cols)}")
            st.stop()

        # Konversi kolom harga ke numerik
        katalog_dama_df['HARGA'] = pd.to_numeric(katalog_dama_df['HARGA'], errors='coerce').fillna(0)

        # Bersihkan dan normalisasi kolom teks untuk pencocokan
        for col in ['NAMA PRODUK', 'JENIS AL QUR\'AN', 'WARNA', 'UKURAN', 'PAKET']:
            # Isi NaN dengan string kosong sebelum operasi string
            katalog_dama_df[col] = katalog_dama_df[col].fillna('').astype(str).str.strip().str.upper()
            # Hapus spasi ganda
            katalog_dama_df[col] = katalog_dama_df[col].str.replace(r'\s+', ' ', regex=True)

    except FileNotFoundError:
        st.error("Error: File 'KATALOG_DAMA.xlsx' tidak ditemukan.")
        st.stop()
    except Exception as e:
        st.error(f"Error saat membaca file 'KATALOG_DAMA.xlsx': {e}")
        st.stop()
        
    st.header("1. Import File Anda")

    if marketplace_choice == "Shopee":
        col1, col2 = st.columns(2)
        with col1:
            uploaded_order = st.file_uploader("1. Import file order-all.xlsx", type="xlsx")
            uploaded_income = st.file_uploader("2. Import file income dilepas.xlsx", type="xlsx")
        with col2:
            uploaded_iklan = st.file_uploader("3. Import file iklan produk", type="csv")
            uploaded_seller = st.file_uploader("4. Import file seller conversion", type="csv")
        # Inisialisasi variabel lain agar tidak error
        uploaded_income_tiktok = None
        uploaded_semua_pesanan = None
        uploaded_pdfs = None

    elif marketplace_choice == "TikTok":
        col1, col2 = st.columns(2)
        with col1:
            uploaded_income_tiktok = st.file_uploader("1. Import file Income (Order details & Reports)", type="xlsx")
            uploaded_semua_pesanan = st.file_uploader("2. Import file semua pesanan.xlsx", type="xlsx")
            product_data_file = st.file_uploader("3. Import file Product Data.xlsx", type="xlsx")
        with col2:
            # --- TAMBAHKAN KONDISI DI SINI ---
            # Hanya tampilkan uploader creator order jika BUKAN DAMA.ID STORE
            # if store_choice != "DAMA.ID STORE":
            #     uploaded_creator_order = st.file_uploader("3. Import file creator order-all.xlsx", type="xlsx")
            # else:
            #     # Jika DAMA.ID STORE, pastikan variabelnya ada tapi None
            #     uploaded_creator_order = None
            #     st.info("File 'creator order-all.xlsx' tidak diperlukan untuk DAMA.ID STORE.") # Opsional: beri info
            label_creator = "3. Import file creator order-all.xlsx"
            if store_choice == "DAMA.ID STORE":
                label_creator += " (Opsional)"
                
            uploaded_creator_order = st.file_uploader(label_creator, type="xlsx")
            # ---------------------------------

            uploaded_pdfs = st.file_uploader(
                # Sesuaikan nomor urut jika creator order disembunyikan
                f"{'4.' if store_choice != 'DAMA.ID STORE' else '3.'} Import Nota Resi Ekspedisi (bisa lebih dari satu)",
                type="pdf",
                accept_multiple_files=True
            )
        # Inisialisasi variabel lain agar tidak error
        uploaded_order = None
        uploaded_income = None
        uploaded_iklan = None
        uploaded_seller = None

    st.markdown("---")
    
    # Kondisi untuk menampilkan tombol proses
    # show_shopee_button = marketplace_choice == "Shopee" and uploaded_order and uploaded_income and uploaded_iklan and uploaded_seller
    shopee_base_files = marketplace_choice == "Shopee" and uploaded_order and uploaded_income and uploaded_iklan
    # Tentukan status tombol berdasarkan toko
    if shopee_base_files and store_choice == "DAMA.ID STORE":
        show_shopee_button = True # DAMA.ID STORE siap, seller conversion opsional
    elif shopee_base_files: # Toko Shopee lain (Human/Pacific)
        show_shopee_button = uploaded_seller # Wajib untuk Human/Pacific
    else:
        show_shopee_button = False
        
    # show_tiktok_button = marketplace_choice == "TikTok" and uploaded_income_tiktok and uploaded_semua_pesanan and uploaded_creator_order and uploaded_pdfs
    tiktok_base_files = marketplace_choice == "TikTok" and uploaded_income_tiktok and uploaded_semua_pesanan
    
    show_tiktok_button = False # Inisialisasi
    if tiktok_base_files and store_choice == "DAMA.ID STORE":
        # DAMA.ID STORE: creator_order & pdfs opsional
        show_tiktok_button = True
    elif tiktok_base_files and store_choice in ["Human Store", "Pacific Bookstore"]:
        # Human Store: creator_order & pdfs wajib
        show_tiktok_button = uploaded_creator_order

    if show_shopee_button or show_tiktok_button:
        button_label = f"🚀 Mulai Proses untuk {marketplace_choice} - {store_choice}"
        if st.button(button_label):
            progress_bar = st.progress(0, text="Mempersiapkan proses...")
            status_text = st.empty()
            
            try:
                # --- LOGIKA PEMBACAAN FILE ---
                if marketplace_choice == "Shopee":
                    # --- ALUR PROSES SHOPEE (KODE LAMA ANDA) ---
                    status_text.text("Membaca file Shopee...")
                    order_all_df = pd.read_excel(uploaded_order, dtype={'Harga Setelah Diskon': str, 'Total Harga Produk': str})
                    income_dilepas_df = pd.read_excel(uploaded_income, sheet_name='Income', skiprows=5)
                    # if store_choice == "Human Store":
                    #     service_fee_df = pd.read_excel(uploaded_income, sheet_name='Service Fee Details', skiprows=1)
                    iklan_produk_df = pd.read_csv(uploaded_iklan, skiprows=7)
                    # seller_conversion_df = pd.read_csv(uploaded_seller)
                    if uploaded_seller:
                        seller_conversion_df = pd.read_csv(uploaded_seller)
                    else:
                        # Buat DataFrame kosong jika file tidak ada
                        # Ini penting agar DAMA.ID STORE tidak error
                        seller_conversion_df = pd.DataFrame(columns=['Kode Pesanan', 'Pengeluaran(Rp)'])
                    progress_bar.progress(20, text="File dimuat. Membersihkan format angka...")

                    # ... (Kode pembersihan data keuangan Anda tetap di sini) ...
                    # --- Langkah 1: Bersihkan file order-all secara khusus ---
                    cols_to_clean_order = ['Harga Setelah Diskon', 'Total Harga Produk']
                    for col in cols_to_clean_order:
                      if col in order_all_df.columns:
                          # Gunakan fungsi baru yang spesifik
                          order_all_df[col] = clean_order_all_numeric(order_all_df[col])
    
                    # --- Langkah 2: Bersihkan file-file lainnya dengan fungsi lama ---
                    other_financial_data_to_clean = [
                        (income_dilepas_df, ['Voucher dari Penjual', 'Biaya Administrasi', 'Biaya Proses Pesanan', 'Total Penghasilan']),
                        (iklan_produk_df, ['Biaya', 'Omzet Penjualan']),
                        (seller_conversion_df, ['Pengeluaran(Rp)'])
                    ]
    
                    for df, cols in other_financial_data_to_clean:
                        for col in cols:
                            if col in df.columns:
                              # Gunakan fungsi lama yang umum
                              df[col] = clean_and_convert_to_numeric(df[col])
                
                    # --- LOGIKA PEMROSESAN BERDASARKAN TOKO ---
                    status_text.text("Menyusun sheet 'REKAP' (Shopee)...")
                    if store_choice == "Human Store":
                        rekap_processed = process_rekap(order_all_df, income_dilepas_df, seller_conversion_df)
                    elif store_choice == "Pacific Bookstore": # Hanya Pacific yang pakai logic ini
                        rekap_processed = process_rekap_pacific(order_all_df, income_dilepas_df, seller_conversion_df)
                    elif store_choice == "DAMA.ID STORE": # Panggil fungsi baru untuk DAMA
                        rekap_processed = process_rekap_dama(order_all_df, income_dilepas_df, seller_conversion_df)
                    else: # Pengaman jika ada pilihan store lain
                        st.error(f"Pilihan toko '{store_choice}' tidak dikenali.")
                        st.stop()
                    progress_bar.progress(40, text="Sheet 'REKAP' selesai.")
                    
                    status_text.text("Menyusun sheet 'IKLAN' (Shopee)...")
                    iklan_processed = process_iklan(iklan_produk_df)
                    progress_bar.progress(60, text="Sheet 'IKLAN' selesai.")
    
                    status_text.text("Menyusun sheet 'SUMMARY' (Shopee)...")
                    if store_choice == "DAMA.ID STORE":
                        summary_processed = process_summary_dama(rekap_processed, iklan_processed, katalog_dama_df, harga_custom_tlj_df)
                    else: # Human Store atau Pacific Bookstore
                        summary_processed = process_summary(rekap_processed, iklan_processed, katalog_df, harga_custom_tlj_df, store_type=store_choice)
                    progress_bar.progress(80, text="Sheet 'SUMMARY' selesai.")
                    
                    file_name_output = f"Rekapanku_Shopee_{store_choice}.xlsx"
                    sheets = {
                        'SUMMARY': summary_processed, 'REKAP': rekap_processed, 'IKLAN': iklan_processed,
                        'sheet order-all': order_all_df, 'sheet income dilepas': income_dilepas_df,
                        'sheet biaya iklan': iklan_produk_df, 'sheet seller conversion': seller_conversion_df
                    }
                    # if store_choice == "Human Store": sheets['sheet service fee'] = service_fee_df
    
                elif marketplace_choice == "TikTok":
                    # --- ALUR PROSES TIKTOK BARU ---
                    status_text.text("Membaca file TikTok...")
                    # Baca sheet 'Order details' dan langsung bersihkan kolomnya
                    order_details_df = pd.read_excel(uploaded_income_tiktok, sheet_name='Order details', header=0)
                    order_details_df = clean_columns(order_details_df)
                    order_details_df.columns = [col.upper() for col in order_details_df.columns]
                    # Baca sheet 'Reports' dan langsung bersihkan kolomnya
                    reports_df = pd.read_excel(uploaded_income_tiktok, sheet_name='Reports', header=0)
                    reports_df = clean_columns(reports_df)
                    reports_df.columns = [col.upper() for col in reports_df.columns]
                    if product_data_file:
                        # Load file product data
                        product_data_df = pd.read_excel(product_data_file)
                        # Pastikan nama kolom konsisten
                        product_data_df.columns = [col.upper().strip() for col in product_data_df.columns]
                    else:
                        product_data_df = pd.DataFrame()
                    # Baca 'semua pesanan' dan langsung bersihkan kolomnya
                    # 1. Baca file tanpa header, sehingga semua baris (termasuk header asli) menjadi data
                    wb = load_workbook(uploaded_semua_pesanan, data_only=True)
                    ws = wb.active
                    # Ambil semua baris sebagai list of values
                    data = [list(row) for row in ws.iter_rows(values_only=True)]
                    data = [r for r in data if any(r)]  # hapus baris kosong
                    # Gabungkan 2 baris pertama jadi header
                    # Gunakan hanya baris pertama sebagai header asli (Order ID, Order Status, dst)
                    final_header = [str(x).strip() if x else "" for x in data[0]]
                    
                    # Cek apakah baris kedua berisi "Platform unique order ID" → hapus kalau iya
                    if len(data) > 1 and any("Platform unique order ID" in str(x) for x in data[1]):
                        data_rows = data[2:]  # Lewati baris kedua
                    else:
                        data_rows = data[1:]
                    # Buat DataFrame
                    semua_pesanan_df = pd.DataFrame(data_rows, columns=final_header)
                    # Bersihkan kolom (hapus spasi dan karakter aneh)
                    semua_pesanan_df.columns = semua_pesanan_df.columns.str.strip()
                    semua_pesanan_df = clean_columns(semua_pesanan_df)
                    semua_pesanan_df.columns = [col.upper() for col in semua_pesanan_df.columns]
                    if uploaded_creator_order:
                        # Jika file di-upload (Human Store), baca filenya
                        creator_order_all_df = clean_columns(pd.read_excel(uploaded_creator_order))
                        creator_order_all_df.columns = [col.upper() for col in creator_order_all_df.columns]
                    else:
                        # Jika DAMA.ID STORE (file=None), buat DataFrame kosong
                        # Tambahkan 'SKU' ke daftar kolom agar merge tidak error
                        creator_order_all_df = pd.DataFrame(columns=['ID PESANAN', 'PRODUK', 'Variasi_Clean', 'PEMBAYARAN KOMISI AKTUAL', 'SKU'])
                    progress_bar.progress(20, text="File Excel TikTok dimuat dan kolom dibersihkan.")
                    
                    # status_text.text(f"Memproses {len(uploaded_pdfs)} file PDF nota resi...")
                    # pdf_data = [parse_pdf_receipt(pdf) for pdf in uploaded_pdfs if pdf is not None]
                    # pdf_data = [data for data in pdf_data if data is not None] # Hapus hasil yang gagal
                    pdf_data = [] # Inisialisasi list kosong
                    if uploaded_pdfs: # Hanya proses jika PDF di-upload
                        status_text.text(f"Memproses {len(uploaded_pdfs)} file PDF nota resi...")
                        pdf_data = [parse_pdf_receipt(pdf) for pdf in uploaded_pdfs if pdf is not None]
                        pdf_data = [data for data in pdf_data if data is not None] # Hapus hasil yang gagal
                    else:
                        # Jika tidak ada PDF (kasus DAMA.ID STORE opsional)
                        status_text.text("Melewati pemrosesan PDF nota resi...")
                    progress_bar.progress(40, text="File PDF selesai diproses.")
    
                    status_text.text("Menyusun sheet 'REKAP' (TikTok)...")
                    rekap_processed = process_rekap_tiktok(order_details_df, semua_pesanan_df, creator_order_all_df, store_choice)
                    progress_bar.progress(60, text="Sheet 'REKAP' selesai.")
                    
                    # Untuk SUMMARY, kita perlu EKSPEDISI dulu, tapi EKSPEDISI perlu agregasi dari SUMMARY.
                    # Jadi, kita buat summary sementara dulu.
                    summary_temp_for_ekspedisi = rekap_processed.copy()
                    
                    status_text.text("Menyusun sheet 'EKSPEDISI'...")
                    ekspedisi_processed = process_ekspedisi_tiktok(summary_temp_for_ekspedisi, pdf_data)
                    progress_bar.progress(70, text="Sheet 'EKSPEDISI' selesai.")
    
                    status_text.text("Menyusun sheet 'SUMMARY' (TikTok)...")
                    # summary_processed = process_summary_tiktok(rekap_processed, katalog_df, harga_custom_tlj_df, ekspedisi_processed)
                    summary_processed = process_summary_tiktok(rekap_processed, katalog_df, harga_custom_tlj_df, ekspedisi_processed, product_data_df, store_choice)
                    progress_bar.progress(85, text="Sheet 'SUMMARY' selesai.")
    
                    file_name_output = f"Rekapanku_TikTok_{store_choice}.xlsx"
                    sheets = {
                        'SUMMARY': summary_processed,
                        'REKAP': rekap_processed,
                        'EKSPEDISI': ekspedisi_processed,
                        'sheet Order details': order_details_df,
                        'sheet Reports': reports_df,
                        'sheet semua pesanan': semua_pesanan_df,
                        'sheet creator order-all': creator_order_all_df,
                        'sheet Iklan': product_data_df
                    }

                # ... (Sisa kode untuk membuat file Excel dan tombol download tetap sama) ...
                status_text.text("Menyiapkan file output untuk diunduh...")
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    
                    # --- SEMUA FORMATTING VISUAL DIDEFINISIKAN DI SINI ---
                    workbook = writer.book
                    
                    # --- PERUBAHAN 1: Format Judul diubah menjadi rata kiri (align: 'left') ---
                    title_format = workbook.add_format({'bold': True, 'fg_color': '#4472C4', 'font_color': 'white', 'align': 'left', 'valign': 'vcenter', 'font_size': 14})
                    
                    # Format Header Kolom (biru muda, bold, border)
                    header_format = workbook.add_format({'bold': True, 'fg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
                    
                    # --- PERUBAHAN 2: Tambahkan format border untuk sel data ---
                    cell_border_format = workbook.add_format({'border': 1})
                    
                    # Format Persen (0.00%) DENGAN BORDER
                    percent_format = workbook.add_format({'num_format': '0.00%', 'border': 1})
                    
                    # Format 1 Angka Desimal (0.0) DENGAN BORDER
                    one_decimal_format = workbook.add_format({'num_format': '0.0', 'border': 1})
                    
                    # Format Baris Total (kuning, bold)
                    total_fmt = workbook.add_format({'bold': True, 'fg_color': '#FFFF00', 'border': 1})
                    total_fmt_percent = workbook.add_format({'bold': True, 'fg_color': '#FFFF00', 'num_format': '0.00%', 'border': 1})
                    total_fmt_decimal = workbook.add_format({'bold': True, 'fg_color': '#FFFF00', 'num_format': '0.0', 'border': 1})

                    # --- PROSES SETIAP SHEET ---
                    for sheet_name, df in sheets.items():
                        # --- PERUBAHAN 3: Ubah startrow menjadi 3 untuk memberi ruang 2 baris header ---
                        start_row_data = 3 if sheet_name in ['SUMMARY', 'REKAP', 'IKLAN'] else 1
                        
                        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row_data, header=False)
                        worksheet = writer.sheets[sheet_name]
                        
                        start_row_header = 0
                        if sheet_name in ['SUMMARY', 'REKAP', 'IKLAN']:
                            # --- PERUBAHAN 4: Buat judul dinamis dan merge 2 baris ---
                            judul_sheet = f"{sheet_name} {store_choice.upper()} {marketplace_choice}"
                            worksheet.merge_range(0, 0, 1, len(df.columns) - 1, judul_sheet, title_format) # merge dari baris 0 hingga 1
                            start_row_header = 2 # Header kolom sekarang mulai di baris ke-3 (index 2)
                        
                        for col_num, value in enumerate(df.columns.values):
                            worksheet.write(start_row_header, col_num, value, header_format)

                        # Terapkan formatting KHUSUS untuk sheet SUMMARY, REKAP, dan IKLAN
                        if sheet_name in ['SUMMARY', 'REKAP', 'IKLAN']:
                            # --- PERUBAHAN 5: Terapkan border ke semua sel data ---
                            # (row_start, col_start, row_end, col_end, format)
                            worksheet.conditional_format(start_row_data, 0, start_row_data + len(df) - 1, len(df.columns) - 1, 
                                                         {'type': 'no_blanks', 'format': cell_border_format})

                        if sheet_name == 'SUMMARY':
                            persen_col = df.columns.get_loc('Persentase')
                            penjualan_hari_col = df.columns.get_loc('Penjualan Per Hari')
                            buku_pesanan_col = df.columns.get_loc('Jumlah buku per pesanan')
                            
                            # --- PERUBAHAN 6: Terapkan format persen ke seluruh kolom, bukan hanya baris total ---
                            # Terapkan format mulai dari baris data pertama hingga baris sebelum total
                            # (worksheet.set_column(col_start, col_end, width, format))
                            # worksheet.set_column(persen_col, persen_col, 12, percent_format) # Format ini sudah termasuk border
                            for row_idx in range(len(df) - 1): # -1 agar tidak menyentuh baris 'Total'
                                excel_row = start_row_data + row_idx
                                cell_value = df.iloc[row_idx, persen_col]
                                worksheet.write(excel_row, persen_col, cell_value, percent_format)
                            
                            # Atur lebar kolomnya secara terpisah
                            worksheet.set_column(persen_col, persen_col, 12)
                            worksheet.set_column(penjualan_hari_col, penjualan_hari_col, 18, one_decimal_format)
                            worksheet.set_column(buku_pesanan_col, buku_pesanan_col, 22, one_decimal_format)
                            
                            last_row = len(df) + start_row_header
                            for col_num in range(len(df.columns)):
                                cell_value = df.iloc[-1, col_num]
                                current_fmt = total_fmt
                                if col_num == persen_col:
                                    current_fmt = total_fmt_percent
                                elif col_num in [penjualan_hari_col, buku_pesanan_col]:
                                    current_fmt = total_fmt_decimal
                                
                                if pd.notna(cell_value):
                                    worksheet.write(last_row, col_num, cell_value, current_fmt)
                                else:
                                    worksheet.write_blank(last_row, col_num, None, current_fmt)

                        # TAMBAHKAN BLOK BARU INI
                        if sheet_name == 'IKLAN':
                            # Cek jika baris terakhir adalah baris TOTAL
                            last_row_idx = len(df) - 1
                            if not df.empty and df.iloc[last_row_idx]['Nama Iklan'] == 'TOTAL':
                                # Terapkan format total (kuning, bold, border) ke setiap sel di baris ini
                                for col_num in range(len(df.columns)):
                                    cell_value = df.iloc[last_row_idx, col_num]
                                    worksheet.write(start_row_data + last_row_idx, col_num, cell_value, total_fmt)
                        
                        # Atur lebar kolom otomatis untuk semua sheet
                        for i, col in enumerate(df.columns):
                            column_len = max(df[col].astype(str).map(len).max(), len(col))
                            worksheet.set_column(i, i, column_len + 2)
                
                output.seek(0)
                progress_bar.progress(100, text="Proses Selesai!")
                status_text.success("✅ Proses Selesai! File Anda siap diunduh.")

                st.header("3. Download Hasil")
                st.download_button(
                    label=f"📥 Download File Output ({file_name_output})",
                    data=output,
                    file_name=file_name_output,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Terjadi kesalahan saat pemrosesan: {e}")
                st.exception(e)
else:
    st.info("Silakan pilih toko terlebih dahulu untuk melanjutkan.")
